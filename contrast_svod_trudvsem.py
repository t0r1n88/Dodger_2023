"""
Скрипт для подсчета изменения в сводных таблицах данных полученных с помощью Кассандры
"""
import pandas

from cass_support_functions import write_df_to_excel_color_selection # получаем функцию для записи в листы Excel
import pandas as pd
from tkinter import messagebox
import math
import numpy as np
import time
import openpyxl
from openpyxl.styles import Font, PatternFill


class FirstNotSheets(Exception):
    """
    Класс для обработки случаев когда не хватает обязательных листов в первом файле
    """
    pass

class SecondNotSheets(Exception):
    """
    Класс для обработки случаев когда не хватает обязательных листов в втором файле
    """
    pass

class NotColumnVac(Exception):
    """
    Класс для проверки наличия колонки Количество вакансий
    """

class NotColumnPay(Exception):
    """
    Класс для проверки наличия колонок средняя и медианная зп
    """
    pass


def split_cell(cell: str, user_sep: str, number_split: int):
    """
    Функция для извлечения части строки
    """
    lst_cell = str(cell).split(user_sep)  # получаем список
    if len(lst_cell) >= number_split + 1:
        out_value = lst_cell[number_split].strip()
        return out_value
    else:
        return cell


def sum_category(row: pd.Series, user_sep: str):
    """
    Суммирование колонок для ID
    """
    out_id = f'{user_sep}'.join(row.tolist())  # создаем выходную строку

    return out_id

def convert_int(value):
    """
    Функция для конвертации в инт
    """
    try:
        return int(value)
    except:
        return 0


def union_column(row):
    """
    Функция для преобразования двух колонок получившихся после мерджа в одну
    """
    # убираем нан
    lst_value = [value for value in (row[0],row[1]) if not pandas.isna(value)]
    # если значение есть в первой или во второй таблице
    if len(lst_value) == 1:
        return lst_value[0]
    elif len(lst_value) == 2:
        # возвращаем значение из второй таблицы
        return row[1]
    else:
        return 'Проверьте количество передаваемых колонок в коде'


def prepare_diff_svod_trudvsem(first_file:str, second_file:str, end_folder:str,type_contrast:str):
    """
    Функция для подсчета разницы между двумя сводами
    :param first_file: первый свод
    :param second_file: второй свод
    :param end_folder: куда будет сохранен результ
    :param type_contrast: тип обработки Сравнение только отраслей или отраслей и работодателей
    """
    try:
        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)
        current_date = time.strftime('%d_%m_%Y', t)
        # Список обязательных листов которые должны быть в файле
        lst_svod_sheets = ['Вакансии по отраслям','Вакансии по муниципалитетам','Муниципалитеты отрасли','Отрасли муниципалитеты', 'Вакансии по работодателям','Вакансии для динамики', 'Зарплата по отраслям',
                           'Категории ЗП по отраслям',
                           'Зарплата по работодателям','Категории ЗП по работодателям',
                           'Образование по отраслям', 'Образование по работодателям', 'График работы по отраслям',
                           'График работы по работодателям',
                           'Тип занятости по отраслям', 'Тип занятости по работодателям', 'Квоты по отраслям',
                           'Квоты по работодателям','Вакансии для соц.кат.',
                           'Требуемый опыт по отраслям', 'Требуемый опыт по работодателям']

        lst_not_standard_sheets = ['Вакансии для динамики','Зарплата по отраслям',
                                   'Зарплата по работодателям']  # список листов требующих нестандартной обработки

        dct_df = dict() # создаем словарь для хранения датафреймов

        #  получаем листы для проверки
        first_wb = openpyxl.load_workbook(first_file)
        first_wb_sheets = first_wb.sheetnames
        first_wb.close()
        second_wb = openpyxl.load_workbook(second_file)
        second_wb_sheets = second_wb.sheetnames
        second_wb.close()

        # Проводим проверки на наличие нужных листов
        diff_first_file_sheets = set(lst_svod_sheets).difference(set(first_wb_sheets))
        if len(diff_first_file_sheets) != 0:
            raise FirstNotSheets

        diff_second_file_sheets = set(lst_svod_sheets).difference(set(second_wb_sheets))
        if len(diff_second_file_sheets) != 0:
            raise SecondNotSheets
        if type_contrast == 'No':
            # Обрабатываем листы
            for name_sheet in lst_svod_sheets:
                # Если свод стандартный по количеству вакансий
                if name_sheet not in lst_not_standard_sheets:
                    first_df = pd.read_excel(first_file, sheet_name=name_sheet)  # первый файл для сравнения
                    second_df = pd.read_excel(second_file, sheet_name=name_sheet)  # второй файл для сравнения

                    # получаем колонки которые будем использовать для создания ID
                    first_key_columns = list(first_df.columns)
                    if 'Количество вакансий' not in first_key_columns: # проверяем наличие колонки с вакансиями
                        raise NotColumnVac
                    first_key_columns.remove('Количество вакансий')

                    second_key_columns = list(second_df.columns)
                    if 'Количество вакансий' not in second_key_columns: # проверяем наличие колонки с вакансиями
                        raise NotColumnVac
                    second_key_columns.remove('Количество вакансий')

                    # Делаем колонки на основе котороых будут создаваться ID строковыми
                    first_df[first_key_columns] = first_df[first_key_columns].astype(str)
                    second_df[second_key_columns] = second_df[second_key_columns].astype(str)

                    # Создаем колонки по которым будет вестись объединение
                    first_df['ID'] = first_df[first_key_columns].apply(lambda x: sum_category(x, '+&+'), axis=1)
                    second_df['ID'] = second_df[second_key_columns].apply(lambda x: sum_category(x, '+&+'), axis=1)

                    # Проводим внешнее слияние
                    merge_df = first_df.merge(second_df, how='outer', left_on=['ID'], right_on=['ID'], indicator=True)

                    # удаляем лишние колонки
                    del_columns = [column for column in merge_df.columns if ('вакансий' not in column) & ('ID' not in column)]
                    merge_df.drop(columns=del_columns, inplace=True)

                    # Упорядочиваем колонки
                    merge_df = merge_df.reindex(
                        columns=['ID', 'Количество вакансий_x', 'Количество вакансий_y'])  # меняем местами
                    merge_df.columns = ['Показатель', 'Первая таблица', 'Вторая таблица']

                    merge_df.fillna(0, inplace=True) # заполняем наны

                    merge_df[['Первая таблица', 'Вторая таблица']] = merge_df[['Первая таблица', 'Вторая таблица']].applymap(convert_int) # приводик колонки с числами к флоат

                    # Создаем колонки с подчетом разниц
                    merge_df['Разница'] = merge_df['Вторая таблица'] - merge_df['Первая таблица']
                    merge_df['Абсолютная разница'] = abs(merge_df['Вторая таблица'] - merge_df['Первая таблица'])
                    merge_df['Изменение в %'] = round(
                        ((merge_df['Вторая таблица'] - merge_df['Первая таблица']) / merge_df['Первая таблица']) * 100, 2)

                    merge_df['Отношение второй таблицы к первой %'] = round(
                        (merge_df['Вторая таблица'] / merge_df['Первая таблица']) * 100, 2)

                    merge_df.sort_values(by='Показатель', inplace=True)  # Сортируем по показателю
                    dct_df[name_sheet] = merge_df  # сохраняем в словарь

                else:
                    # вычисляем динамику изменения вакансий
                    if name_sheet == 'Вакансии для динамики':
                        first_df = pd.read_excel(first_file, sheet_name=name_sheet)  # первый файл для сравнения
                        second_df = pd.read_excel(second_file, sheet_name=name_sheet)  # второй файл для сравнения
                        # Проводим внешнее слияние
                        merge_df = first_df.merge(second_df, how='outer', left_on=['ID вакансии'], right_on=['ID вакансии'], indicator=True)

                        # Создаем объединенные колонки
                        merge_df['Работодатель'] = merge_df[['Краткое название работодателя_x','Краткое название работодателя_y']].apply(union_column,axis=1)
                        merge_df['Вакансия'] = merge_df[['Вакансия_x','Вакансия_y']].apply(union_column,axis=1)
                        merge_df['Ссылка на вакансию'] = merge_df[['Ссылка на вакансию_x','Ссылка на вакансию_y']].apply(union_column,axis=1)
                        merge_df.drop(columns=['ID вакансии','_merge','Краткое название работодателя_x','Краткое название работодателя_y',
                                               'Вакансия_x','Вакансия_y','Ссылка на вакансию_x','Ссылка на вакансию_y'],inplace=True) # удаляем лишние колонки

                        # Переименовываем колонки
                        merge_df.rename(columns={'Количество рабочих мест_x':'Первая таблица','Количество рабочих мест_y':'Вторая таблица'},inplace=True)
                        merge_df.fillna(0, inplace=True)  # заполняем наны

                        merge_df[['Первая таблица', 'Вторая таблица']] = merge_df[
                            ['Первая таблица', 'Вторая таблица']].applymap(
                            convert_int)  # приводим колонки с числами к флоат

                        # Создаем колонки с подчетом разниц
                        merge_df['Разница'] = merge_df['Вторая таблица'] - merge_df['Первая таблица']
                        merge_df['Абсолютная разница'] = abs(merge_df['Вторая таблица'] - merge_df['Первая таблица'])
                        merge_df['Изменение в %'] = round(
                            ((merge_df['Вторая таблица'] - merge_df['Первая таблица']) / merge_df[
                                'Первая таблица']) * 100, 2)

                        merge_df['Отношение второй таблицы к первой %'] = round(
                            (merge_df['Вторая таблица'] / merge_df['Первая таблица']) * 100, 2)

                        merge_df.sort_values(by='Работодатель', inplace=True)  # Сортируем по показателю
                        merge_df = merge_df.reindex(columns=['Работодатель','Вакансия','Первая таблица','Вторая таблица','Разница',
                                                  'Абсолютная разница','Изменение в %','Отношение второй таблицы к первой %','Ссылка на вакансию'])


                        dct_df[name_sheet] = merge_df  # сохраняем в словарь

                    else:
                        # обрабатываем листы с зарплатой
                        first_df = pd.read_excel(first_file, sheet_name=name_sheet)  # первый файл для сравнения
                        second_df = pd.read_excel(second_file, sheet_name=name_sheet)  # второй файл для сравнения

                        # получаем колонки которые будем использовать для создания ID
                        first_key_columns = list(first_df.columns)
                        if 'Средняя ариф. минимальная зп' not in first_key_columns or 'Медианная минимальная зп' not in first_key_columns: # проверяем наличие колонки с вакансиями
                            raise NotColumnPay
                        first_key_columns.remove('Средняя ариф. минимальная зп')
                        first_key_columns.remove('Медианная минимальная зп')

                        second_key_columns = list(second_df.columns)
                        if 'Средняя ариф. минимальная зп' not in second_key_columns or 'Медианная минимальная зп' not in second_key_columns: # проверяем наличие колонки с вакансиями
                            raise NotColumnPay
                        second_key_columns.remove('Средняя ариф. минимальная зп')
                        second_key_columns.remove('Медианная минимальная зп')

                        # Создаем колонки по которым будет вестись объединение
                        first_df['ID'] = first_df[first_key_columns].apply(lambda x: sum_category(x, '-'), axis=1)
                        second_df['ID'] = second_df[second_key_columns].apply(lambda x: sum_category(x, '-'), axis=1)

                        # Проводим внешнее слияние
                        merge_df = first_df.merge(second_df, how='outer', left_on=['ID'], right_on=['ID'], indicator=True)

                        # удаляем лишние колонки
                        del_columns = [column for column in merge_df.columns if ('минимальная зп' not in column) & ('ID' not in column)]
                        merge_df.drop(columns=del_columns, inplace=True)

                        # Упорядочиваем колонки
                        merge_df = merge_df.reindex(
                            columns=['ID', 'Средняя ариф. минимальная зп_x','Средняя ариф. минимальная зп_y', 'Медианная минимальная зп_x',
                                     'Медианная минимальная зп_y'])  # меняем местами
                        merge_df.columns = ['Показатель', 'Средняя ариф. минимальная. Первая таблица', 'Средняя ариф. минимальная. Вторая таблица',
                                            'Медианная минимальная. Первая таблица','Медианная минимальная. Вторая таблица']

                        merge_df.fillna(0, inplace=True) # заполняем наны

                        merge_df[['Средняя ариф. минимальная. Первая таблица', 'Средняя ариф. минимальная. Вторая таблица',
                                            'Медианная минимальная. Первая таблица','Медианная минимальная. Вторая таблица']] = merge_df[['Средняя ариф. минимальная. Первая таблица', 'Средняя ариф. минимальная. Вторая таблица',
                                            'Медианная минимальная. Первая таблица','Медианная минимальная. Вторая таблица']].applymap(convert_int) # приводик колонки с числами к int

                        # Создаем колонки с подчетом разниц ср.ариф.
                        merge_df['Сред.ариф.мин. Разница'] = merge_df['Средняя ариф. минимальная. Вторая таблица'] - merge_df['Средняя ариф. минимальная. Первая таблица']
                        merge_df['Сред.ариф.мин. Абсолютная разница'] = abs(merge_df['Средняя ариф. минимальная. Вторая таблица'] - merge_df['Средняя ариф. минимальная. Первая таблица'])
                        merge_df['Сред.ариф.мин. Изменение в %'] = round(
                            ((merge_df['Средняя ариф. минимальная. Вторая таблица'] - merge_df['Средняя ариф. минимальная. Первая таблица']) / merge_df['Средняя ариф. минимальная. Первая таблица']) * 100, 2)

                        merge_df['Сред.ариф.мин. Отношение второй таблицы к первой %'] = round(
                            (merge_df['Средняя ариф. минимальная. Вторая таблица'] / merge_df['Средняя ариф. минимальная. Первая таблица']) * 100, 2)

                        # Создаем колонки с подсчетом разниц медиан
                        merge_df['Медианнная.мин. Разница'] = merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                            'Медианная минимальная. Первая таблица']

                        merge_df['Медианнная.мин. Абсолютная разница'] = abs(
                            merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                                'Медианная минимальная. Первая таблица'])

                        merge_df['Медианнная.мин. Изменение в %'] = round(
                            ((merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                                'Медианная минимальная. Первая таблица']) / merge_df[
                                 'Медианная минимальная. Первая таблица']) * 100, 2)

                        merge_df['Медианнная.мин. Отношение второй таблицы к первой %'] = round(
                            (merge_df['Медианная минимальная. Вторая таблица'] / merge_df[
                                'Медианная минимальная. Первая таблица']) * 100, 2)
                        #


                        merge_df.sort_values(by='Показатель', inplace=True)  # Сортируем по показателю
                        dct_df[name_sheet] = merge_df  # сохраняем в словарь




            # Создаем словарь с параметрами записи
            dct_change = {'number_column':3,'font':Font(color='FF000000'),
                          'fill':PatternFill(fill_type='solid', fgColor='ffa500'),
                          'find_value':'-'}
            dct_grow = {'number_column':3,'font':Font(color='FF000000'),
                          'fill':PatternFill(fill_type='solid', fgColor='90ee90'),
                          'find_value':'+'}

            change_wb = write_df_to_excel_color_selection(dct_df,False,[dct_change,dct_grow],lst_not_standard_sheets)
            change_wb.save(f'{end_folder}/Изменения от {current_time}.xlsx')
        else:
            for name_sheet in lst_svod_sheets:
                # Если свод стандартный по количеству вакансий
                if name_sheet not in lst_not_standard_sheets and 'работодателям' not in name_sheet and name_sheet !='Вакансии для динамики' and 'муниципалитетам' not in name_sheet:
                    first_df = pd.read_excel(first_file, sheet_name=name_sheet)  # первый файл для сравнения
                    second_df = pd.read_excel(second_file, sheet_name=name_sheet)  # второй файл для сравнения

                    # получаем колонки которые будем использовать для создания ID
                    first_key_columns = list(first_df.columns)
                    if 'Количество вакансий' not in first_key_columns:  # проверяем наличие колонки с вакансиями
                        raise NotColumnVac
                    first_key_columns.remove('Количество вакансий')

                    second_key_columns = list(second_df.columns)
                    if 'Количество вакансий' not in second_key_columns:  # проверяем наличие колонки с вакансиями
                        raise NotColumnVac
                    second_key_columns.remove('Количество вакансий')

                    # Делаем колонки на основе котороых будут создаваться ID строковыми
                    first_df[first_key_columns] = first_df[first_key_columns].astype(str)
                    second_df[second_key_columns] = second_df[second_key_columns].astype(str)

                    # Создаем колонки по которым будет вестись объединение
                    first_df['ID'] = first_df[first_key_columns].apply(lambda x: sum_category(x, '+&+'), axis=1)
                    second_df['ID'] = second_df[second_key_columns].apply(lambda x: sum_category(x, '+&+'), axis=1)

                    # Проводим внешнее слияние
                    merge_df = first_df.merge(second_df, how='outer', left_on=['ID'], right_on=['ID'], indicator=True)

                    # удаляем лишние колонки
                    del_columns = [column for column in merge_df.columns if
                                   ('вакансий' not in column) & ('ID' not in column)]
                    merge_df.drop(columns=del_columns, inplace=True)

                    # Упорядочиваем колонки
                    merge_df = merge_df.reindex(
                        columns=['ID', 'Количество вакансий_x', 'Количество вакансий_y'])  # меняем местами
                    merge_df.columns = ['Показатель', 'Первая таблица', 'Вторая таблица']

                    merge_df.fillna(0, inplace=True)  # заполняем наны

                    merge_df[['Первая таблица', 'Вторая таблица']] = merge_df[
                        ['Первая таблица', 'Вторая таблица']].applymap(convert_int)  # приводик колонки с числами к инту

                    # Создаем колонки с подчетом разниц
                    merge_df['Разница'] = merge_df['Вторая таблица'] - merge_df['Первая таблица']
                    merge_df['Абсолютная разница'] = abs(merge_df['Вторая таблица'] - merge_df['Первая таблица'])
                    merge_df['Изменение в %'] = round(
                        ((merge_df['Вторая таблица'] - merge_df['Первая таблица']) / merge_df['Первая таблица']) * 100,
                        2)

                    merge_df['Отношение второй таблицы к первой %'] = round(
                        (merge_df['Вторая таблица'] / merge_df['Первая таблица']) * 100, 2)

                    merge_df.sort_values(by='Показатель', inplace=True)  # Сортируем по показателю
                    dct_df[name_sheet] = merge_df  # сохраняем в словарь

                else:
                    if 'работодателям' not in name_sheet and name_sheet!='Вакансии для динамики' and 'муниципалитетам' not in name_sheet:

                        # обрабатываем нестандартные листы
                        first_df = pd.read_excel(first_file, sheet_name=name_sheet)  # первый файл для сравнения
                        second_df = pd.read_excel(second_file, sheet_name=name_sheet)  # второй файл для сравнения

                        # получаем колонки которые будем использовать для создания ID
                        first_key_columns = list(first_df.columns)
                        if 'Средняя ариф. минимальная зп' not in first_key_columns or 'Медианная минимальная зп' not in first_key_columns:  # проверяем наличие колонки с вакансиями
                            raise NotColumnPay
                        first_key_columns.remove('Средняя ариф. минимальная зп')
                        first_key_columns.remove('Медианная минимальная зп')

                        second_key_columns = list(second_df.columns)
                        if 'Средняя ариф. минимальная зп' not in second_key_columns or 'Медианная минимальная зп' not in second_key_columns:  # проверяем наличие колонки с вакансиями
                            raise NotColumnPay
                        second_key_columns.remove('Средняя ариф. минимальная зп')
                        second_key_columns.remove('Медианная минимальная зп')

                        # Создаем колонки по которым будет вестись объединение
                        first_df['ID'] = first_df[first_key_columns].apply(lambda x: sum_category(x, '-'), axis=1)
                        second_df['ID'] = second_df[second_key_columns].apply(lambda x: sum_category(x, '-'), axis=1)

                        # Проводим внешнее слияние
                        merge_df = first_df.merge(second_df, how='outer', left_on=['ID'], right_on=['ID'], indicator=True)

                        # удаляем лишние колонки
                        del_columns = [column for column in merge_df.columns if
                                       ('минимальная зп' not in column) & ('ID' not in column)]
                        merge_df.drop(columns=del_columns, inplace=True)

                        # Упорядочиваем колонки
                        merge_df = merge_df.reindex(
                            columns=['ID', 'Средняя ариф. минимальная зп_x', 'Средняя ариф. минимальная зп_y',
                                     'Медианная минимальная зп_x',
                                     'Медианная минимальная зп_y'])  # меняем местами
                        merge_df.columns = ['Показатель', 'Средняя ариф. минимальная. Первая таблица',
                                            'Средняя ариф. минимальная. Вторая таблица',
                                            'Медианная минимальная. Первая таблица',
                                            'Медианная минимальная. Вторая таблица']

                        merge_df.fillna(0, inplace=True)  # заполняем наны

                        merge_df[['Средняя ариф. минимальная. Первая таблица', 'Средняя ариф. минимальная. Вторая таблица',
                                  'Медианная минимальная. Первая таблица', 'Медианная минимальная. Вторая таблица']] = \
                        merge_df[['Средняя ариф. минимальная. Первая таблица', 'Средняя ариф. минимальная. Вторая таблица',
                                  'Медианная минимальная. Первая таблица', 'Медианная минимальная. Вторая таблица']].applymap(convert_int)  # приводик колонки с числами к инту

                        # Создаем колонки с подчетом разниц ср.ариф.
                        merge_df['Сред.ариф.мин. Разница'] = merge_df['Средняя ариф. минимальная. Вторая таблица'] - \
                                                             merge_df['Средняя ариф. минимальная. Первая таблица']
                        merge_df['Сред.ариф.мин. Абсолютная разница'] = abs(
                            merge_df['Средняя ариф. минимальная. Вторая таблица'] - merge_df[
                                'Средняя ариф. минимальная. Первая таблица'])
                        merge_df['Сред.ариф.мин. Изменение в %'] = round(
                            ((merge_df['Средняя ариф. минимальная. Вторая таблица'] - merge_df[
                                'Средняя ариф. минимальная. Первая таблица']) / merge_df[
                                 'Средняя ариф. минимальная. Первая таблица']) * 100, 2)

                        merge_df['Сред.ариф.мин. Отношение второй таблицы к первой %'] = round(
                            (merge_df['Средняя ариф. минимальная. Вторая таблица'] / merge_df[
                                'Средняя ариф. минимальная. Первая таблица']) * 100, 2)

                        # Создаем колонки с подсчетом разниц медиан
                        merge_df['Медианнная.мин. Разница'] = merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                            'Медианная минимальная. Первая таблица']

                        merge_df['Медианнная.мин. Абсолютная разница'] = abs(
                            merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                                'Медианная минимальная. Первая таблица'])

                        merge_df['Медианнная.мин. Изменение в %'] = round(
                            ((merge_df['Медианная минимальная. Вторая таблица'] - merge_df[
                                'Медианная минимальная. Первая таблица']) / merge_df[
                                 'Медианная минимальная. Первая таблица']) * 100, 2)

                        merge_df['Медианнная.мин. Отношение второй таблицы к первой %'] = round(
                            (merge_df['Медианная минимальная. Вторая таблица'] / merge_df[
                                'Медианная минимальная. Первая таблица']) * 100, 2)
                        #

                        merge_df.sort_values(by='Показатель', inplace=True)  # Сортируем по показателю
                        dct_df[name_sheet] = merge_df  # сохраняем в словарь

            # Создаем словарь с параметрами записи
            dct_change = {'number_column': 3, 'font': Font(color='FF000000'),
                          'fill': PatternFill(fill_type='solid', fgColor='ffa500'),
                          'find_value': '-'}
            dct_grow = {'number_column': 3, 'font': Font(color='FF000000'),
                        'fill': PatternFill(fill_type='solid', fgColor='90ee90'),
                        'find_value': '+'}

            change_wb = write_df_to_excel_color_selection(dct_df, False, [dct_change, dct_grow],
                                                          lst_not_standard_sheets)
            change_wb.save(f'{end_folder}/Изменения от {current_time}.xlsx')

    except FirstNotSheets:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f'В первом файле не хватает листов {diff_first_file_sheets}')

    except SecondNotSheets:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f'Во втором файле не хватает листов {diff_second_file_sheets}')

    except NotColumnVac:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f' На листе {name_sheet} не найдена колонка Количество вакансий')

    except NotColumnPay:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f' На листе {name_sheet} не найдены колонки  Средняя ариф. минимальная зп или Медианная минимальная зп')

    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    except OSError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Укажите в качестве конечной папки, папку в корне диска с коротким названием. Проблема может быть\n '
                             f'в слишком длинном пути к создаваемому файлу')
    else:
        messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                            'Данные успешно обработаны.Ошибок не обнаружено')


if __name__ == '__main__':
    main_first_file = 'data/Аналитика по вакансиям региона/15_03_2024/Свод по региону Бурятия от 14 марта.xlsx'
    main_first_file = 'data/Республика Бурятия/Аналитика по вакансиям региона/28_05_2024/Свод по региону от 11_42_58.xlsx'
    main_second_file = 'data/Аналитика по вакансиям региона/16_03_2024/Свод по региону Бурятия от 16 марта.xlsx'
    main_second_file = 'data/Республика Бурятия/Аналитика по вакансиям региона/28_05_2024/2 файл.xlsx'
    main_end_folder = 'data'

    prepare_diff_svod_trudvsem(main_first_file, main_second_file, main_end_folder,'No')

    print('Lindy Booth !!!')
































"""
Вспомогательные функции и исключения
"""
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# Классы для исключений
class BadHeader(Exception):
    """
    Класс для проверки правильности заголовка
    """
    pass


class CheckBoxException(Exception):
    """
    Класс для вызовы исключения в случае если неправильно выставлены чекбоксы
    """
    pass


class NotFoundValue(Exception):
    """
    Класс для обозначения того что значение не найдено
    """
    pass


class ShapeDiffierence(Exception):
    """
    Класс для обозначения несовпадения размеров таблицы
    """
    pass


class ColumnsDifference(Exception):
    """
    Класс для обозначения того что названия колонок не совпадают
    """
    pass


def convert_to_int(value):
    """
    Функция для конвертации в инт
    """
    try:
        return int(value)
    except ValueError:
        return 0




def write_df_to_excel_color_selection(dct_df:dict,write_index:bool,lst_color_select:list,exlude_sheets:list)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel с выделением цветом нужных строк
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :param lst_color_select: параметры для выделение цветом строк по значению. Список словарей
    :param exlude_sheets: список листов для которых нужна особая обработка без расширения колонок
    :return: объект Workbook с записанными датафреймами
    """

    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    # Создаем словарь для отрицательных значений среднего арифметического
    dct_negative = {'number_column': 5, 'font': Font(color='FF000000'),
                'fill': PatternFill(fill_type='solid', fgColor='ffa500'),
                'find_value': '-'}
    # Создаем словарь для положительных значений среднего арифметического
    dct_positive = {'number_column': 5, 'font': Font(color='FF000000'),
                'fill': PatternFill(fill_type='solid', fgColor='90ee90'),
                'find_value': '-'}



    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        none_check = None # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df,index=write_index,header=True):
            if len(row) == 1 and not row[0]: # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)
        count_index += 1
        # ширина по содержимому
        # сохраняем по ширине колонок
        if name_sheet not in exlude_sheets:
            for column in wb[name_sheet].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb[name_sheet].column_dimensions[column_name].width = adjusted_width

        else:
            # делаем перенос строк чтобы было удобнее
            wb[name_sheet].column_dimensions['A'].width = 50 #
            wb[name_sheet]['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['B1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['C1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['D1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['E1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['F1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['G1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['H1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['I1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['J1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['K1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['L1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['M1'].alignment = openpyxl.styles.Alignment(wrapText=True)
            wb[name_sheet]['N1'].alignment = openpyxl.styles.Alignment(wrapText=True)


        # Форматирование строк
        # Итерируемся по словарям с параметрами
        if name_sheet not in exlude_sheets:
            for param_dct in lst_color_select:
                if param_dct['find_value'] == '-': # если нужно выделить отрицательные значения
                    font = param_dct['font']  # Получаем цвет шрифта
                    fill = param_dct['fill'] # получаем заливку
                    if name_sheet != 'Вакансии для динамики':
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            if param_dct['find_value'] in str(row[param_dct['number_column']].value): # делаем ячейку строковой и проверяем наличие искомого слова
                                for cell in row: # применяем стиль если условие сработало
                                    cell.font = font
                                    cell.fill = fill
                    else:
                        # Поскольку на этом листе колонка разницы не третья а четвертая то добавляем единицу
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            if param_dct['find_value'] in str(row[param_dct['number_column'] + 1].value): # делаем ячейку строковой и проверяем наличие искомого слова
                                for cell in row: # применяем стиль если условие сработало
                                    cell.font = font
                                    cell.fill = fill
                elif param_dct['find_value'] == '+': # если нужно выделить значения больше нуля
                    font = param_dct['font']  # Получаем цвет шрифта
                    fill = param_dct['fill'] # получаем заливку
                    if name_sheet != 'Вакансии для динамики':
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            try:
                                if int(row[param_dct['number_column']].value) > 0:
                                    for cell in row: # применяем стиль если условие сработало
                                        cell.font = font
                                        cell.fill = fill
                            except:
                                continue
                    else:
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            try:
                                if int(row[param_dct['number_column'] + 1].value) > 0:
                                    for cell in row: # применяем стиль если условие сработало
                                        cell.font = font
                                        cell.fill = fill
                            except:
                                continue
        else:
            if name_sheet == 'Вакансии для динамики':
                for param_dct in lst_color_select:
                    if param_dct['find_value'] == '-': # если нужно выделить отрицательные значения
                        font = param_dct['font']  # Получаем цвет шрифта
                        fill = param_dct['fill'] # получаем заливку
                            # Поскольку на этом листе колонка разницы не третья а четвертая то добавляем единицу
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            if param_dct['find_value'] in str(row[param_dct['number_column'] + 1].value): # делаем ячейку строковой и проверяем наличие искомого слова
                                for cell in row: # применяем стиль если условие сработало
                                    cell.font = font
                                    cell.fill = fill
                    elif param_dct['find_value'] == '+': # если нужно выделить значения больше нуля
                        font = param_dct['font']  # Получаем цвет шрифта
                        fill = param_dct['fill'] # получаем заливку
                        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                                        min_col=0, max_col=df.shape[1]):  # Перебираем строки
                            try:
                                if int(row[param_dct['number_column'] + 1].value) > 0:
                                    for cell in row: # применяем стиль если условие сработало
                                        cell.font = font
                                        cell.fill = fill
                            except:
                                continue
            else:
                # Обрабатываем листы с зарплатой
                for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                                    min_col=0, max_col=df.shape[1]):  # Перебираем строки
                    try:
                        value = int(row[dct_negative['number_column']].value)
                        if value > 0:
                            for cell in row:  # применяем стиль если условие сработало
                                cell.font = dct_positive['font']
                                cell.fill = dct_positive['fill']
                        elif value < 0:
                            for cell in row:  # применяем стиль если условие сработало
                                cell.font = dct_negative['font']
                                cell.fill = dct_negative['fill']
                        else:
                            continue
                    except:
                        continue

    return wb


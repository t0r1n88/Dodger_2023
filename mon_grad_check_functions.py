"""
Модуль для проверочных функций мониторинга занятости выпускников для сайта СССР
"""
from cass_check_functions import extract_code_nose
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def create_check_tables_mon_grad(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности чтобы можно было проверить правильность внесенных
    данных
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}
    # инвертируем словарь так, чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})
    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    used_name_sheet = set()  # Множество для хранения названий листов
    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            code = extract_code_nose(code_spec)
            # проверяем есть ли такой лист. На случай когда коды одинаковые а описание специальности разное
            if code not in used_name_sheet:
                wb.create_sheet(title=code, index=idx)
                used_name_sheet.add(code)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            code = extract_code_nose(code_spec)
            # Конвертируем в датафрейм
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            # Удаляем колонки без названий
            temp_code_df.drop(columns=[name_column for name_column in temp_code_df.columns if 'Unnamed' in name_column],
                              inplace=True)

            temp_code_df = temp_code_df.reset_index()  # извлекаем название организации
            temp_code_df.rename(columns={'index': 'Наименование'},
                                inplace=True)  # переименовываем колонку с названием организации

            for r in dataframe_to_rows(temp_code_df, index=False, header=True):
                wb[code].append(r)
            wb[code].column_dimensions['A'].width = 20
            wb[code].column_dimensions['B'].width = 40

    # Удаляем лист Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb


def check_first_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 2 = 3 + 31 + 32 + 60 + 61 + 62+ 63 + 64 + 65 + 66 + 67 + 68 + 69 + 70.
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['3', '31', '32', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['2'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]  # получаем значение проверки
    if df.iloc[0, -1] == 'Неправильно':
        # получаем значения из колонок сумма и проверочной колонки
        first_value = df['2'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]

        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 2 = 3 + 31 + 32 + 60 + 61 + 62+ 63 + 64 + 65 + 66 + 67 + 68 + 69 + 70.'
                                            f' Графа 2 = {first_value} ,сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_second_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 3 = сумма значений граф с 4 по 30
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['4', '5', '6', '7', '8', '9',
                         '10', '11', '12', '13', '14', '15',
                         '16', '17', '18', '19', '20', '21',
                         '22', '23', '24', '25', '26', '27', '28', '29', '30']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['3'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['3'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 3 = сумма значений граф с 4 по 30. Графа 3 = {first_value}, сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_third_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 3 >= графы 3.1
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])
    df['Результат'] = df['3'] >= df['3.1']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['3'].tolist()[0]
        second_value = df['3.1'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 3 больше или равно графы 3.1. Графа 3 = {first_value}, графа 3.1 = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_fourth_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 3 >= графы 3.2
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])
    df['Результат'] = df['3'] >= df['3.2']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['3'].tolist()[0]
        second_value = df['3.2'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 3 больше или равно графы 3.2. Графа 3 = {first_value}, графа 3.2 = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_fifth_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 32 = сумма значений граф с 33 по 59.
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['33', '34', '35', '36', '37', '38',
                         '39', '40', '41', '42', '43', '44',
                         '45', '46', '47', '48', '49', '50',
                         '51', '52', '53', '54', '55', '56', '57', '58', '59']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['32'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['32'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 32 = сумма значений граф с 33 по 59. Графа 32 = {first_value}, сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_six_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 32 >= графы 32.1
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])
    df['Результат'] = df['32'] >= df['32.1']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['32'].tolist()[0]
        second_value = df['32.1'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 32 больше или равно графы 32.1. Графа 32 = {first_value}, графа 32.1 = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_seventh_error_grad(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 32 >= графы 32.2
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])
    df['Результат'] = df['32'] >= df['32.2']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df.iloc[0, 0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['32'].tolist()[0]
        second_value = df['32.2'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Не выполняется условие: Графа 32 больше или равно графы 32.2. Графа 32 = {first_value}, графа 32.2 = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_error_mon_grad_spo(df: pd.DataFrame, name_file: str, correction: int):
    """
    Точка входа для проверки датафрейма занятости выпускников на арифметические ошибки
    :param correction: число строк на которое нужно увеличить результат чтобы показывалась правильная строка у ошибки
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    border = 0
    for i in range(1, len(df) + 1):
        row_df = df.iloc[border, :].to_frame().transpose()  # получаем датафрейм строку

        # Проводим проверку Графа 2 = 3 + 31 + 32 + 60 + 61 + 62+ 63 + 64 + 65 + 66 + 67 + 68 + 69 + 70.
        first_error_df_grad = check_first_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, first_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 3 = сумма значений граф с 4 по 30.
        second_error_df_grad = check_second_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, second_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 3.1 <= Графа 3
        third_error_df_grad = check_third_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, third_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 3.2 <= Графа 3
        fourth_error_df_grad = check_fourth_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, fourth_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 32 = сумма значений граф с 33 по 59
        fifth_error_df_grad = check_fifth_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, fifth_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 32 >= Графа 32.1
        six_error_df_grad = check_six_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, six_error_df_grad], axis=0, ignore_index=True)
        # Проводим проверку Графа 32 >= Графа 32.2
        seventh_error_df_grad = check_seventh_error_grad(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, seventh_error_df_grad], axis=0, ignore_index=True)

        border += 1

    return error_df


def check_first_error_grad_target(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 5 = сумма граф 6 + 7 + 8 + 9 + 10 + 11 + 12.
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['6', '7', '8', '9', '10', '11',
                         '12']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['5'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df['1'].tolist()[0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['5'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Лист Выпуск-Целевое. Не выполняется условие: Графа 5 = сумма значений граф с 6 по 12. Графа 5 = {first_value}, сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_error_mon_grad_target(spo_df: pd.DataFrame, target_df: pd.DataFrame, name_file, correction):
    """
    Функция для проверки правильности заполнения листа 2 Целевой выпуск
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    group_first_check_df = target_df.groupby('1').agg({'5': 'sum', '6': 'sum'}).reset_index()
    group_first_check_df.rename(
        columns={'1': 'Код и наименование', '5': 'Численность целевиков', '6': 'Трудоустроено целевиков'}, inplace=True)
    spo_df = spo_df[['1', '2', '3']]
    spo_df.rename(
        columns={'1': 'Код и наименование', '2': 'Суммарный выпуск', '3': 'Всего трудоустроено'}, inplace=True)
    # проверяем на совпадение специальностей
    both_in_two_tables_df = pd.merge(group_first_check_df, spo_df, how='outer', left_on='Код и наименование',
                                     right_on='Код и наименование', indicator=True)
    # получаем все специальности которых нет на листе СПО-1
    not_spo_sheet_df = both_in_two_tables_df[both_in_two_tables_df['_merge'] == 'left_only']
    if len(not_spo_sheet_df) != 0:
        # Получаем строку с перечислением ошибочных специальностей
        error_lst = not_spo_sheet_df['Код и наименование'].tolist()
        for error_spec in error_lst:
            temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                         data=[[name_file, f'{error_spec}',
                                                f'Профессия, специальность отсутствует на листе Выпуск-СПО']])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # Проверяем на размер суммарного выпуска
    # получаем все специальности которые есть в обоих листах
    both_spec_df = both_in_two_tables_df[both_in_two_tables_df['_merge'] == 'both']
    if len(both_spec_df) != 0:
        # Перебираем построчно
        for row in both_spec_df.itertuples():
            name_spec = row[1]
            quantity_target = int(row[2])  # численость целевиков по специальности
            worker_target = int(row[3])  # трудоустроено целевиков по специальности
            all_release = int(row[4])  # суммарный выпуск по специальности
            all_worker = int(row[5])  # всего трудоустроено

            # Проверки
            """
            Суммарная численность выпускников по каждой профессии, специальности на вкладке «2. Выпуск – Целевое» 
            не может превышать суммарный выпуск
             по этой профессии, специальности на вкладке «1. Выпуск – СПО»
            """
            if quantity_target > all_release:
                temp_error_df = pd.DataFrame(
                    columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                    data=[[name_file, f'{name_spec}',
                           f'Лист Выпуск-Целевое. Суммарная численность целевиков по специальности больше чем суммарный выпуск специальности '
                           f'указанный в графе 2 на листе Выпуск-СПО. Целевиков- {quantity_target} а суммарная численность - {all_release}']])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            """
            Численность трудоустроенных выпускников по каждой профессии, специальности на вкладке «2. Выпуск – Целевое» 
            не может превышать численность трудоустроенных выпускников по этой профессии, специальности на вкладке «1. Выпуск – СПО».
            """

            if worker_target > all_worker:
                temp_error_df = pd.DataFrame(
                    columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                    data=[[name_file, f'{name_spec}',
                           f'Лист Выпуск-Целевое. Численность  трудоустроенных целевиков по специальности в графе 6 больше чем количество трудоустроенных '
                           f'указанное в графе 3 на листе Выпуск-СПО. Трудоустроено целевиков- {worker_target} всего трудоустроено - {all_worker}']])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
    # Проверяем сумму целевиков
    """
    Графа 5 = сумма граф 6 + 7 + 8 + 9 + 10 + 11 + 12.
    """
    border = 0
    for i in range(1, len(target_df) + 1):
        row_df = target_df.iloc[border, :].to_frame().transpose()  # получаем датафрейм строку
        first_error_df_grad_target = check_first_error_grad_target(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, first_error_df_grad_target], axis=0, ignore_index=True)
        border += 1

    return error_df


def check_first_error_grad_prof(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 8 = сумма граф 9 + 10 + 11 + 12+ 13.
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['9', '10', '11',
                         '12','13']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['8'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df['1'].tolist()[0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['8'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Лист Выпуск-Профессионалитет. Не выполняется условие: Графа 8 = сумма значений граф с 9 по 13. Графа 8 = {first_value}, сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df

def check_second_error_grad_prof(df: pd.DataFrame, name_file: str, number_row: int):
    """
    Функция для проверки условия Графа 6 = сумма граф 7 и 8.
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    check_sum_columns = ['7', '8']
    df['Сумма'] = df[check_sum_columns].sum(axis=1)
    df['Результат'] = df['6'] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    name_spec = df['1'].tolist()[0]
    if df.iloc[0, -1] == 'Неправильно':
        first_value = df['6'].tolist()[0]
        second_value = df['Сумма'].tolist()[0]
        temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'],
                                     data=[[name_file, f'Строка {number_row}- {name_spec}',
                                            f'Лист Выпуск-Профессионалитет. Не выполняется условие: Графа 6 = сумма значений граф  7 и 8. Графа 6 = {first_value}, сумма граф = {second_value}']])
        return temp_error_df

    return temp_error_df


def check_error_mon_grad_prof(prof_df:pd.DataFrame,name_file, correction):
    """
    Функция для подсчета данных с листа профессионалитет
    """

    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    border = 0
    for i in range(1, len(prof_df) + 1):
        row_df = prof_df.iloc[border, :].to_frame().transpose()  # получаем датафрейм строку
        # проверяем Графа 8 = сумма граф  9 + 10 + 11 + 12+13
        first_error_df_grad_prof = check_first_error_grad_prof(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, first_error_df_grad_prof], axis=0, ignore_index=True)
        # Проверяем Графа 6 =  графы 7 +8
        second_error_df_grad_prof = check_second_error_grad_prof(row_df.copy(), name_file, correction + i)
        error_df = pd.concat([error_df, second_error_df_grad_prof], axis=0, ignore_index=True)
        border += 1

    return error_df



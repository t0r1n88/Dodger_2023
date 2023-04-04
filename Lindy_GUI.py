import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
# pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_folder_data():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data
    path_folder_data = filedialog.askdirectory()

def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()


def select_files_data_xlsx():
    """
    Функция для выбора нескоьких файлов с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global files_data_xlsx
    # Получаем путь файлы
    files_data_xlsx = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def check_data(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0


def check_data_note(cell):
    if cell is np.nan:
        return 'Не заполнено'
    return str(cell)


def check_first_error(df: pd.DataFrame, name_file):
    """
    Функция для проверки гр. 09 и гр. 10 < гр. 08
    """

    df['Сумма'] = df['09'] + df['10']
    # Проводим проверку
    df['Результат'] = df['08'] >= df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 9, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 09 и гр. 10 <= гр. 08 '
    return temp_error_df


def check_second_error(df: pd.DataFrame, name_file):
    """
    Функция для проверки правильности введеденных данных
    (гр. 07= гр.08 + сумма(с гр.11 по гр.32))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # конвертируем в инт
    all_sum_cols = list(df)
    # удаляем колонки 07, 09, 10
    all_sum_cols.remove('07')
    all_sum_cols.remove('09')
    all_sum_cols.remove('10')
    # получаем сумму колонок 08, 11:32
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['07'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 9, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 07 = гр.08 + сумма(с гр.11 по гр.32)'
    return temp_error_df


def check_third_error(df: pd.DataFrame, name_file, border):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр. 02 + стр. 04
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    foo_df = pd.DataFrame(columns=['02', '04', '06', ])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['Сумма'] = foo_df['02'] + foo_df['04']
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + 9} - {border + 23} колонка {str(x)}', raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр. 02 + стр. 04 '
    return temp_error_df


def check_fourth_error(df: pd.DataFrame, name_file, border):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13

    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """

    foo_df = pd.DataFrame(columns=['06', '07', '08', '09', '10', '11', '12', '13'])

    # Добавляем данные в датафрейм
    foo_df['06'] = df.iloc[5, :]
    foo_df['07'] = df.iloc[6, :]
    foo_df['08'] = df.iloc[7, :]
    foo_df['09'] = df.iloc[8, :]
    foo_df['10'] = df.iloc[9, :]
    foo_df['11'] = df.iloc[10, :]
    foo_df['12'] = df.iloc[11, :]
    foo_df['13'] = df.iloc[12, :]

    sum_col = ['07', '08', '09', '10', '11', '12', '13']
    foo_df['Сумма'] = foo_df[sum_col].sum(axis=1)
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + 9} - {border + 23} колонка {str(x)}', raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13 '

    return temp_error_df


def check_fifth_error(df: pd.DataFrame, name_file, border):
    """
    Функция для проверки правильности введеденных данных
    стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    foo_df = pd.DataFrame(columns=['05', '06', '14'])

    # Добавляем данные в датафрейм
    foo_df['05'] = df.iloc[4, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['14'] = df.iloc[13, :]

    foo_df['Результат'] = (foo_df['14'] <= foo_df['05']) & (foo_df['14'] <= foo_df['06'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df[
        'index'].tolist()  # делаем список, прибавляем для того чтобы номера строк совпадали с строками в файле
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк  {border + 9} - {border + 23} колонка {str(x)}', raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 14<=стр. 06, стр. 14<=стр 05'
    return temp_error_df


def check_error(df: pd.DataFrame,name_file):
    """
    Функция для проверки данных
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 6:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)
        # прибавляем border
        border += 15
    # Возвращаем датафрейм с ошибками
    return error_df


def create_check_tables(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

                # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        wb.create_sheet(title=code_spec, index=idx)
    del wb['Sheet']  # удаляем лишний лист

    for code_spec in code_spec_dct.keys():
        temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
        temp_code_df = temp_code_df.stack()
        temp_code_df = temp_code_df.to_frame()

        temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
        temp_code_df[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
        temp_code_df[
            'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
        temp_code_df[
            'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
        temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
        temp_code_df['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
        temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
        temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
        temp_code_df[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
        temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(lambda x: x.get('Колонка 16'))
        temp_code_df['Неформальная занятость (нелегальная)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
        temp_code_df[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
        temp_code_df[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
        temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = temp_code_df[0].apply(
            lambda x: x.get('Колонка 20'))
        temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
        temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
            lambda x: x.get('Колонка 22'))
        temp_code_df[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
        temp_code_df[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
        temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
            lambda x: x.get('Колонка 25'))
        temp_code_df[
            'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
        temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
        temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
            lambda x: x.get('Колонка 28'))
        temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
        temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
        temp_code_df[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 31'))
        temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
        temp_code_df['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = \
        temp_code_df[0].apply(lambda x: x.get('Колонка 33'))

        finish_code_spec_df = temp_code_df.drop([0], axis=1)

        finish_code_spec_df = finish_code_spec_df.reset_index()

        finish_code_spec_df.rename(
            columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении',
               'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
            ,
               'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
               'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
               'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
               'Строка 12': 'расстройствами аутистического спектра',
               'Строка 13': 'с инвалидностью вследствие  других причин',
               'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
               'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
               }
        finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
            wb[code_spec].append(r)
        wb[code_spec].column_dimensions['A'].width = 20
        wb[code_spec].column_dimensions['B'].width = 40

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)
    wb.save(f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')


def processing_data_employment():
    """
    Фугкция для обработки данных
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    try:
        for file in os.listdir(path_folder_data):
            name_file = file.split('.xlsx')[0]
            print(name_file)
            df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=7, dtype=str)
            # получаем  часть с данными
            mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
            # проверяем есть ли строка полностью состоящая из nan
            if True in mask:
                df = df.iloc[:mask.idxmax()] # если есть то отсекаем все что ниже такой строки
                # Проверка на размер таблицы, должно бьть кратно 15
            if df.shape[0] % 15 != 0:
                temp_error_df = pd.DataFrame(data=[
                    [f'{name_file}', '', 'КОЛИЧЕСТВО СТРОК С ДАННЫМИ НЕ КРАТНО 15 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            check_code_lst = df['03'].tolist()  # получаем список кодов специальностей
            # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
            if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            # Проверка на непрерывность кода специальности, то есть на 15 строк должен быть только один код
            border_check_code = 0  # счетчик обработанных страниц
            quantity_check_code = len(check_code_lst) // 15  # получаем сколько специальностей в таблице
            flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 15 строк
            flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
            for i in range(quantity_check_code):
                # получаем множество отбрасывая np.nan
                temp_set = set([code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 15] if
                                code_spec is not np.nan])

                if len(temp_set) != 1:
                    flag_error_code_spec = True
                if ' ' in temp_set:
                    flag_error_space_spec = True
                border_check_code += 15

            if flag_error_space_spec:
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            if flag_error_code_spec:
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД СПЕЦИАЛЬНОСТИ НА КАЖДЫЕ 15 СТРОК !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            df.columns = list(map(str, df.columns))
            # Заполняем пока пропуски в 15 ячейке для каждой специальности
            df['06'] = df['06'].fillna('15 ячейка')

            # Проводим проверку на корректность данных, отправляем копию датафрейма
            file_error_df = check_error(df.copy(), name_file)
            error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)

            # очищаем от нан и возможнных пустых пробелов
            code_spec = [spec for spec in df['03'].unique() if spec is not np.nan]
            code_spec = [spec for spec in code_spec if spec != ' ']

            # Создаем список для строк
            row_cat = [f'Строка {i}' for i in range(1, 16)]
            # Создаем список для колонок
            column_cat = [f'Колонка {i}' for i in range(7, 34)]

            # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
            spec_dict = {}
            for row in row_cat:
                spec_dict[row] = {key: 0 for key in column_cat}

            # Изменяем последний ключ на строковый поскольку там будут хранится примечания
            for row, value in spec_dict.items():
                for col, data in value.items():
                    if col == 'Колонка 33':
                        spec_dict[row][col] = ''
            # Создаем словарь среднего уровня содержащй данные по всем специальностям
            poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

            high_level_dct[name_file] = copy.deepcopy(poo_dct)

            """
            В итоге получается такая структура
            {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

            """

            current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

            idx_row = 1  # счетчик обработанных строк

            # Итерируемся по полученному датафрейму через itertuples
            for row in df.itertuples():
                # если счетчик колонок больше 15 то уменьшаем его до единицы
                if idx_row > 15:
                    idx_row = 1

                # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                if (row[3] is not np.nan) and (row[3] != ' '):
                    # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                    if row[3] != current_code:
                        current_code = row[3]

                data_row = row[7:34]  # получаем срез с нужными данными

                for idx_col, value in enumerate(data_row, start=1):
                    if idx_col + 6 == 33:
                        # сохраняем примечания в строке
                        high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                            f'Колонка {idx_col + 6}'] = f'{name_file} {check_data_note(value)};'

                    else:
                        high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                            f'Колонка {idx_col + 6}'] += check_data(value)

                idx_row += 1

        create_check_tables(high_level_dct)

        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        if col == 'Колонка 33':
                            itog_df[code_spec][row][col] += check_data_note(value)
                        else:
                            itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()

        frame['Всего'] = frame[0].apply(lambda x: x.get('Колонка 7'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
        frame[0].apply(lambda x: x.get('Колонка 8'))
        frame[
            'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
        frame[0].apply(lambda x: x.get('Колонка 9'))
        frame[
            'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
        frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Индивидуальные предприниматели'] = frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = frame[
            0].apply(lambda x: x.get('Колонка 12'))
        frame['Продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 13'))
        frame['Проходят службу в армии по призыву'] = frame[0].apply(lambda x: x.get('Колонка 14'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
        frame[0].apply(lambda x: x.get('Колонка 15'))
        frame['Находятся в отпуске по уходу за ребенком'] = frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Неформальная занятость (нелегальная)'] = frame[0].apply(lambda x: x.get('Колонка 17'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
        frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
        frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Иные причины нахождения под риском нетрудоустройства'] = frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Смерть, тяжелое состояние здоровья'] = frame[0].apply(lambda x: x.get('Колонка 21'))
        frame['Находятся под следствием, отбывают наказание'] = frame[0].apply(lambda x: x.get('Колонка 22'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
        frame[0].apply(lambda x: x.get('Колонка 23'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
        frame[0].apply(lambda x: x.get('Колонка 24'))
        frame['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = frame[0].apply(
            lambda x: x.get('Колонка 25'))
        frame[
            'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
        frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['будут трудоустроены'] = frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['будут осуществлять предпринимательскую деятельность'] = frame[0].apply(lambda x: x.get('Колонка 28'))
        frame['будут самозанятыми'] = frame[0].apply(lambda x: x.get('Колонка 29'))
        frame['будут призваны в армию'] = frame[0].apply(lambda x: x.get('Колонка 30'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
        frame[0].apply(lambda x: x.get('Колонка 31'))
        frame['будут продолжать обучение'] = frame[0].apply(lambda x: x.get('Колонка 32'))
        frame['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = frame[
            0].apply(lambda x: x.get('Колонка 33'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении',
               'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
            ,
               'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
               'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
               'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
               'Строка 12': 'расстройствами аутистического спектра',
               'Строка 13': 'с инвалидностью вследствие  других причин',
               'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
               'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
               }
        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])
        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        finish_df.to_excel(f'{path_to_end_folder}/Итоговая таблица от {current_time}.xlsx', index=False)

        # Создаем документ
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ от {current_time}.xlsx')

    except NameError:
        messagebox.showerror('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1',
                             f'Закройте открытые файлы Excel {e.args}')

    else:
        messagebox.showinfo('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1',
                            'Данные успешно обработаны')


if __name__ == '__main__':
    window = Tk()
    window.title('ЛИНДИ Подсчет данных по трудоустройству выпускников ver 2.1')
    window.geometry('700x860')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_employment = ttk.Frame(tab_control)
    tab_control.add(tab_employment, text='Подсчет данных по трудоустройству')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_employment,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_employment,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data = Button(tab_employment, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                             command=select_folder_data
                             )
    btn_choose_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_employment, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_employment, text='3) Обработать данные', font=('Arial Bold', 20),
                                  command=processing_data_employment
                                  )
    btn_proccessing_data.grid(column=0, row=4, padx=10, pady=10)

    window.mainloop()
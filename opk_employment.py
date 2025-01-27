# -*- coding: utf-8 -*-
"""
Скрипт для подсчета данных центров карьеры
"""
from check_opk_functions import * # импортируем функции проверки файлов ОПК
from cass_check_functions import * # импортируем функции общих проверок
from support_functions import * # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random



def prepare_opk_employment(path_folder,path_to_end_folder):
    """
    Функция для обработки полной таблицы занятости выпускников в ОПК
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    all_form2 = pd.DataFrame(columns=['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11'])

    try:
        for file in os.listdir(path_folder):
            if not file.startswith('~$') and file.endswith('.xls'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Файл с расширением XLS (СТАРЫЙ ФОРМАТ EXCEL)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                # Проверяем наличие листов с названиями Форма 1 и Форма 2
                wb_1 = openpyxl.load_workbook(f'{path_folder}/{file}')
                if not {'Форма 1', 'Форма 2'}.issubset(set(wb_1.sheetnames)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте наличие листов с названием Форма 1 и Форма 2! Не должно быть пробелов в начале и в конце названия ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df_form1 = pd.read_excel(f'{path_folder}/{file}', skiprows=8, dtype=str,
                                         sheet_name='Форма 1')  # общие данные

                # Находим строку с номерами колонок, так как вполне возможно в файле остались примеры
                temp_wb = openpyxl.load_workbook(f'{path_folder}/{file}',
                                                 read_only=True)  # открываем файл в режиме чтения
                temp_ws = temp_wb['Форма 2']
                threshold_form2 = 5
                for row in temp_ws.iter_rows(0):  # перебираем значения в первой колонке
                    for cell in row:
                        if cell.value == '01':
                            threshold_form2 = cell.row
                temp_wb.close()  # закрываем файл чтобы потом не было ошибок
                form2_df = pd.read_excel(f'{path_folder}/{file}', skiprows=threshold_form2 - 1, dtype=str,
                                         sheet_name='Форма 2')  # подробные данные по ОПК
                # создаем множество колонок наличие которых мы проверяем
                check_cols = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                              '16',
                              '17',
                              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                              '33',
                              '34', '35', '36',
                              '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51',
                              '52',
                              '53', '54', '55',
                              '56', '57', '58', '59', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70',
                              '71',
                              '72', '73', '74',
                              '75', '76', '77', '78', '79', '80'}
                if not check_cols.issubset(set(df_form1.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы на листе Форма 1.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 9 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # проверяем корректность формы 2
                check_cols_form2 = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11'}
                if not check_cols_form2.issubset(set(form2_df.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы на листе Форма 2.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться ВЫШЕ списка трудоустроенных.! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df_form1 = df_form1[df_form1['04'] != '03']  # фильтруем строки с проверкой
                form2_df.dropna(axis=0, inplace=True, how='all')  # убираем все пустые строки
                # заполняем пустые строки в колонке 01
                form2_df['01'] = form2_df['01'].fillna('Не заполнено')
                form2_df = form2_df[
                    ~form2_df['01'].str.contains('Проверка', case=False)]  # фильруем строки с проверкой на листе 2
                form2_df = form2_df[
                    ~form2_df['01'].str.contains('Не заполнено', case=False)]  # фильруем строки с проверкой на листе 2

                df_form1 = df_form1.loc[:, '01':'78']  # отсекаем возможную первую колонку и колонки с примечаниями
                # получаем  часть с данными
                mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
                if mask[0]:  # если пустая строка идет первой то удаляем ее и обновляем маску
                    df_form1.drop(axis=0, index=0, inplace=True)
                    mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # Находим индекс первой пустой строки, если он есть,получаем список с значениями где есть пустые строки
                empty_row_index = np.where(df_form1.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df_form1 = df_form1.iloc[:row_index]
                quantity_spec = df_form1.shape[0] // 2  # получаем количество специальностей в файле

                check_two_rows_spec = df_form1['04'].tolist() == ['01',
                                                                  '02'] * quantity_spec  # проверяем чтобы колонка 04 состояла только из 01 и 02
                if not check_two_rows_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Возможно пропущена строка 01 или 02. Для каждой спец./проф. должны быть  только строки 01 и 02 не считая строки 03 с проверкой. Проверьте наличие ПУСТОЙ строки после таблицы! Последняя строка проверки должна быть заполнена полностью! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                check_code_lst = df_form1['03'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue

                # Проверка на непрерывность кода специальности, то есть на 2 строки должен быть только один код
                border_check_code = 0  # счетчик обработанных страниц
                quantity_check_code = len(check_code_lst) // 2  # получаем сколько специальностей в таблице
                flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 2 строки
                flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
                for i in range(quantity_check_code):
                    # получаем множество отбрасывая np.nan
                    temp_set = set([code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 2]])
                    if len(temp_set) != 1:
                        flag_error_code_spec = True
                    if ' ' in temp_set:
                        flag_error_space_spec = True
                    border_check_code += 2

                if flag_error_space_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if flag_error_code_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД и Название СПЕЦИАЛЬНОСТИ/ПРОФЕССИИ НА КАЖДЫЕ 2 СТРОКИ (не считая строки с проверкой)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                """
                ПРОВЕРКИ
                в том числе проверка кода специальности

                """
                tup_correct = (10, 12)  # создаем кортеж  с поправками
                file_error_df = check_error_opk(df_form1.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)

                # проводим кросс проверки между 2 формами
                file_cross_error_df = check_cross_error_opk(df_form1.copy(), form2_df.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_cross_error_df], axis=0, ignore_index=True)

                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if file_cross_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки связанные с сравнением данных из формы 1 и формы 2!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df_form1['03'] = df_form1['03'].apply(extract_code_full)  # очищаем от текста в кодах
                if 'Не найден код специальности' in df_form1['03'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 03 Код специальности.Вместо кода присутствует дата,слитное написание кода и названия, вместе с кодом есть название,пробел перед кодом и т.п.!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                """
                Окончание блока проверок
                """
                all_form2 = pd.concat([all_form2, form2_df], axis=0, ignore_index=True)

                code_spec = [spec for spec in df_form1['03'].unique()]

                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 3)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(6, 79)]

                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}

                # Изменяем последний ключ на строковый поскольку там будут хранится примечания
                for row, value in spec_dict.items():
                    for col, data in value.items():
                        if col == 'Колонка 78':
                            spec_dict[row][col] = ''

                # Создаем словарь среднего уровня содержащй данные по всем специальностям
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

                high_level_dct[name_file] = copy.deepcopy(poo_dct)

                """
                В итоге получается такая структура
                {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                """

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк
                # Итерируемся по полученному датафрейму через itertuples
                for row in df_form1.itertuples():
                    # если счетчик колонок больше 15 то уменьшаем его до единицы
                    if idx_row > 2:
                        idx_row = 1
                        # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[3] is not np.nan) and (row[3] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[3] != current_code:
                            current_code = row[3]
                    data_row = row[6:79]  # получаем срез с нужными данными
                    for idx_col, value in enumerate(data_row, start=1):
                        if idx_col + 5 == 78:
                            # сохраняем примечания в строке
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 5}'] = f'{name_file} {check_data_note(value)};'

                        else:
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 5}'] += check_data(value)
                    idx_row += 1
        # создаем файл для проверки. Специальности с разбивкой по организациям
        wb_opk_check_tables = create_check_tables_opk(high_level_dct)
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        wb_opk_check_tables.save(f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')


        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        if col == 'Колонка 78':
                            itog_df[code_spec][row][col] += check_data_note(value)
                        else:
                            itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()

        frame['Суммарный выпуск 2023 г.'] = frame[0].apply(lambda x: x.get('Колонка 6'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 7'))
        frame['Трудоустроены на предприятия оборонно-промышленного комплекса*'] = \
            frame[0].apply(lambda x: x.get('Колонка 8'))
        frame['Трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 9'))
        frame['Трудоустроены на предприятия сельского хозяйства'] = \
            frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Трудоустроены на предприятия металлургии'] = \
            frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Трудоустроены на предприятия железнодорожного транспорта'] = \
            frame[0].apply(lambda x: x.get('Колонка 12'))
        frame['Трудоустроены на предприятия легкой промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 13'))
        frame['Трудоустроены на предприятия химической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 14'))
        frame['Трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 15'))
        frame['Трудоустроены на предприятия фармацевтической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Трудоустроены на предприятия отрасли информационных технологий'] = \
            frame[0].apply(lambda x: x.get('Колонка 17'))
        frame['Трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Трудоустроены на предприятия транспортной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Трудоустроены на предприятия горнодобывающей отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 21'))
        frame[
            'Трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 22'))
        frame['Трудоустроены на предприятия лесной промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 23'))
        frame['Трудоустроены на предприятия строительной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 24'))
        frame[
            'Трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 25'))
        frame['Трудоустроены на предприятия индустрии робототехники'] = \
            frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['Трудоустроены на предприятия в отрасли образования'] = \
            frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['Трудоустроены на предприятия в медицинской отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 28'))
        frame[
            'Трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
            frame[0].apply(lambda x: x.get('Колонка 29'))
        frame['Трудоустроены на предприятия в отрасли искусства'] = \
            frame[0].apply(lambda x: x.get('Колонка 30'))
        frame['Трудоустроены на предприятия иная отрасль'] = \
            frame[0].apply(lambda x: x.get('Колонка 31'))

        frame['Индивидуальные предприниматели'] = \
            frame[0].apply(lambda x: x.get('Колонка 32'))
        frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессиональный доход)'] = \
            frame[0].apply(lambda x: x.get('Колонка 33'))
        frame['Продолжили обучение'] = \
            frame[0].apply(lambda x: x.get('Колонка 34'))
        frame['Проходят службу в армии по призыву'] = \
            frame[0].apply(lambda x: x.get('Колонка 35'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
            frame[0].apply(lambda x: x.get('Колонка 36'))
        frame['Находятся в отпуске по уходу за ребенком'] = \
            frame[0].apply(lambda x: x.get('Колонка 37'))
        frame[
            'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 38'))
        frame['Будут трудоустроены на предприятия оборонно-промышленного комплекса* '] = \
            frame[0].apply(lambda x: x.get('Колонка 39'))
        frame['Будут трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 40'))
        frame['Будут трудоустроены на предприятия сельского хозяйства'] = \
            frame[0].apply(lambda x: x.get('Колонка 41'))
        frame['Будут трудоустроены на предприятия металлургии'] = \
            frame[0].apply(lambda x: x.get('Колонка 42'))
        frame['Будут трудоустроены на предприятия железнодорожного транспорта'] = \
            frame[0].apply(lambda x: x.get('Колонка 43'))
        frame['Будут трудоустроены на предприятия легкой промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 44'))
        frame['Будут трудоустроены на предприятия химической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 45'))
        frame['Будут трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 46'))
        frame['Будут трудоустроены на предприятия фармацевтической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 47'))
        frame['Будут трудоустроены на предприятия отрасли информационных технологий'] = \
            frame[0].apply(lambda x: x.get('Колонка 48'))
        frame['Будут трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 49'))
        frame[
            'Будут трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 50'))

        frame['Будут трудоустроены на предприятия транспортной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 51'))
        frame['Будут трудоустроены на предприятия горнодобывающей отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 52'))
        frame[
            'Будут трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 53'))
        frame['Будут трудоустроены на предприятия лесной промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 54'))
        frame['Будут трудоустроены на предприятия строительной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 55'))
        frame[
            'Будут трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 56'))
        frame['Будут трудоустроены на предприятия индустрии робототехники'] = \
            frame[0].apply(lambda x: x.get('Колонка 57'))
        frame['Будут трудоустроены на предприятия в отрасли образования'] = \
            frame[0].apply(lambda x: x.get('Колонка 58'))
        frame['Будут трудоустроены на предприятия в медицинской отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 59'))
        frame[
            'Будут трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
            frame[0].apply(lambda x: x.get('Колонка 60'))
        frame['Будут трудоустроены на предприятия в отрасли искусства'] = \
            frame[0].apply(lambda x: x.get('Колонка 61'))
        frame['Будут трудоустроены на предприятия иная отрасль'] = \
            frame[0].apply(lambda x: x.get('Колонка 62'))
        frame['Будут индивидуальными предпринимателями'] = \
            frame[0].apply(lambda x: x.get('Колонка 63'))
        frame['Будут самозанятыми'] = \
            frame[0].apply(lambda x: x.get('Колонка 64'))
        frame['Будут продолжать обучение'] = \
            frame[0].apply(lambda x: x.get('Колонка 65'))
        frame['Будут призваны в армию'] = \
            frame[0].apply(lambda x: x.get('Колонка 66'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
            frame[0].apply(lambda x: x.get('Колонка 67'))
        frame['Будут находиться в отпуске по уходу за ребенком'] = \
            frame[0].apply(lambda x: x.get('Колонка 68'))
        frame['Неформальная занятость (теневой сектор экономики)'] = \
            frame[0].apply(lambda x: x.get('Колонка 69'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
            frame[0].apply(lambda x: x.get('Колонка 70'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
            frame[0].apply(lambda x: x.get('Колонка 71'))
        frame['Иные причины нахождения под риском нетрудоустройства'] = \
            frame[0].apply(lambda x: x.get('Колонка 72'))
        frame['Смерть, тяжелое состояние здоровья'] = \
            frame[0].apply(lambda x: x.get('Колонка 73'))
        frame['Находятся под следствием, отбывают наказание'] = \
            frame[0].apply(lambda x: x.get('Колонка 74'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            frame[0].apply(lambda x: x.get('Колонка 75'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
            frame[0].apply(lambda x: x.get('Колонка 76'))
        frame[
            'Иное в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 0,5% выпускников - укажите причины в графе "принимаемые меры"'] = \
            frame[0].apply(lambda x: x.get('Колонка 77'))
        frame[
            'Принимаемые меры по содействию занятости, в том числе по трудоустройству выпускников на предприятия оборонно-промышленного комплекса тезисно - вид меры, охват выпускников мерой'] = \
            frame[0].apply(lambda x: x.get('Колонка 78'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из строки 01: имеют договор о целевом обучении'}

        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        # создаем итоговый отчет по ОПК
        opk_df = finish_df[['Суммарный выпуск 2023 г.',
                            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                            'Трудоустроены на предприятия оборонно-промышленного комплекса*',
                            'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                            'Будут трудоустроены на предприятия оборонно-промышленного комплекса* ']]
        opk_df = opk_df.sum(axis=0).to_frame()  # суммируем данные
        opk_df = opk_df.transpose()  # разворачиваем из колонки в строку

        # Создаем сумму по всем колонкам
        all_sum_df = finish_df.iloc[:, 2:].sum(axis=0).to_frame()  # суммируем данные
        all_sum_df = all_sum_df.transpose()  # разворачиваем из колонки в строку

        # добавляем строки с проверкой
        count = 0
        for i in range(2, len(finish_df) + 1, 2):
            new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
            new_row.iloc[:, 1] = 'Проверка (строка не редактируется) Строку требуется подставлять после каждого кода'
            new_row.iloc[:, 2:] = 'проверка пройдена'

            # Вставка новой строки через каждые 15 строк
            finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(
                drop=True)
            count += 1

        lst_number_row = ['01', '02', '03']
        multipler = len(finish_df) // 3  # получаем количество специальностей/профессий
        # вставляем новую колонку
        finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))

        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        # создаем сокращенный датафрейм по специальностям из плоной таблицы
        opk_finish_df = finish_df[['Код специальности', 'Номер строки',
                                   'Наименование показателей (категория выпускников)',
                                   'Суммарный выпуск 2023 г.',
                                   'Трудоустроены на предприятия оборонно-промышленного комплекса*',
                                   'Будут трудоустроены на предприятия оборонно-промышленного комплекса* ']]

        opk_finish_df.to_excel(
            f'{path_to_end_folder}/Трудоустройство по специальностям (ОПК) от {current_time}.xlsx',
            index=False)
        finish_df.to_excel(
            f'{path_to_end_folder}/Полная таблица Трудоустройство по отраслям от {current_time}.xlsx',
            index=False)

        # считаем сколько целевиков
        target_df = finish_df[finish_df['Номер строки'] == '02']
        target_df = target_df.iloc[:, 2:].sum(axis=0).to_frame()  # суммируем данные
        target_df = target_df.transpose()  # разворачиваем из колонки в строку
        # сохраняем результирующие датафреймы в один файл
        with pd.ExcelWriter(f'{path_to_end_folder}/Итоги по ОПК,целевикам и всем колонкам {current_time}.xlsx',
                            engine='openpyxl') as writer:
            opk_df.to_excel(writer, sheet_name='Итог по ОПК')
            all_sum_df.to_excel(writer, sheet_name='Итог по всем колонкам')
            target_df.to_excel(writer, sheet_name='Итог по целевикам')

        # обрабатываем список из второй формы
        all_form2.columns = ['Субъект Российской Федерации', 'Код и наименование профессии, специальности',
                             'Наименование образовательной организации среднего профессионального образования, в которой обучался выпускник',
                             'Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса',
                             'Категория выпускников (выпускники, которые уже трудоустроены/выпускники, которые будут трудоустроены)',
                             'Наличие договора о целевом обучении (имеющие договор о целевом обучении/не имеющие договор о целевом обучении) с предприятием оборонно-промышленного комплекса',
                             'ИНН предприятия',
                             'Наименование предприятия/филиала предприятия/структурного подразделения предприятия оборонно-промышленного комплекса, на котором трудоустроен выпускник/планирует трудоустройство',
                             'Наименование профессии/должности, по которой трудоустраивается (планирует трудоустройство) выпускник',
                             'Трудоустройство в соответствии с освоенной профессией, специальностью (да/нет)',
                             'Сложности при трудоустройстве (например, при взаимодействии с предприятием, наличие рисков расторжения договора о целевом обучении и т.д.), реализуемые меры (описательная часть. В отсутствие сложностей - пропустите графу)']

        all_form2['Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса']=all_form2['Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса'].apply(check_data)
        #проверяем на пустоту
        if all_form2.shape[0] != 0:
            # Создаем сводную таблицу
            #переименываем колонки для удобства
            all_form2.columns = ['Регион', 'Специальность', 'Наименование', 'Количество', 'Трудоустройство',
                                 'Целевой договор', 'ИНН', 'Предприятие', 'Должность', 'Трудоустройство по специальности',
                                 'Сложности']
            # делаем категориальными значения в некоторых колонках
            all_form2['Трудоустройство'] = all_form2['Трудоустройство'].astype('category')
            all_form2['Трудоустройство'] = all_form2["Трудоустройство"].cat.set_categories(["уже трудоустроены", "будут трудоустроены"])

            all_form2['Целевой договор'] = all_form2['Целевой договор'].astype('category')
            all_form2['Целевой договор'] = all_form2["Целевой договор"].cat.set_categories(["нет", "заключили договор о целевом обучении"])

            all_form2['Трудоустройство по специальности'] = all_form2['Трудоустройство по специальности'].astype('category')
            all_form2["Трудоустройство по специальности"]=all_form2["Трудоустройство по специальности"].cat.set_categories(["нет", "да"])


            out_svod_all_form2 = all_form2.pivot_table(index=['Специальность'],
                                                       values=['Количество'],
                                                       columns=['Трудоустройство', 'Целевой договор'],
                                                       aggfunc={'Количество': sum},
                                                       margins=True)

            out_svod_all_form2.fillna(0,inplace=True)
            out_svod_all_form2 = out_svod_all_form2.applymap(int)

            out_svod_all_form2.rename(index={'All': 'Итого'}, columns={'All': 'Итого'}, inplace=True)
        else:
            out_svod_all_form2 = pd.DataFrame(columns=['нет данных'])


        with pd.ExcelWriter(f'{path_to_end_folder}/Общий список и сводная таблица по форме 2 от {current_time}.xlsx',
                            engine='openpyxl') as writer:
            all_form2.to_excel(writer, sheet_name='Общий список',index=False)
            out_svod_all_form2.to_excel(writer,sheet_name='Сводная таблица')


        error_df.to_excel(f'{path_to_end_folder}/Ошибки ОПК от {current_time}.xlsx', index=False)
    # except NameError:
    #     messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
    #                          f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    # except:
    #     messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
    #                          f'При обработке файла {name_file} возникла ошибка !!!\n'
    #                          f'Проверьте файл на соответствие шаблону.')
    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                                'Данные успешно обработаны.')

if __name__ == '__main__':
    path_data = 'data/example/Форма ОПК по отраслям( 2 строки)'
    path_end = 'data/result/Форма ОПК по отраслям( 2 строки)'
    prepare_opk_employment(path_data, path_end)
    print('Lindy Booth')
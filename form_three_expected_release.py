# -*- coding: utf-8 -*-
"""
Скрипт для обработки данных Формы 1 пятистрочной мониторинга занятости выпускников
"""
from check_functions import * # импортируем функции проверки
from support_functions import * # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random


def prepare_form_three_employment(path_folder_data:str,path_to_end_folder):
    """
    Фугкция для обработки данных формы 3 пять строк ожидаемый выпуск
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем словарь верхнего уровня для хранения пары ключ значение где ключ это код специальности а значение- код и наименование
    dct_code_and_name = dict()
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    tup_correct = (7, 11)  # создаем кортеж  с поправками где 6 это первая строка с данными а 10 строка где заканчивается первый диапазон

    for file in os.listdir(path_folder_data):
        if not file.startswith('~$') and not file.endswith('.xlsx'):
            name_file = file.split('.xls')[0]
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue
        if not file.startswith('~$') and file.endswith('.xlsx'):
            name_file = file.split('.xlsx')[0]
            print(name_file)
            # получаем название первого листа
            temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}', read_only=True)
            lst_temp_sheets = temp_wb.sheetnames  # получаем листы в файле
            temp_wb.close()
            if 'Форма 3 ожидаемый выпуск' not in lst_temp_sheets:  # проверяем наличие листа с названием в файле
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Не найден лист с названием Форма 3 ожидаемый выпуск !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=5, dtype=str)
            df.columns = list(map(str, df.columns))  # делаем названия колонок строковыми
            # создаем множество колонок наличие которых мы проверяем
            check_cols = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                          '16', '17',
                          '18', '19', '20', '21', '22', '23', '25', '26', '27', '28']
            if check_cols != list(df.columns):
                diff_cols = set(list(df.columns)).difference(set(check_cols))
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{diff_cols}',
                                                    'Возможно измененная версия формы 3.Строка с номерами колонок (01,02, ... 23,25,26,27,28 как в исходной форме 3 ожидаемый выпуск)\n должна находиться на 6 строке!\n'
                                                    ' указанные колонки являются лишними.Колонки с названимем Unnamed означаеют что на листе есть данные без заголовка в виде цифр на 6 строке  ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            df = df.loc[:, '02':'22']  # отсекаем колонки с регионом и проверками
            # получаем  часть с данными
            mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
            # проверяем есть ли строка полностью состоящая из nan
            empty_row_index = np.where(df.isna().all(axis=1))
            if empty_row_index[0].tolist():
                row_index = empty_row_index[0][0]
                df = df.iloc[:row_index]
            #     # Проверка на размер таблицы, должно бьть кратно 5
            count_spec = df.shape[0] // 5  # количество специальностей
            check_code_lst = df['02'].tolist()  # получаем список кодов специальностей
            # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
            if True in mask.tolist():
                if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В колонке 02 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
            # Проверка на непрерывность кода специальности, то есть на 5 строк должен быть только один код и на пустые ячейки
            border_check_code = 0  # начало отсчета
            quantity_check_code = len(check_code_lst) // 5  # получаем сколько специальностей в таблице
            correction = 0  # размер поправки на случай если есть строка проверки
            sameness_error_df = check_sameness_column(check_code_lst, 5, border_check_code, quantity_check_code,
                                                      tup_correct, correction, name_file, 'Код и наименование')

            blankness_error_df = check_blankness_column(check_code_lst, 5, border_check_code, quantity_check_code,
                                                        tup_correct, correction, name_file, 'Код и наименование')

            # проверяем на арифметические ошибки
            file_error_df = check_error_form_three(df.copy(), name_file, tup_correct)
            # добавляем в получившийся датафейм ошибки однородности диапазона
            file_error_df = pd.concat([file_error_df, sameness_error_df], axis=0, ignore_index=True)
            file_error_df = pd.concat([file_error_df, blankness_error_df], axis=0, ignore_index=True)

            # добавляем в словарь в полные имена из кода и наименования
            for full_name in df['02'].tolist():
                code = extract_code_nose(full_name)  # получаем только цифры
                dct_code_and_name[code] = full_name
            # очищаем от текста чтобы названия листов не обрезались
            df['02'] = df['02'].apply(extract_code_nose)  # очищаем от текста в кодах
            if 'error' in df['02'].values:
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Некорректные значения в колонке 02 Код и наименование профессии/специальности.Вместо кода присутствует дата, и т.п. проверьте правильность заполнения колонки 02!!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                file_error_df = pd.concat([file_error_df, temp_error_df], axis=0, ignore_index=True)
            # добавляем в основной файл с ошибками
            error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
            if file_error_df.shape[0] != 0:
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
                # Создание словаря для хранения данных файла
            code_spec = [spec for spec in df['02'].unique()]  # получаем список специальностей которые есть в файле
            # Создаем список для строк
            row_cat = [f'Строка {i}' for i in range(1, 6)]
            # Создаем список для колонок
            column_cat = [f'Колонка {i}' for i in range(5, 23)]  #
            # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
            spec_dict = {}
            for row in row_cat:
                spec_dict[row] = {key: 0 for key in column_cat}
            poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}
            high_level_dct[name_file] = copy.deepcopy(poo_dct)

            current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

            idx_row = 1  # счетчик обработанных строк

            # Итерируемся по полученному датафрейму через itertuples
            for row in df.itertuples():
                # если счетчик колонок больше 5 то уменьшаем его до единицы
                if idx_row > 5:
                    idx_row = 1
                # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                if (row[1] is not np.nan) and (row[1] != ' '):
                    # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                    if row[1] != current_code:
                        current_code = row[1]
                data_row = row[4:22]  # получаем срез с нужными данными

                for idx_col, value in enumerate(data_row, start=1):
                    high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                        f'Колонка {idx_col + 4}'] += check_data(value)
                #
                idx_row += 1

    t = time.localtime()  # получаем текущее время
    current_time = time.strftime('%H_%M_%S', t)
    wb_check_tables = create_check_tables_form_three(high_level_dct)  # проверяем данные по каждой специальности
    wb_check_tables.save(
        f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')



    # Создаем документ
    wb = openpyxl.Workbook()
    for r in dataframe_to_rows(error_df, index=False, header=True):
        wb['Sheet'].append(r)

    wb['Sheet'].column_dimensions['A'].width = 50
    wb['Sheet'].column_dimensions['B'].width = 40
    wb['Sheet'].column_dimensions['C'].width = 50
    wb.save(f'{path_to_end_folder}/ОШИБКИ Форма 3 Ожидаемый выпуск от {current_time}.xlsx')
    wb.close()



if __name__ == '__main__':
    main_data_folder = 'data/example/Форма 3'
    main_result_folder = 'data/result/Форма 3'
    prepare_form_three_employment(main_data_folder,main_result_folder)

    print('Lindy Booth')
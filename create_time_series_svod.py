"""
Скрипт для создания временных рядов из сводных таблиц созданных Кассандрой
"""
import re

import openpyxl
import pandas as pd
from tkinter import messagebox
import os
import time
import gc


class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass

class NotReqSheets(Exception):
    """
    Обработка случаев когда нет обязательных листов
    """
    pass

class NotReqColumns(Exception):
    """
    Обработка случаев когда нет обязательных колонок
    """
    pass


def parse_date_str(date_str):
    try:
        day, month, year = map(int, date_str.split('.'))
        return year * 10000 + month * 100 + day  # Сортируемое число
    except:
        return 99999999  # Для некорректных дат



def preparing_data(data_folder:str,required_columns:dict,dct_index_svod:dict,error_df:pd.DataFrame,set_error_name_file:set,dct_second_cols:dict,dct_rename_value:dict,dct_abbr:dict):
    """
    Функция для проверки исходных файлов на базовые ошибки и создания списков встречающихся индексов (первой колонки)
    """




    for dirpath, dirnames, filenames in os.walk(data_folder):
        for file in filenames:
            if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm')):
                try:
                    if file.endswith('.xlsx'):
                        name_file = file.split('.xlsx')[0].strip()
                    else:
                        name_file = file.split('.xlsm')[0].strip()
                    # проверяем на правильность даты в названии
                    result_date = re.search(r'\d{2}_\d{2}_\d{4}', name_file)
                    if result_date:
                        file_date = result_date.group()
                        file_date = file_date.replace('_', '.')
                    else:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}',
                                   f'В названии файла отсутствует дата в правильном формате. Требуется формат DD_MM_YYYY т.е. 25.12.2025'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        set_error_name_file.add(name_file)
                        continue
                    # открываем файл для проверки наличия листов и колонок
                    temp_wb = openpyxl.load_workbook(f'{dirpath}/{file}', read_only=True)
                    temp_wb_sheets = set(temp_wb.sheetnames)
                    temp_wb.close()
                    diff_sheets = set(required_columns.keys()).difference(set(temp_wb_sheets))
                    if len(diff_sheets) != 0:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}',
                                   f'Отсутствуют обязательные листы {diff_sheets}. Данные с этих листов не обработаны.'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)

                    # Собираем возможные индексы которые могут встретиться
                    for sheet, lst_cols in required_columns.items():
                        if sheet not in temp_wb_sheets:
                            continue
                        temp_req_df = pd.read_excel(f'{dirpath}/{file}', sheet_name=sheet)
                        if 'Полное название работодателя' in temp_req_df.columns:
                            temp_req_df['Полное название работодателя'] = temp_req_df['Полное название работодателя'].apply(
                                lambda x: x.upper() if isinstance(x, str) else x).replace(dct_abbr, regex=True)
                            temp_req_df = temp_req_df.rename(columns={'Полное название работодателя':'Краткое название работодателя'})
                        if len(temp_req_df) == 0:
                            continue
                        diff_cols = set(lst_cols).difference(set(temp_req_df.columns))
                        if len(diff_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}',
                                       f'На листе {sheet} отсутствуют обязательные колонки {diff_cols}'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            continue


                        # Открываем файл для обработки
                        df = pd.read_excel(f'{dirpath}/{file}', sheet_name=sheet)  # открываем файл
                        if 'Краткое название работодателя' in df.columns:
                            df['Краткое название работодателя'] = df[
                                'Краткое название работодателя'].apply(
                                lambda x: x.upper() if isinstance(x, str) else x).replace(dct_abbr, regex=True)

                        if sheet not in dct_second_cols:
                            dct_index_svod[sheet].update(df[df.columns[0]].unique())
                        else:
                            # получаем данные не из первой колонки
                            # проводим замену устаревших категорий
                            if sheet in dct_rename_value:
                                df[dct_second_cols[sheet][0]] = df[dct_second_cols[sheet][0]].replace(dct_rename_value[sheet])
                            dct_index_svod[sheet].update(df[dct_second_cols[sheet][0]].unique())


                except:
                    temp_error_df = pd.DataFrame(
                        data=[[f'{name_file}',
                               f'Не удалось обработать файл. Возможно файл поврежден'
                               ]],
                        columns=['Название файла',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                         ignore_index=True)
                    set_error_name_file.add(name_file)
                    continue
    return dct_index_svod,error_df,set_error_name_file




def create_dash_df(dct_dash_df:dict,dash_temp_df:pd.DataFrame,sheet:str,result_date:str,dash_special_treatment:dict,dct_value_rename:dict,dct_filter:dict,dct_exclude_filter:dict,dct_vac_df:dict):
    """
    Функция для заполнения словаря для развернутых датафреймов предназначенных для построения графиков по датам
    """
    # Создаем для дашборда
    if sheet not in dash_special_treatment: # Если обычный лист не требующий группировки
        dash_base_df = dct_dash_df[sheet]
        dash_temp_df = dash_temp_df[dash_temp_df[dash_temp_df.columns[0]] != 'Итого']
        dash_temp_df = dash_temp_df[dash_temp_df[dash_temp_df.columns[0]] != 'ИТОГО']

        if 'Квотируемое место' in dash_temp_df.columns:
            dash_temp_df.drop(columns=['Квотируемое место'],inplace=True) # удаляем колонку для квот. Зачем вообще я ее делал?

        if sheet != 'Вакансии для динамики':
            dash_temp_df = dash_temp_df.assign(Данные_на=result_date)
            dash_base_df = pd.concat([dash_base_df, dash_temp_df])
            dash_base_df.fillna(0, inplace=True)
            dct_dash_df[sheet] = dash_base_df
            if sheet == 'Вакансии по отраслям':
                # Создаем лист с подсчетом по общему количеству вакансий
                dash_temp_df = dash_temp_df[dash_temp_df['Сфера деятельности'] != 'Итого']
                itog_dash_vac = dash_temp_df['Количество вакансий'].sum()
                temp_dash_itog_df = pd.DataFrame(columns=['Количество вакансий', 'Данные_на'], data=[[itog_dash_vac, result_date]])
                itog_dash_base_df = dct_dash_df['Всего вакансий']
                itog_dash_base_df = pd.concat([itog_dash_base_df, temp_dash_itog_df])
                itog_dash_base_df.fillna(0, inplace=True)
                dct_dash_df['Всего вакансий'] = itog_dash_base_df
        else:
            if len(dct_filter) != 0:
                dct_dash_df[sheet],dct_vac_df[sheet] = create_dyn_vac_df(dash_temp_df, dash_base_df, dct_filter, dct_exclude_filter, result_date,dct_vac_df[sheet])


    else:
        if sheet == 'Зарплата по работодателям': # заменяем устаревшие названия
            dash_temp_df = dash_temp_df.groupby('Краткое название работодателя').agg({'Средняя ариф. минимальная зп':'mean','Медианная минимальная зп':'median'})
            dash_temp_df['Средняя ариф. минимальная зп'] = dash_temp_df['Средняя ариф. минимальная зп'].apply(lambda x: round(x, 0))
            dash_temp_df['Медианная минимальная зп'] = dash_temp_df['Медианная минимальная зп'].apply(lambda x: round(x, 0))
            dash_temp_df = dash_temp_df.reset_index()
            dash_temp_df = dash_temp_df.assign(Данные_на=result_date)
            # добавляем в базовый датафрейм
            base_dash_df = dct_dash_df[sheet]
            base_dash_df = pd.concat([base_dash_df,dash_temp_df])
            base_dash_df.fillna(0,inplace=True)
            dct_dash_df[sheet] = base_dash_df
        else:
            dash_temp_df[dash_special_treatment[sheet][0]] = dash_temp_df[dash_special_treatment[sheet][0]].replace(dct_value_rename[sheet])
            dash_temp_df = dash_temp_df.groupby(dash_special_treatment[sheet][0]).agg({dash_special_treatment[sheet][1]:'sum'})
            dash_temp_df = dash_temp_df.reset_index()
            dash_temp_df = dash_temp_df.assign(Данные_на=result_date)
            base_dash_df = dct_dash_df[sheet]
            base_dash_df = pd.concat([base_dash_df,dash_temp_df])
            base_dash_df.fillna(0,inplace=True)
            dct_dash_df[sheet] = base_dash_df


def create_dyn_vac_df(dash_temp_df:pd.DataFrame,dash_base_df:pd.DataFrame,dct_filter:dict,dct_exclude_filter:dict,result_date,df_vac:pd.DataFrame):
    """
    Функция для создания свода по динамике вакансий
    """

    temp_df = pd.DataFrame(columns=['Вакансия', 'Количество вакансий', 'Данные_на'])

    for key, lst_vac in dct_filter.items():
        dash_temp_df['Вакансия'] = dash_temp_df['Вакансия'].fillna('Не заполнено')
        temp_filter_df = dash_temp_df[dash_temp_df['Вакансия'].str.contains('|'.join(lst_vac), case=False,
                                                                            regex=True)]  # отбираем если содержит в себе список значений
        # Проводим дополнительную фильтрацию
        if len(dct_exclude_filter[key]) != 0:
            temp_filter_df = temp_filter_df[
                ~temp_filter_df['Вакансия'].str.contains('|'.join(dct_exclude_filter[key]), case=False, regex=True)]

        row_temp_filter_df = pd.DataFrame(columns=['Вакансия', 'Количество вакансий', 'Данные_на'],
                                          data=[[','.join(lst_vac), sum(temp_filter_df['Количество рабочих мест']),
                                                 result_date]])
        dash_base_df = pd.concat([dash_base_df, row_temp_filter_df])
        dash_base_df.fillna(0, inplace=True)


        temp_df = pd.concat([temp_df,row_temp_filter_df])
    # Обрабатываем для добавления
    temp_df.set_index('Вакансия',inplace=True)
    temp_df.drop(columns=['Данные_на'], inplace=True)
    temp_df.columns = [result_date]
    df_vac = df_vac.join(temp_df)

    return dash_base_df,df_vac









def processing_time_series(data_folder,end_folder,param_filter:str):
    """
    Функция для формирования временных рядов
    """
    # Словарь для аббревиатур
    dct_abbr = {'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ':'ООО','КРАЕВОЕ ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ УЧРЕЖДЕНИЕ':'КГАУ',
                'КРАЕВОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'КГБУ','ОТКРЫТОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО':'ОАО',
                'ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ':'ИП','МУНИЦИПАЛЬНОЕ УНИТАРНОЕ ПРЕДПРИЯТИЕ':'МУП',
                'ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ УЧРЕЖДЕНИЕ':'ГАУ','МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МАОУ',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'ФГБУ',
                'ОБЛАСТНОЕ ГОСУДАРСТВЕННОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'ОГКУ','УПРАВЛЕНИЕ ИМУЩЕСТВЕННЫХ ОТНОШЕНИЙ':'УИО',
                'МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ ДОПОЛНИТЕЛЬНОГО ОБРАЗОВАНИЯ':'МБУДО','ФЕДЕРАЛЬНОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'ФБУ',
                'ЗАКРЫТОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО':'ЗАО','МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МАДОУ',
                'МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МБОУ','ГЛАВНОЕ УПРАВЛЕНИЕ ФЕДЕРАЛЬНОЙ СЛУЖБЫ СУДЕБНЫХ ПРИСТАВОВ':'ГУ ФССП',
                'ПУБЛИЧНОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО':'ПАО',
                'ФИЛИАЛ АКЦИОНЕРНОГО ОБЩЕСТВА':'ФИЛИАЛ АО',
                'МУНИЦИПАЛЬНОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'МКУ','ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ УНИТАРНОЕ ПРЕДПРИЯТИЕ':'ФГУП',
                'МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МБДОУ',
                'МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ДОПОЛНИТЕЛЬНОГО ОБРАЗОВАНИЯ':'МАОУ ДО',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ФГБДОУ',
                'МУНИЦИПАЛЬНОЕ КАЗЁННОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МКОУ','СЕЛЬСКОХОЗЯЙСТВЕННЫЙ ПРОИЗВОДСТВЕННЫЙ КООПЕРАТИВ':'СПК',
                'ОБЛАСТНОЕ ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ УЧРЕЖДЕНИЕ':'ОГАУ','ГЛАВНОЕ УПРАВЛЕНИЕ':'ГУ','ОБЛАСТНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'ОГБУ',
                'МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'МБУ','ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'ФГКУ',
                'ГОСУДАРСТВЕННОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'ГКУ',
                'АВТОНОМНОЕ СТАЦИОНАРНОЕ УЧРЕЖДЕНИЕ':'АСУ','МУНИЦИПАЛЬНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МОУ',
                'МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МБОУ',
                'МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МАОУ',
                'ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ГБПОУ','БЮДЖЕТНОЕ ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'БДОУ',
                'ФЕДЕРАЛЬНОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'ФКУ','ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ВЫСШЕГО ОБРАЗОВАНИЯ':'ФГАОУ ВО',
                'ОБЛАСТНОЕ ГОСУДАРСТВЕННОЕ КАЗЁННОЕ УЧРЕЖДЕНИЕ':'ОГКУ','АВТОНОМНАЯ НЕКОММЕРЧЕСКАЯ ПРОФЕССИОНАЛЬНАЯ ОБРАЗОВАТЕЛЬНАЯ ОРГАНИЗАЦИЯ':'АНПОО',
                'ПРОИЗВОДСТВЕННЫЙ КООПЕРАТИВ':'ПК','ГОСУДАРСТВЕННОЕ УЧРЕЖДЕНИЕ ЗДРАВООХРАНЕНИЯ':'ГУЗ',
                'ФЕДЕРАЛЬНОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ ЗДРАВООХРАНЕНИЯ':'ФБУЗ',
                'БЮДЖЕТНОЕ ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'БПОУ','ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ГАПОУ',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ВЫСШЕГО ОБРАЗОВАНИЯ':'ФГБОУ ВО','МУНИЦИПАЛЬНОЕ КАЗЕННОЕ ПРЕДПРИЯТИ':'МКУ',
                'МУНИЦИПАЛЬНОЕ КАЗЁННОЕ УЧРЕЖДЕНИЕ':'МКУ',
                'ФИЛИАЛ ФЕДЕРАЛЬНОГО ГОСУДАРСТВЕННОГО БЮДЖЕТНОГО ОБРАЗОВАТЕЛЬНОГО УЧРЕЖДЕНИЯ ВЫСШЕГО ОБРАЗОВАНИЯ':'ФИЛИАЛ ФГБОУ ВО',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ ПРЕДПРИЯТИЕ':'ФГП',
                'АКЦИОНЕРНОЕ ОБЩЕСТВО': 'АО',
                'ГОСУДАРСТВЕННОЕ СПЕЦИАЛЬНОЕ УЧЕБНО-ВОСПИТАТЕЛЬНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ГСУВ ОО',
                'АВТОНОМНОЕ УЧРЕЖДЕНИЕ РЕСПУБЛИКИ':'АУР','Войсковая часть':'ВЧ',
                'ФИЛИАЛ ФЕДЕРАЛЬНОГО ГОСУДАРСТВЕННОГО БЮДЖЕТНОГО УЧРЕЖДЕНИЯ':'ФИЛИАЛ ФГБУ','МУНИЦИПАЛЬНОЕ УЧРЕЖДЕНИЕ':'МУ',
                'МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ':'МО','РЕСПУБЛИКАНСКОЕ ГОСУДАРСТВЕННОЕ УЧРЕЖДЕНИЕ':'РГУ',
                'МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ УЧРЕЖДЕНИЕ':'МАУ',
                'АВТОНОМНОЕ УЧРЕЖДЕНИЕ СОЦИАЛЬНОГО ОБСЛУЖИВАНИЯ':'АУ СО',
                'АВТОНОМНОЕ УЧРЕЖДЕНИЕ КУЛЬТУРЫ': 'АУ КУЛЬТУРЫ',
                'ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ ЧАСТНОЕ УЧРЕЖДЕНИЕ':'ПОЧУ',
                'МУНИЦИПАЛЬНОЕ ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МДОУ','ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ДОУ',
                'ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ГБОУ','ЧАСТНОЕ УЧРЕЖДЕНИЕ ЗДРАВООХРАНЕНИЯ':'ЧУЗ',
                'МУНИЦИПАЛЬНОГО РАЙОНА':'МР','ТЕРРИТОРИАЛЬНОЕ УПРАВЛЕНИЕ':'ТУ',
                'ФИЛИАЛ ПУБЛИЧНОГО АКЦИОНЕРНОГО ОБЩЕСТВА':'ФИЛИАЛ ПАО','ВОЙСКОВАЯ ЧАСТЬ':'ВЧ',
                'ФИЛИАЛ ФЕДЕРАЛЬНОГО ГОСУДАРСТВЕННОГО АВТОНОМНОГО УЧРЕЖДЕНИЯ':'ФИЛИАЛ ФГАУ',
                'ФИЛИАЛ ФЕДЕРАЛЬНОГО ГОСУДАРСТВЕННОГО УНИТАРНОГО ПРЕДПРИЯТИЯ':'ФИЛИАЛ ФГУП',
                'СРЕДНЯЯ ОБЩЕОБРАЗОВАТЕЛЬНАЯ ШКОЛА':'СОШ',
                'ФИЛИАЛ ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'ФИЛИАЛ ФГБУ',
                'ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ СОЦИАЛЬНОГО ОБСЛУЖИВАНИЯ':'ГБУ СО',
                'МИНИСТЕРСТВА ВНУТРЕННИХ ДЕЛ РОССИЙСКОЙ ФЕДЕРАЦИИ':'МВД РФ',
                'АВТОНОМНОЕ УЧРЕЖДЕНИЕ': 'АУ',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ': 'ФГА', 'ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ': 'ГБУ',
                'БЮДЖЕТНОЕ УЧРЕЖДЕНИЕ':'БУ','ДОПОЛНИТЕЛЬНОГО ОБРАЗОВАНИЯ':'ДО',
                'АВТОНОМНАЯ НЕКОММЕРЧЕСКАЯ ОБЩЕОБРАЗОВАТЕЛЬНАЯ ОРГАНИЗАЦИЯ':'АНОО',
                'АВТОНОМНАЯ НЕКОММЕРЧЕСКАЯ ОРГАНИЗАЦИЯ':'АНО',
                'ГОСУДАРСТВЕННОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ':'ГОКУ',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ КАЗЕННОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ФГКОУ',
                'ВЫСШЕГО ОБРАЗОВАНИЯ':'ВО',
                'ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ПОУ',
                'ЧАСТНОЕ ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ЧПОУ',
                'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ НАУЧНОЕ УЧРЕЖДЕНИЕ':'ФГБНУ',
                'МУНИЦИПАЛЬНОЕ КАЗЁННОЕ':'МК',
                'СЕЛЬСКОЕ ПОТРЕБИТЕЛЬСКОЕ ОБЩЕСТВО':'СПО',
                'ЧАСТНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ЧОУ',
                'МУНИЦИПАЛЬНОЕ КАЗЕННОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'МКОУ',
                'ОСНОВНАЯ ОБЩЕОБРАЗОВАТЕЛЬНАЯ ШКОЛА':'ООШ',
                'ДЕТСКАЯ ШКОЛА ИСКУССТВ':'ДШИ',
                'ДЕТСКИЙ ОЗДОРОВИТЕЛЬНО-ОБРАЗОВАТЕЛЬНЫЙ ЦЕНТР':'ДООЦ',
                'ЦЕНТР ЗАНЯТОСТИ НАСЕЛЕНИЯ':'ЦЗН',
                'ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОЗДОРОВИТЕЛЬНОЕ ОБЩЕОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ':'ГБООУ',
                'МУНИЦИПАЛЬНАЯ БЮДЖЕТНАЯ ДОШКОЛЬНАЯ ОБРАЗОВАТЕЛЬНАЯ ОРГАНИЗАЦИЯ':'МБДОУ',
                'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕСТВЕННОСТЬЮ':'ООО',
                }

    t = time.localtime()  # получаем текущее время и дату
    current_time = time.strftime('%H_%M_%S', t)
    current_date = time.strftime('%d_%m_%Y', t)
    # Обязательные листы
    error_df = pd.DataFrame(
        columns=['Название файла', 'Описание ошибки'])  # датафрейм для ошибок

    dct_filter_vac = dict() # словарь для хранения подготовленных списков с вакансиями которые нужно искать
    dct_exclude_filter_vac = dict() # словарь для хранения списков со значениями которые нужно отбросить в уже отфильтрованных данных
    lst_for_index_vac_df = [] # список для хранения наименований вакансий для последующего соединения

    # Проверяем заполнение файла со списком вакансий динамику по которым нужно получить
    try:
        if param_filter != '' and param_filter != 'Не выбрано':
            df_param_filter = pd.read_excel(param_filter,dtype=str,usecols='A:B')
            if len(df_param_filter) != 0:
                df_param_filter = df_param_filter.replace(r'^\s*$', pd.NA, regex=True).dropna(subset=df_param_filter.columns[0])
                # Создаем словарь по строкам с указанием вакансий которые есть в этой строке
                for idx,row in enumerate(df_param_filter.itertuples(),1):
                    # создаем списки вакансий которые будут искаться
                    lst_temp = row[1].split(',')
                    lst_temp = [value.strip().lower() for value in lst_temp if value]
                    dct_filter_vac[idx] = lst_temp

                    lst_for_index_vac_df.append(','.join(lst_temp))
                    # создаем списки для дополнительной фильтрации
                    if isinstance(row[2],str):
                        lst_dop_temp = row[2].split(',')
                        lst_dop_temp = [value.strip().lower() for value in lst_dop_temp if value]
                        dct_exclude_filter_vac[idx] = lst_dop_temp
                    else:
                        dct_exclude_filter_vac[idx] = []


        lst_files = []  # список для файлов
        for dirpath, dirnames, filenames in os.walk(data_folder):
            lst_files.extend(filenames)
        # отбираем файлы
        lst_xlsx = [file for file in lst_files if not file.startswith('~$') and file.endswith('.xlsx')]
        quantity_files = len(lst_xlsx)  # считаем сколько xlsx файлов в папке
        # Обрабатываем в зависимости от количества файлов в папке
        if quantity_files == 0:
            raise NotFile
        else:
            required_columns = {'Вакансии по отраслям':['Сфера деятельности','Количество вакансий'],
                                'Вакансии по муниципалитетам':['Муниципалитет','Количество вакансий'],
                                'Вакансии для динамики':['Краткое название работодателя','Вакансия','Количество рабочих мест','ID вакансии','Ссылка на вакансию'],
                                'Вакансии по работодателям':['Краткое название работодателя','Количество вакансий'],
                                'Образование по отраслям':['Сфера деятельности','Образование','Количество вакансий'],
                                'График работы по отраслям':['Сфера деятельности','График работы','Количество вакансий'],
                                'Тип занятости по отраслям':['Сфера деятельности','Тип занятости','Количество вакансий'],
                                'Квоты по отраслям':['Сфера деятельности','Количество вакансий'],
                                'Квоты по работодателям':['Краткое название работодателя','Количество вакансий'],
                                'Требуемый опыт по отраслям':['Сфера деятельности','Требуемый опыт работы в годах','Количество вакансий'],}

            drop_columns = {'Квоты по отраслям':['Квотируемое место'],'Квоты по работодателям':['Квотируемое место']} # для простых листов где надо удалить лишние колонки


            # словарь для листов где нужно обрабатывать группировать не первую колонку колонку
            second_cols_sheets = {'Образование по отраслям':['Образование','Количество вакансий'],
                                  'График работы по отраслям':['График работы','Количество вакансий'],
                                  'Тип занятости по отраслям':['Тип занятости','Количество вакансий'],
                                  'Требуемый опыт по отраслям':['Требуемый опыт работы в годах','Количество вакансий']}


            special_treatment = {
                                 'Образование по отраслям':['Образование','Количество вакансий'],
                                 'График работы по отраслям':['График работы','Количество вакансий'],
                                 'Тип занятости по отраслям':['Тип занятости','Количество вакансий'],
                                 'Требуемый опыт по отраслям': ['Требуемый опыт работы в годах', 'Количество вакансий']
                                 } # листы которые нужно обработать по особому

            dash_special_treatment = {
                                 'Образование по отраслям':['Образование','Количество вакансий'],
                                 'График работы по отраслям':['График работы','Количество вакансий'],
                                 'Тип занятости по отраслям':['Тип занятости','Количество вакансий'],
                                 'Требуемый опыт по отраслям': ['Требуемый опыт работы в годах', 'Количество вакансий']
                                 } # листы которые нужно обработать по особому



            # dupl_special_treatment = {'Зарплата по работодателям':{'Средняя ариф. минимальная зп':'Средняя ариф. минимальная зп Раб','Медианная минимальная зп':'Медианная минимальная зп Раб'}}
            dupl_special_treatment = {}

            # Словарь для листов из двух колонок где первая это индекс по которым нужно провести группировку
            # dct_one_group_sheet = {'Зарплата по работодателям':{'Средняя ариф. минимальная зп':'mean','Медианная минимальная зп':'median'
            #                                                     }}
            dct_one_group_sheet = {}
            # словарь для замены устаревших категорий
            dct_value_rename = {'Образование по отраслям':{'Высшее':'Высшее образование','Высшее-бакалавриат':'Высшее образование — бакалавриат',
                                                           'Высшее-подготовка кадров высшей квалификации':'Высшее образование — подготовка кадров высшей квалификации',
                                                           'Высшее-специалитет, магистратура':'Высшее образование — специалитет, магистратура',
                                                           'Среднее общее':'Среднее общее образование','Среднее профессиональное':'Среднее профессиональное образование',
                                                          },
                                'График работы по отраслям':{'Ненормированный рабочий день':'Ненормированный рабочий день'},
                                'Тип занятости по отраслям': {'Временная': 'Временная'},
                                'Требуемый опыт по отраслям':{0:0}
                                }
            dct_rename = {'Вакансии по отраслям':'Вакансии по отраслям',
                          'Выбранные вакансии': 'Выбранные вакансии',
                          'Вакансии по муниципалитетам':'Вакансии по муниципалитетам',
                          'Вакансии по работодателям':'Вакансии по работодателям',
                          'Вакансии для динамики':'Динамика по вакансиям',
                          'Образование по отраслям': 'Образование Вак',
                          'График работы по отраслям': 'График работы Вак',
                          'Тип занятости по отраслям': 'Тип занятости Вак',
                          'Требуемый опыт по отраслям': 'Опыт Вак',
                          'Образование':'Образование Вак',
                          'График работы':'График работы Вак',
                          'Тип занятости':'Тип занятости Вак',
                          'Квоты по отраслям':'Квоты по отраслям',
                          'Квоты по работодателям':'Квоты по работодателям',
                          'Требуемый опыт работы в годах':'Опыт Вак',
                          'Всего вакансий':'Всего вакансий'

                          } # словарь для переименования
            dct_index_svod = {key:set() for key in required_columns.keys()} # словарь для хранения всех значений сводов которые могут встретиться в файлах
            set_error_name_file = set() # множество для хранения названий файлов с ошибками

            dct_index_svod,error_df,set_error_name_file = preparing_data(data_folder,required_columns,dct_index_svod,error_df,set_error_name_file,second_cols_sheets,dct_value_rename,dct_abbr) # Проверяем на ошибки
            # Создаем словарь с базовыми датафреймами
            dct_base_df = dict()
            # Создаем словарь для хранения датафреймов для сводов
            dct_dash_df = {'Всего вакансий':pd.DataFrame(columns=['Количество вакансий','Данные_на']),
                           'Вакансии по отраслям':pd.DataFrame(columns=['Сфера деятельности','Количество вакансий','Данные_на']),
                           'Вакансии по муниципалитетам':pd.DataFrame(columns=['Муниципалитет','Количество вакансий','Данные_на']),
                           'Вакансии для динамики':pd.DataFrame(columns=['Вакансия','Количество вакансий','Данные_на']),
                           'Вакансии по работодателям':pd.DataFrame(columns=['Краткое название работодателя','Количество вакансий','Данные_на']),
                           'Образование по отраслям':pd.DataFrame(columns=['Образование','Количество вакансий','Данные_на']),
                           'График работы по отраслям':pd.DataFrame(columns=['График работы','Количество вакансий','Данные_на']),
                           'Тип занятости по отраслям':pd.DataFrame(columns=['Тип занятости','Количество вакансий','Данные_на']),
                           'Квоты по отраслям':pd.DataFrame(columns=['Сфера деятельности','Количество вакансий','Данные_на']),
                           'Квоты по работодателям':pd.DataFrame(columns=['Краткое название работодателя','Количество вакансий','Данные_на']),
                           'Требуемый опыт по отраслям':pd.DataFrame(columns=['Требуемый опыт работы в годах','Количество вакансий','Данные_на']),}

            # Датафрейм для сбора данных по выбранным вакансиям
            dct_vac_df = {'Вакансии для динамики':pd.DataFrame(index=sorted([value for value in lst_for_index_vac_df]))}



            # Создаем ключи
            for name_sheet,set_index in dct_index_svod.items():
                dct_base_df[name_sheet] = pd.DataFrame(index=sorted([value for value in set_index if value != 'Итого']))

            # добавляем ключ для подсчета общего количества вакансий
            dct_base_df['Всего вакансий'] = pd.DataFrame(index=['Вакансий по региону'])
            # добавляем ключ для подсчета вакансий
            for dirpath, dirnames, filenames in os.walk(data_folder):
                for file in filenames:
                    if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm')):
                        if file.endswith('.xlsx'):
                            name_file = file.split('.xlsx')[0].strip()
                        else:
                            name_file = file.split('.xlsm')[0].strip()
                        if name_file not in set_error_name_file:
                            print(name_file)  # обрабатываемый файл
                            # ха повторяющийся код ну и ладно
                            result_date = re.search(r'\d{2}_\d{2}_\d{4}', name_file)
                            result_date = result_date.group().replace('_','.')

                            for sheet, lst_cols in required_columns.items():
                                temp_wb = openpyxl.load_workbook(f'{dirpath}/{file}', read_only=True)
                                lst_sheets = temp_wb.sheetnames
                                temp_wb.close()
                                if sheet not in lst_sheets:
                                    continue
                                temp_req_df = pd.read_excel(f'{dirpath}/{file}', sheet_name=sheet)

                                if 'Полное название работодателя' in temp_req_df.columns:
                                    temp_req_df['Полное название работодателя'] = temp_req_df[
                                        'Полное название работодателя'].apply(
                                        lambda x: x.upper() if isinstance(x, str) else x).replace(dct_abbr, regex=True)
                                    temp_req_df = temp_req_df.rename(
                                        columns={'Полное название работодателя': 'Краткое название работодателя'})
                                if len(temp_req_df) == 0:
                                    continue
                                diff_cols = set(required_columns[sheet]).difference(set(temp_req_df.columns))
                                if len(diff_cols) !=0 :
                                    continue

                                if 'Краткое название работодателя' in temp_req_df.columns:
                                    temp_req_df['Краткое название работодателя'] = temp_req_df['Краткое название работодателя'].apply(
                                        lambda x: x.upper() if isinstance(x, str) else x).replace(dct_abbr, regex=True)
                                # Заполняем словарь для дашборда
                                dash_temp_df = temp_req_df.copy() # создаем копию

                                create_dash_df(dct_dash_df,dash_temp_df,sheet,result_date,dash_special_treatment,dct_value_rename,dct_filter_vac,dct_exclude_filter_vac,dct_vac_df)

                                # Делаем первую колонку индексом
                                temp_req_df.set_index(temp_req_df.columns[0],inplace=True)



                                if sheet not in special_treatment:
                                    if sheet == 'Вакансии для динамики':
                                        continue
                                    if sheet not in drop_columns:
                                        temp_req_df.columns = [result_date]
                                    else:
                                        # удаляем лишние колонки в простых листах
                                        temp_req_df.drop(columns=drop_columns[sheet],inplace=True,errors='ignore')
                                        temp_req_df.columns = [result_date]

                                    base_df = dct_base_df[sheet] # получаем базовый датафрейм
                                    base_df= base_df.join(temp_req_df)
                                    base_df.fillna(0,inplace=True)
                                    dct_base_df[sheet] = base_df


                                    if sheet == 'Вакансии по отраслям':
                                        # заполняем лист Всего вакансий
                                        prom_df = temp_req_df[temp_req_df.index != 'Итого']
                                        itog_vac = prom_df[result_date].sum()
                                        temp_itog_df = pd.DataFrame(columns=[result_date],data=[itog_vac],index=['Вакансий по региону'])
                                        itog_base_df = dct_base_df['Всего вакансий']
                                        itog_base_df = itog_base_df.join(temp_itog_df)
                                        itog_base_df.fillna(0, inplace=True)
                                        dct_base_df['Всего вакансий'] = itog_base_df


                                else:
                                    if sheet not in second_cols_sheets:
                                        # Создаем отдельные датафреймы
                                        for name_column in special_treatment[sheet]:
                                            if sheet not in dupl_special_treatment:
                                                # перебираем список и проверяем есть уже такой базовый датафрейм, если нет то создаем
                                                if name_column not in dct_base_df:
                                                    temp_treatement_df = temp_req_df[[name_column]].copy()
                                                    if sheet not in dct_one_group_sheet:
                                                        temp_treatement_df.columns = [result_date]
                                                    else:
                                                        # группируем
                                                        temp_treatement_df = temp_treatement_df.groupby(level=0).agg(dct_one_group_sheet[sheet][name_column])
                                                        temp_treatement_df.columns = [result_date]
                                                        if dct_one_group_sheet[sheet][name_column] in ('mean', 'median'):
                                                            # округляем если функция средняя или медиана
                                                            temp_treatement_df[result_date] = temp_treatement_df[result_date].apply(lambda x: round(x, 0))
                                                    dct_base_df[name_column] = temp_treatement_df
                                                else:
                                                    base_treatment_df = dct_base_df[name_column]  # получаем базовый датафрейм
                                                    temp_treatement_df = temp_req_df[[name_column]].copy()
                                                    if sheet not in dct_one_group_sheet:
                                                        temp_treatement_df.columns = [result_date]
                                                    else:
                                                        # группируем
                                                        temp_treatement_df = temp_treatement_df.groupby(level=0).agg(dct_one_group_sheet[sheet][name_column])
                                                        temp_treatement_df.columns = [result_date]
                                                        if dct_one_group_sheet[sheet][name_column] in ('mean', 'median'):
                                                            # округляем если функция средняя или медиана
                                                            temp_treatement_df[result_date] = temp_treatement_df[result_date].apply(lambda x: round(x, 0))
                                                    base_treatment_df = base_treatment_df.join(temp_treatement_df)
                                                    base_treatment_df.fillna(0, inplace=True)
                                                    dct_base_df[name_column] = base_treatment_df

                                            else:
                                                if dupl_special_treatment[sheet][name_column] not in dct_base_df:
                                                    temp_treatement_df = temp_req_df[[name_column]].copy()
                                                    if sheet not in dct_one_group_sheet:
                                                        temp_treatement_df.columns = [result_date]
                                                    else:
                                                        # группируем
                                                        temp_treatement_df = temp_treatement_df.groupby(level=0).agg(dct_one_group_sheet[sheet][name_column])
                                                        temp_treatement_df.columns = [result_date]
                                                        if dct_one_group_sheet[sheet][name_column] in ('mean','median'):
                                                            # округляем если функция средняя или медиана
                                                            temp_treatement_df[result_date] = temp_treatement_df[result_date].apply(lambda x:round(x,0))
                                                    dct_base_df[dupl_special_treatment[sheet][name_column]] = temp_treatement_df
                                                else:
                                                    base_treatment_df = dct_base_df[dupl_special_treatment[sheet][name_column]]  # получаем базовый датафрейм
                                                    temp_treatement_df = temp_req_df[[name_column]].copy()
                                                    if sheet not in dct_one_group_sheet:
                                                        temp_treatement_df.columns = [result_date]
                                                    else:
                                                        # группируем
                                                        temp_treatement_df = temp_treatement_df.groupby(level=0).agg(dct_one_group_sheet[sheet][name_column])
                                                        temp_treatement_df.columns = [result_date]
                                                        if dct_one_group_sheet[sheet][name_column] in ('mean','median'):
                                                            # округляем если функция средняя или медиана
                                                            temp_treatement_df[result_date] = temp_treatement_df[result_date].apply(lambda x:round(x,0))
                                                    base_treatment_df = base_treatment_df.join(temp_treatement_df)
                                                    base_treatment_df.fillna(0, inplace=True)
                                                    dct_base_df[dupl_special_treatment[sheet][name_column]] = base_treatment_df
                                    else:
                                        if second_cols_sheets[sheet][0] not in dct_base_df:
                                            temp_base_df = dct_base_df[sheet].copy() # делаем копию
                                            temp_req_df = temp_req_df.reset_index() # вытаскиваем первую колонку из индекса
                                            temp_req_df.drop(columns=temp_req_df.columns[0],inplace=True) # удаляем колонку бывшую индексом
                                            # заменяем устаревшие названия категорий
                                            temp_req_df[second_cols_sheets[sheet][0]] = temp_req_df[second_cols_sheets[sheet][0]].replace(dct_value_rename[sheet])
                                            temp_req_df.set_index(second_cols_sheets[sheet][0],inplace=True)
                                            temp_treatement_df = temp_req_df.groupby(level=0).agg('sum')
                                            temp_treatement_df.columns = [result_date]
                                            temp_base_df = temp_base_df.join(temp_treatement_df)
                                            dct_base_df[second_cols_sheets[sheet][0]] = temp_base_df
                                        else:
                                            base_treatment_df = dct_base_df[second_cols_sheets[sheet][0]]  # получаем базовый датафрейм
                                            temp_req_df = temp_req_df.reset_index()  # вытаскиваем первую колонку из индекса
                                            temp_req_df.drop(columns=temp_req_df.columns[0],
                                                             inplace=True)  # удаляем колонку бывшую индексом
                                            # заменяем устаревшие названия категорий
                                            temp_req_df[second_cols_sheets[sheet][0]] = temp_req_df[second_cols_sheets[sheet][0]].replace(dct_value_rename[sheet])
                                            temp_req_df.set_index(second_cols_sheets[sheet][0], inplace=True)
                                            temp_treatement_df = temp_req_df.groupby(level=0).agg('sum')
                                            temp_treatement_df.columns = [result_date]
                                            base_treatment_df = base_treatment_df.join(temp_treatement_df)
                                            base_treatment_df.fillna(0, inplace=True)
                                            dct_base_df[second_cols_sheets[sheet][0]] = base_treatment_df

            # Сохраняем в горизонтальном виде
            # переносим лист Всего вакансий в начало
            dct_base_df['Выбранные вакансии'] = dct_vac_df['Вакансии для динамики']
            new_order = ['Всего вакансий','Выбранные вакансии','Вакансии по отраслям','Вакансии по муниципалитетам',
                         'Вакансии по работодателям','Образование',
                         'График работы','Тип занятости','Требуемый опыт работы в годах',
                         'Квоты по отраслям','Квоты по работодателям',
                         'Образование по отраслям','График работы по отраслям',
                         'Тип занятости по отраслям','Требуемый опыт по отраслям']
            dct_base_df = {key: dct_base_df[key] for key in new_order}

            with pd.ExcelWriter(f'{end_folder}/Горизонтальный вид {current_time}.xlsx',engine='xlsxwriter') as writer:
                for sheet_name, df in dct_base_df.items():
                    if sheet_name in special_treatment:
                        continue
                    # Преобразуем и сортируем колонки-даты
                    date_cols = []

                    for col in df.columns:
                        date_obj = pd.to_datetime(col, format='%d.%m.%Y')
                        date_cols.append((date_obj, col))

                    # Сортируем даты
                    date_cols.sort(key=lambda x: x[0])
                    # Создаем новые названия колонок
                    new_columns =[date for _, date in date_cols]

                    # Переупорядочиваем DataFrame
                    df = df[new_columns]

                    # Преобразуем названия колонок-дат
                    for date_obj, old_name in date_cols:
                        df = df.rename(columns={old_name: date_obj})
                    if len(df) != 0:
                        df.columns = df.columns.strftime('%d.%m.%Y')
                    df.to_excel(writer,sheet_name=dct_rename[sheet_name],index=True)
                    gc.collect()

            # Вертикальный вид
            with pd.ExcelWriter(f'{end_folder}/Вертикальный вид {current_time}.xlsx',engine='xlsxwriter') as writer:
                for sheet_name, df in dct_base_df.items():
                    if sheet_name in special_treatment:
                        continue

                    # Преобразуем и сортируем колонки-даты
                    date_cols = []

                    for col in df.columns:
                        date_obj = pd.to_datetime(col, format='%d.%m.%Y')
                        date_cols.append((date_obj, col))

                    # Сортируем даты
                    date_cols.sort(key=lambda x: x[0])
                    # Создаем новые названия колонок
                    new_columns =[date for _, date in date_cols]

                    # Переупорядочиваем DataFrame
                    df = df[new_columns]

                    # Преобразуем названия колонок-дат
                    for date_obj, old_name in date_cols:
                        df = df.rename(columns={old_name: date_obj})
                    if len(df) != 0:
                        df.columns = df.columns.strftime('%d.%m.%Y')
                        if len(df) < 10000: # транспонируем только если не больше уровня
                            df = df.transpose()
                    df.to_excel(writer,sheet_name=dct_rename[sheet_name],index=True)
                    gc.collect()

            del dct_base_df
            # Формат для дашборда
            with pd.ExcelWriter(f'{end_folder}/Для сводов {current_time}.xlsx',engine='xlsxwriter') as writer:
                for sheet_name, df in dct_dash_df.items():

                    df.to_excel(writer, sheet_name=dct_rename[sheet_name], index=False)
            # Сохраняем по отдельности
            if not os.path.exists(f'{end_folder}/Отдельные своды'):
                os.makedirs(f'{end_folder}/Отдельные своды')

            for sheet_name, df in dct_dash_df.items():
                df.to_excel(f'{end_folder}/Отдельные своды/{dct_rename[sheet_name]}.xlsx',index=False)




            error_df.to_excel(f'{end_folder}/Ошибки_{current_time}.xlsx',index=False)

    except NotFile:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 f'В выбранной папке отсутствуют файлы Excel с расширением xlsx')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except OSError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Укажите в качестве конечной папки, папку в корне диска с коротким названием. Проблема может быть\n '
                             f'в слишком длинном пути к создаваемому файлу')

    else:
        messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                            'Данные успешно обработаны.Ошибок не обнаружено')


















if __name__ == '__main__':
    main_data_folder = 'data/Своды'
    # main_data_folder = 'data/СВОД Бурятия'
    main_end_folder = 'data/РЕЗУЛЬТАТ'
    main_filter_file = 'data/Свод для динамики вакансий.xlsx'
    # main_filter_file = 'Не выбрано'
    start_time = time.time()
    processing_time_series(main_data_folder,main_end_folder,main_filter_file)
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Время выполнения: {execution_time} секунд")
    print('Lindy Booth')








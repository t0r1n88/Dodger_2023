# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time

pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random


# Классы для исключений
class BadHeader(Exception):
    """
    Класс для проверки правильности заголовка
    """
    pass


class CheckBoxException(Exception):
    """
    Класс для вызовы исключения в случае если неправильно выставлены чекбоксы
    """
    pass


class NotFoundValue(Exception):
    """
    Класс для обозначения того что значение не найдено
    """
    pass


class ShapeDiffierence(Exception):
    """
    Класс для обозначения несовпадения размеров таблицы
    """
    pass


class ColumnsDifference(Exception):
    """
    Класс для обозначения того что названия колонок не совпадают
    """
    pass


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_folder_data():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data
    path_folder_data = filedialog.askdirectory()


def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()


# для обработки отчетов ЦК
def select_folder_data_ck():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data_ck
    path_folder_data_ck = filedialog.askdirectory()


def select_end_folder_ck():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder_ck
    path_to_end_folder_ck = filedialog.askdirectory()


def select_folder_data_opk():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data_opk
    path_folder_data_opk = filedialog.askdirectory()


def select_end_folder_opk():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder_opk
    path_to_end_folder_opk = filedialog.askdirectory()


def select_files_data_xlsx():
    """
    Функция для выбора нескоьких файлов с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global files_data_xlsx
    # Получаем путь файлы
    files_data_xlsx = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def check_data(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0


def check_data_note(cell):
    if cell is np.nan:
        return 'Не заполнено'
    return str(cell)


def check_first_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки гр. 09 и гр. 10 < гр. 08
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]

    # Проводим проверку
    df['Результат'] = (df['08'] >= df['09']) & (df['08'] >= df['10'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 09 <= гр. 08  или гр. 10 <= гр. 08'
    return temp_error_df


def check_second_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 07= гр.08 + сумма(с гр.11 по гр.32))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # конвертируем в инт
    all_sum_cols = list(df)
    # удаляем колонки 07, 09, 10
    all_sum_cols.remove('07')
    all_sum_cols.remove('09')
    all_sum_cols.remove('10')
    # получаем сумму колонок 08, 11:32
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['07'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 07 = гр.08 + сумма(с гр.11 по гр.32)'
    return temp_error_df


def check_third_error(df: pd.DataFrame, name_file, border, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр. 02 + стр. 04
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '06', ])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['Сумма'] = foo_df['02'] + foo_df['04']
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр. 02 + стр. 04 '
    return temp_error_df


def check_fourth_error(df: pd.DataFrame, name_file, border, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13

    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['06', '07', '08', '09', '10', '11', '12', '13'])

    # Добавляем данные в датафрейм
    foo_df['06'] = df.iloc[5, :]
    foo_df['07'] = df.iloc[6, :]
    foo_df['08'] = df.iloc[7, :]
    foo_df['09'] = df.iloc[8, :]
    foo_df['10'] = df.iloc[9, :]
    foo_df['11'] = df.iloc[10, :]
    foo_df['12'] = df.iloc[11, :]
    foo_df['13'] = df.iloc[12, :]

    sum_col = ['07', '08', '09', '10', '11', '12', '13']
    foo_df['Сумма'] = foo_df[sum_col].sum(axis=1)
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13 '

    return temp_error_df


def check_fifth_error(df: pd.DataFrame, name_file, border, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['05', '06', '14'])

    # Добавляем данные в датафрейм
    foo_df['05'] = df.iloc[4, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['14'] = df.iloc[13, :]

    foo_df['Результат'] = (foo_df['14'] <= foo_df['05']) & (foo_df['14'] <= foo_df['06'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df[
        'index'].tolist()  # делаем список, прибавляем для того чтобы номера строк совпадали с строками в файле
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк  {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 14<=стр. 06, стр. 14<=стр 05'
    return temp_error_df


def check_sixth_error(df: pd.DataFrame, name_file, border, tup_correct: tuple):
    """
    Функция для проверки правильности введеденных данных
    стр 03 <= стр 02 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '03'])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['03'] = df.iloc[2, :]
    foo_df['Результат'] = foo_df['03'] <= foo_df['02']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 03 <= стр. 02 '
    return temp_error_df


def check_seventh_error(df: pd.DataFrame, name_file, border, tup_correct: tuple):
    """
    Функция для проверки правильности введеденных данных
    стр.02 и стр.04 и стр.05 < стр.01
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '05', '01'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['05'] = df.iloc[4, :]

    foo_df['Результат'] = (foo_df['01'] >= foo_df['02']) & (foo_df['01'] >= foo_df['04']) & (
                foo_df['01'] >= foo_df['05'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр.02<= стр.01 или стр.04<= стр.01 или стр.05<= стр.01 '
    return temp_error_df


def check_error(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 6:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.03 <= стр.02
        sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct)
        error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
    # Возвращаем датафрейм с ошибками
    return error_df


def extract_code(value):
    """
    Функция для извлечения кода специальности
    """
    value = str(value)
    re_code = re.compile('^\d{2}?[.]\d{2}?[.]\d{2}$')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.group()
    else:
        return 'error'


def create_check_tables(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df[
                'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df[
                'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Неформальная занятость (нелегальная)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 20'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 22'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 25'))
            temp_code_df[
                'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 28'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 31'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
            temp_code_df['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 33'))

            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении',
                   'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
                ,
                   'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
                   'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
                   'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
                   'Строка 12': 'расстройствами аутистического спектра',
                   'Строка 13': 'с инвалидностью вследствие  других причин',
                   'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
                   'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    wb.save(f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')


def processing_data_employment():
    """
    Фугкция для обработки данных
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    try:
        for file in os.listdir(path_folder_data):
            if not file.startswith('~$') and file.endswith('.xls'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Файл с расширением XLS (СТАРЫЙ ФОРМАТ EXCEL)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=7, dtype=str)
                # проверяем корректность заголовка
                # создаем множество колонок наличие которых мы проверяем
                check_cols = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                              '16', '17',
                              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                              '33'}
                if not check_cols.issubset(set(df.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы в файле.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 8 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df = df[df['05'] != '16']  # фильтруем строки с проверками
                # отсекаем возможный первый столбец с данными ПОО,начинаем датафрейм с колонки 01 и отсекаем колонки с примечаниями
                df = df.loc[:, '01':'33']

                # # получаем  часть с данными
                mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # проверяем есть ли строка полностью состоящая из nan
                empty_row_index = np.where(df.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df = df.iloc[:row_index]
                #     # Проверка на размер таблицы, должно бьть кратно 15
                count_spec = df.shape[0] // 15  # количество специальностей
                df = df.iloc[:count_spec * 15, :]  # отбрасываем строки проверки
                check_code_lst = df['03'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue
                # Проверка на непрерывность кода специальности, то есть на 15 строк должен быть только один код
                border_check_code = 0  # счетчик обработанных страниц
                quantity_check_code = len(check_code_lst) // 15  # получаем сколько специальностей в таблице
                flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 15 строк
                flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
                for i in range(quantity_check_code):
                    temp_set = set(
                        [code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 15]])
                    if len(temp_set) != 1:
                        flag_error_code_spec = True
                    if ' ' in temp_set:
                        flag_error_space_spec = True
                    border_check_code += 15

                if flag_error_space_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if flag_error_code_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД СПЕЦИАЛЬНОСТИ НА КАЖДЫЕ 15 СТРОК (не считая строки с проверкой) !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df.columns = list(map(str, df.columns))
                # Заполняем пока пропуски в 15 ячейке для каждой специальности
                df['06'] = df['06'].fillna('15 ячейка')

                # Проводим проверку на корректность данных, отправляем копию датафрейма
                tup_correct = (9, 23)  # создаем кортеж  с поправками
                file_error_df = check_error(df.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df['03'] = df['03'].apply(extract_code)  # очищаем от текста в кодах
                # Проверяем на наличие слова error что означает что там есть некорректные значения кодов специальности
                if 'error' in df['03'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 03 Код специальности.Вместо кода присутствует дата, вместе с кодом есть название,пробел перед кодом и т.п.!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                code_spec = [spec for spec in df['03'].unique()]

                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 16)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(7, 34)]

                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}

                # Изменяем последний ключ на строковый поскольку там будут хранится примечания
                for row, value in spec_dict.items():
                    for col, data in value.items():
                        if col == 'Колонка 33':
                            spec_dict[row][col] = ''
                # Создаем словарь среднего уровня содержащй данные по всем специальностям
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

                high_level_dct[name_file] = copy.deepcopy(poo_dct)

                """
                В итоге получается такая структура
                {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                """

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк

                # Итерируемся по полученному датафрейму через itertuples
                for row in df.itertuples():
                    # если счетчик колонок больше 15 то уменьшаем его до единицы
                    if idx_row > 15:
                        idx_row = 1

                    # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[3] is not np.nan) and (row[3] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[3] != current_code:
                            current_code = row[3]

                    data_row = row[7:34]  # получаем срез с нужными данными

                    for idx_col, value in enumerate(data_row, start=1):
                        if idx_col + 6 == 33:
                            # сохраняем примечания в строке
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 6}'] = f'{name_file} {check_data_note(value)};'

                        else:
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 6}'] += check_data(value)

                    idx_row += 1

        create_check_tables(high_level_dct)

        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        if col == 'Колонка 33':
                            itog_df[code_spec][row][col] += check_data_note(value)
                        else:
                            itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()

        frame['Всего'] = frame[0].apply(lambda x: x.get('Колонка 7'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 8'))
        frame[
            'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
            frame[0].apply(lambda x: x.get('Колонка 9'))
        frame[
            'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
            frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Индивидуальные предприниматели'] = frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = frame[
            0].apply(lambda x: x.get('Колонка 12'))
        frame['Продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 13'))
        frame['Проходят службу в армии по призыву'] = frame[0].apply(lambda x: x.get('Колонка 14'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 15'))
        frame['Находятся в отпуске по уходу за ребенком'] = frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Неформальная занятость (нелегальная)'] = frame[0].apply(lambda x: x.get('Колонка 17'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
            frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Иные причины нахождения под риском нетрудоустройства'] = frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Смерть, тяжелое состояние здоровья'] = frame[0].apply(lambda x: x.get('Колонка 21'))
        frame['Находятся под следствием, отбывают наказание'] = frame[0].apply(lambda x: x.get('Колонка 22'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            frame[0].apply(lambda x: x.get('Колонка 23'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
            frame[0].apply(lambda x: x.get('Колонка 24'))
        frame['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = frame[0].apply(
            lambda x: x.get('Колонка 25'))
        frame[
            'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
            frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['будут трудоустроены'] = frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['будут осуществлять предпринимательскую деятельность'] = frame[0].apply(lambda x: x.get('Колонка 28'))
        frame['будут самозанятыми'] = frame[0].apply(lambda x: x.get('Колонка 29'))
        frame['будут призваны в армию'] = frame[0].apply(lambda x: x.get('Колонка 30'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 31'))
        frame['будут продолжать обучение'] = frame[0].apply(lambda x: x.get('Колонка 32'))
        frame['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = frame[
            0].apply(lambda x: x.get('Колонка 33'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении',
               'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
            ,
               'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
               'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
               'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
               'Строка 12': 'расстройствами аутистического спектра',
               'Строка 13': 'с инвалидностью вследствие  других причин',
               'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
               'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
               }
        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])
        # добавляем строки с проверкой
        count = 0
        for i in range(15, len(finish_df) + 1, 15):
            new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
            new_row.iloc[:, 1] = 'Проверка (строка не редактируется)'
            new_row.iloc[:, 2:] = 'проверка пройдена'

            # Вставка новой строки через каждые 15 строк
            finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(
                drop=True)
            count += 1
        lst_number_row = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                          '16']
        multipler = len(finish_df) // 16  # получаем количество специальностей/профессий
        # вставляем новую колонку
        finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))
        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        finish_df.to_excel(f'{path_to_end_folder}/Полная таблица  от {current_time}.xlsx', index=False)

        # Создаем файл с 5 строками
        small_finish_df = pd.DataFrame(columns=finish_df.columns)
        one_finish_df = pd.DataFrame(columns=finish_df.columns)

        lst_code_spec = finish_df['Код специальности'].unique()  # получаем список специальностей
        for code_spec in lst_code_spec:
            temp_df = finish_df[finish_df['Код специальности'] == code_spec]
            small_finish_df = pd.concat([small_finish_df, temp_df.iloc[:5, :]], axis=0, ignore_index=True)
            one_finish_df = pd.concat([one_finish_df, temp_df.iloc[:1, :]], axis=0, ignore_index=True)

        with pd.ExcelWriter(f'{path_to_end_folder}/5 строк Трудоустройство от {current_time}.xlsx') as writer:
            small_finish_df.to_excel(writer, sheet_name='5 строк', index=False)
            one_finish_df.to_excel(writer, sheet_name='1 строка (Всего выпускников)', index=False)

        # Создаем документ
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ от {current_time}.xlsx')

    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Закройте открытые файлы Excel {e.args}')
    except:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'При обработке файла {name_file} возникла ошибка !!!\n'
                             f'Проверьте файл на соответствие шаблону')

    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                'Данные успешно обработаны.')


"""
Функции для обработки отчетов ЦК
"""


def check_horizont_all_sum_error(df: pd.DataFrame, tup_exluded_cols: tuple, name_itog_cols, name_file):
    """
    Функция для проверки горизонтальных сумм по всей строке
    сумма в колонке 05 должна быть равна сумме всех колонок за исключением 07 и 15
    """
    # датафрейм для ощибок по горизонтали
    hor_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # получаем список колонок
    all_sum_cols = list(df)
    # удаляем колонки
    for name_cols in tup_exluded_cols:
        all_sum_cols.remove(name_cols)
    # удаляем итоговую колонку
    all_sum_cols.remove(name_itog_cols)

    # получаем сумму колонок за вычетом исключаемых и итоговой колонки
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} = сумма остальных гр. за искл.{tup_exluded_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_horizont_chosen_sum_error(df: pd.DataFrame, tup_checked_cols: list, name_itog_cols, name_file):
    """
    Функция для проверки равенства одиночных или небольших групп колонок
    tup_checked_cols колонки сумму которых нужно сравнить с name_itog_cols чтобы она не превышала это значение
    """
    # Считаем проверяемые колонки
    df['Сумма'] = df[tup_checked_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] >= df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} >= сумма {tup_checked_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_vertical_chosen_sum(df: pd.DataFrame, lst_checked_rows: list, itog_row, name_file):
    """
    Функция для проверки вертикальной суммы заданных строк сумма значений в tupl_checked_row должна быть равной ил меньше чем значение
    в itog_row
    """

    # обрабаотываем список строк чтобы привести его в читаемый вид
    lst_out_rows = list(map(lambda x: x + 1, lst_checked_rows))

    # делаем значения строковыми
    lst_out_rows = list(map(str, lst_out_rows))
    # Добавляем ноль в строки
    lst_out_rows = list(map(lambda x: '0' + x, lst_out_rows))
    # обрабатываем формат выходной строки
    out_itog_row = f'0{str(itog_row + 1)}'

    # создаем временный датафрейм
    foo_df = pd.DataFrame()
    # разворачиваем строки в колонки
    for idx_row in lst_checked_rows:
        foo_df[idx_row] = df.iloc[idx_row, :]

    # добавляем итоговую колонку
    foo_df[itog_row] = df.iloc[itog_row, :]

    # суммируем
    foo_df['Сумма'] = foo_df[lst_checked_rows].sum(axis=1)
    foo_df['Результат'] = foo_df[itog_row] >= foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    error_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    # Добавляем слово колонка
    error_df['index'] = error_df['index'].apply(lambda x: 'Колонка ' + str(x))
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    temp_error_df['Строка или колонка с ошибкой'] = error_df['index']
    temp_error_df[
        'Описание ошибки'] = f'Для указанной колонки сумма в строках {lst_out_rows} превышает значением в строке {out_itog_row} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! '
    temp_error_df['Название файла'] = name_file

    return temp_error_df


def check_error_ck(df: pd.DataFrame, name_file):
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # проводим горизонтальные проверки
    # проверка на общую сумму
    first_error_ck_df = check_horizont_all_sum_error(df.copy(), ('07', '15'), '05', name_file)
    error_df = pd.concat([error_df, first_error_ck_df], axis=0, ignore_index=True)
    # проверяем небольшие группы или одиночные колонки
    second_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['07'], '06', name_file)
    error_df = pd.concat([error_df, second_error_ck_df], axis=0, ignore_index=True)

    # проверяем колонки 14 и 15
    third_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['15'], '14', name_file)
    error_df = pd.concat([error_df, third_error_ck_df], axis=0, ignore_index=True)

    # Проводим вертикальные проверки
    # Сумма овз и целевиков не должна превышать общую численность выпускников. Строки с индексом 1 и 4 должныть меньше или равны строке с индексом 0
    fourth_error_ck_df = check_vertical_chosen_sum(df.copy(), [1, 4], 0, name_file)
    error_df = pd.concat([error_df, fourth_error_ck_df], axis=0, ignore_index=True)

    # Проверяем ОВЗ
    fifth_error_ck_df = check_vertical_chosen_sum(df.copy(), [2, 3], 1, name_file)
    error_df = pd.concat([error_df, fifth_error_ck_df], axis=0, ignore_index=True)

    return error_df


def processing_data_ck_employment():
    """
    Функция для обработки отчетов центров карьеры
    :return:
    """
    # создаем базовый датафрейм заполненный нулями
    base_df = pd.DataFrame(np.zeros((5, 27)))
    base_df = base_df.applymap(int)  # приводим его к инту
    cols_df = ['05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21',
               '22', '23', '24',
               '25', '26', '27', '28', '29', '30', '31']
    base_df.columns = cols_df

    # Создаем общую таблицы для проверки
    general_table = pd.DataFrame(columns=['Название файла'] + cols_df)

    # создаем датафрейм для регистрации ошибок
    base_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    # Создаем датафрейм для хранения строковых данных из колонки 32
    str_df = pd.DataFrame(index=range(5))

    try:
        for file in os.listdir(path_folder_data_ck):
            if not file.startswith('~$') and file.endswith('.xls'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Файл с расширением XLS (СТАРЫЙ ФОРМАТ EXCEL)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                base_error_df = pd.concat([base_error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                temp_df_ck = pd.read_excel(f'{path_folder_data_ck}/{file}', skiprows=5, nrows=5)
                if temp_df_ck.shape[1] != 30:
                    temp_error_df = pd.DataFrame(data=[
                        [f'{name_file}', '',
                         'Количество колонок в таблице не равно 30 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                        columns=['Название файла', 'Строка или колонка с ошибкой',
                                 'Описание ошибки'])
                    base_error_df = pd.concat([base_error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                # Создаем копию датафрейма только с числовыми колонками
                temp_df_int = temp_df_ck.iloc[:, 2:29].copy()
                # заменяем пустые ячейки нулями

                temp_df_int.fillna(0, inplace=True)
                temp_df_int = temp_df_int.applymap(int)

                # проверяем на ошибки
                temp_error_df = check_error_ck(temp_df_int.copy(), name_file)
                # Добавляем в итоговый датафрейм с ошибками
                base_error_df = pd.concat([base_error_df, temp_error_df], axis=0, ignore_index=True)
                # проверяем размер датафрейма с ошибками, если их нет то добавляем в результат.
                if base_error_df.shape[0] == 0:
                    base_df = base_df + temp_df_int  # складываем значения в таблицах
                    # делаем копию промежутчного датафрейма, так как мы будем добавлять новую колонку
                    temp_add_df = temp_df_int.copy()
                    temp_add_df.insert(0, 'Название файла', name_file)
                    temp_add_df['32'] = temp_df_ck.iloc[:, 29]
                    general_table = pd.concat([general_table, temp_add_df], axis=0,
                                              ignore_index=True)  # сохраняем в общую таблицу
                    # Добаввляем принимаемые меры
                    str_df = pd.concat([str_df, temp_df_ck.iloc[:, 29].to_frame().fillna('_')], axis=1,
                                       ignore_index=True)

                else:
                    continue
        # Объдиняем колонки с принимаемыми мерами в одну и добавляем в base df

        base_df['32'] = str_df.apply(lambda x: ';'.join(x), axis=1)
        # Добавляем колонки
        fourth = ['Всего (общая численность выпускников)',
                  'из общей численности выпускников (из строки 01): лица с ОВЗ',
                  'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                  'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                  'Имеют договор о целевом обучении']
        three = ['01', '02', '03', '04', '05']
        base_df.insert(0, '03', three)
        base_df.insert(1, '04', fourth)
        # в общую таблицу
        miultipler = general_table.shape[0] // 5
        general_table.insert(1, '03', three * miultipler)
        general_table.insert(2, '04', fourth * miultipler)

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        base_df.to_excel(f'{path_to_end_folder_ck}/Отчет ЦК Общий результат от {current_time}.xlsx', index=False)
        base_error_df.to_excel(f'{path_to_end_folder_ck}/Отчет ЦК Ошибки от {current_time}.xlsx', index=False)
        general_table.to_excel(f'{path_to_end_folder_ck}/Отчет ЦК Данные из всех таблиц от {current_time}.xlsx',
                               index=False)
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Не найдено значение {e.args}')
    except ValueError as e:
        foo_str = e.args[0].split(':')[1]

        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'В таблице с названием {name_file} в колонках: 05 -31 обнаружено НЕ числовое значение! В этих колонках не должно быть текста, пробелов или других символов, кроме чисел. \n'
                             f'Некорректное значение - {foo_str} !!!\n Исправьте и повторно запустите обработку')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Закройте открытые файлы Excel {e.args}')
    except:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'При обработке файла {name_file} возникла ошибка !!!\n'
                             f'Проверьте файл на соответствие шаблону.')

    else:
        if base_error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Отчет ЦК ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку чтбы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                'Данные успешно обработаны')


"""
обработка модерновой таблицы
"""


def processing_data_employment_modern():
    """
    Фугкция для обработки данных формы №15
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    try:
        for file in os.listdir(path_folder_data):
            if not file.startswith('~$') and file.endswith('.xls'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Файл с расширением XLS (СТАРЫЙ ФОРМАТ EXCEL)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=4, dtype=str)
                # создаем множество колонок наличие которых мы проверяем
                check_cols = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                              '16', '17',
                              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                              '33'}
                if not check_cols.issubset(set(df.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы в файле.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 5 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df = df[df['05'] != '16']  # фильтруем строки с проверками
                # отсекаем возможный первый столбец с данными ПОО,начинаем датафрейм с колонки 01 и отсекаем колонки с примечаниями
                df = df.loc[:, '01':'33']
                # получаем  часть с данными
                mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # проверяем есть ли строка полностью состоящая из nan
                empty_row_index = np.where(df.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df = df.iloc[:row_index]
                #     # Проверка на размер таблицы, должно бьть кратно 15
                count_spec = df.shape[0] // 15  # количество специальностей
                df = df.iloc[:count_spec * 15, :]  # отбрасываем строки проверки

                check_code_lst = df['03'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue
                # Проверка на непрерывность кода специальности, то есть на 15 строк должен быть только один код
                border_check_code = 0  # счетчик обработанных страниц
                quantity_check_code = len(check_code_lst) // 15  # получаем сколько специальностей в таблице
                flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 15 строк
                flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
                for i in range(quantity_check_code):
                    # получаем множество отбрасывая np.nan
                    # temp_set = set([code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 15] if
                    #                 code_spec is not np.nan])
                    temp_set = set(
                        [code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 15]])
                    if len(temp_set) != 1:
                        flag_error_code_spec = True
                    if ' ' in temp_set:
                        flag_error_space_spec = True
                    border_check_code += 15

                if flag_error_space_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if flag_error_code_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД СПЕЦИАЛЬНОСТИ НА КАЖДЫЕ 15 СТРОК (не считая строки с проверкой)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df.columns = list(map(str, df.columns))
                # Заполняем пока пропуски в 15 ячейке для каждой специальности
                df['06'] = df['06'].fillna('15 ячейка')

                # Проводим проверку на корректность данных, отправляем копию датафрейма
                tup_correct = (6, 20)  # создаем кортеж  с поправками
                file_error_df = check_error(df.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df['03'] = df['03'].apply(extract_code)  # очищаем от текста в кодах

                # Проверяем на наличие слова error что означает что там есть некорректные значения кодов специальности
                if 'error' in df['03'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 03 Код специальности.Вместо кода присутствует дата, вместе с кодом есть название,пробел перед кодом и т.п.!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                code_spec = [spec for spec in df['03'].unique()]

                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 16)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(7, 34)]

                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}

                # Изменяем последний ключ на строковый поскольку там будут хранится примечания
                for row, value in spec_dict.items():
                    for col, data in value.items():
                        if col == 'Колонка 33':
                            spec_dict[row][col] = ''
                # Создаем словарь среднего уровня содержащй данные по всем специальностям
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

                high_level_dct[name_file] = copy.deepcopy(poo_dct)

                """
                В итоге получается такая структура
                {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                """

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк

                # Итерируемся по полученному датафрейму через itertuples
                for row in df.itertuples():
                    # если счетчик колонок больше 15 то уменьшаем его до единицы
                    if idx_row > 15:
                        idx_row = 1

                    # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[3] is not np.nan) and (row[3] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[3] != current_code:
                            current_code = row[3]

                    data_row = row[7:34]  # получаем срез с нужными данными

                    for idx_col, value in enumerate(data_row, start=1):
                        if idx_col + 6 == 33:
                            # сохраняем примечания в строке
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 6}'] = f'{name_file} {check_data_note(value)};'

                        else:
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 6}'] += check_data(value)

                    idx_row += 1

        create_check_tables(high_level_dct)

        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        if col == 'Колонка 33':
                            itog_df[code_spec][row][col] += check_data_note(value)
                        else:
                            itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()

        frame['Всего'] = frame[0].apply(lambda x: x.get('Колонка 7'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 8'))
        frame[
            'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
            frame[0].apply(lambda x: x.get('Колонка 9'))
        frame[
            'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
            frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Индивидуальные предприниматели'] = frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = frame[
            0].apply(lambda x: x.get('Колонка 12'))
        frame['Продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 13'))
        frame['Проходят службу в армии по призыву'] = frame[0].apply(lambda x: x.get('Колонка 14'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 15'))
        frame['Находятся в отпуске по уходу за ребенком'] = frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Неформальная занятость (нелегальная)'] = frame[0].apply(lambda x: x.get('Колонка 17'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
            frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Иные причины нахождения под риском нетрудоустройства'] = frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Смерть, тяжелое состояние здоровья'] = frame[0].apply(lambda x: x.get('Колонка 21'))
        frame['Находятся под следствием, отбывают наказание'] = frame[0].apply(lambda x: x.get('Колонка 22'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            frame[0].apply(lambda x: x.get('Колонка 23'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
            frame[0].apply(lambda x: x.get('Колонка 24'))
        frame['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = frame[0].apply(
            lambda x: x.get('Колонка 25'))
        frame[
            'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
            frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['будут трудоустроены'] = frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['будут осуществлять предпринимательскую деятельность'] = frame[0].apply(lambda x: x.get('Колонка 28'))
        frame['будут самозанятыми'] = frame[0].apply(lambda x: x.get('Колонка 29'))
        frame['будут призваны в армию'] = frame[0].apply(lambda x: x.get('Колонка 30'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 31'))
        frame['будут продолжать обучение'] = frame[0].apply(lambda x: x.get('Колонка 32'))
        frame['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = frame[
            0].apply(lambda x: x.get('Колонка 33'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении',
               'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
            ,
               'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
               'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
               'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
               'Строка 12': 'расстройствами аутистического спектра',
               'Строка 13': 'с инвалидностью вследствие  других причин',
               'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
               'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
               }
        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        # добавляем строки с проверкой
        count = 0
        for i in range(15, len(finish_df) + 1, 15):
            new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
            new_row.iloc[:, 1] = 'Проверка (строка не редактируется)'
            new_row.iloc[:, 2:] = 'проверка пройдена'

            # Вставка новой строки через каждые 15 строк
            finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(
                drop=True)
            count += 1
        lst_number_row = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                          '16']
        multipler = len(finish_df) // 16  # получаем количество специальностей/профессий
        # вставляем новую колонку
        finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))
        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        finish_df.to_excel(f'{path_to_end_folder}/Полная таблица Форма №15 от {current_time}.xlsx', index=False)

        # Создаем файл с 5 строками
        small_finish_df = pd.DataFrame(columns=finish_df.columns)
        one_finish_df = pd.DataFrame(columns=finish_df.columns)

        lst_code_spec = finish_df['Код специальности'].unique()  # получаем список специальностей
        for code_spec in lst_code_spec:
            temp_df = finish_df[finish_df['Код специальности'] == code_spec]
            small_finish_df = pd.concat([small_finish_df, temp_df.iloc[:5, :]], axis=0, ignore_index=True)
            one_finish_df = pd.concat([one_finish_df, temp_df.iloc[:1, :]], axis=0, ignore_index=True)

        with pd.ExcelWriter(f'{path_to_end_folder}/5 строк Форма №15 от {current_time}.xlsx') as writer:
            small_finish_df.to_excel(writer, sheet_name='5 строк', index=False)
            one_finish_df.to_excel(writer, sheet_name='1 строка (Всего выпускников)', index=False)

        # small_finish_df.to_excel(f'{path_to_end_folder}/5 строк Форма №15 от {current_time}.xlsx', index=False)

        # Создаем документ
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ Форма №15 от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Закройте открытые файлы Excel {e.args}')
    except:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'При обработке файла {name_file} возникла ошибка !!!\n'
                             f'Проверьте файл на соответсвие шаблону')

    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                'Данные успешно обработаны.')


"""
Функция для ОПК
"""


def extract_code_full(value):
    """
    Функция для извлечения кода специальности из ячейки в которой соединены и код и название специалньости
    """
    value = str(value)
    re_code = re.compile('\d{2}?[.]\d{2}?[.]\d{2}')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.group()
    else:
        return 'Не найден код специальности'


"""
Проверка ошибок
"""


def check_error_opk(df1: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки таблиц по ОПК
    :param df1: датафрейм форма 1
    :param df2: датафрейм форма 2
    :param name_file: название обрабатываемого файла
    :param tup_correct: значаение корректировки
    :return: датафрейм с найденными ошибками
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # делаем датафреймы для простых проверок
    df1 = df1.iloc[:, 5:77]
    df1 = df1.applymap(check_data)

    # получаем количество датафреймов
    quantity = df1.shape[0] // 2
    # счетчик для обработанных строк
    border = 0
    for i in range(1, quantity + 1):
        temp_df = df1.iloc[border:border + 2, :]
        # Проверяем корректность заполнения формы 1
        first_error_opk = check_horizont_sum_opk_all(temp_df.copy(), name_file,
                                                     tup_correct)  # проверяем сумму по строкам
        error_df = pd.concat([error_df, first_error_opk], axis=0, ignore_index=True)

        # проверяем условие  по колонкам строка 02 не должна быть больше строки 01
        second_error_opk = check_vertical_opk_all(temp_df.copy(), border, name_file, tup_correct)
        error_df = pd.concat([error_df, second_error_opk], axis=0, ignore_index=True)

        # проверяем условие чтобы сумма по отраслям была равна колонке 08

        # список колонок которые нужно суммировать для проверки условия 07 = 08:31
        lst_07 = ['08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24',
                  '25', '26', '27', '28', '29', '30', '31']
        third_error_opk = check_horizont_chosen_sum_opk(temp_df.copy(), lst_07, '07', name_file)
        error_df = pd.concat([error_df, third_error_opk], axis=0, ignore_index=True)

        # считаем для будут трудоустроены по отраслям
        lst_038 = ['39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53', '54', '55',
                   '56', '57', '58', '59', '60', '61', '62']
        fourth_error_opk = check_horizont_chosen_sum_opk(temp_df.copy(), lst_038, '38', name_file)
        error_df = pd.concat([error_df, fourth_error_opk], axis=0, ignore_index=True)

        border += 2

    return error_df


def check_cross_error_opk(df1: pd.DataFrame, df2: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки значений между формой 1 и формой 2
    :param df1:
    :param df2:
    :param name_file:
    :param tup_correct:
    :return:
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    df1['08'] = df1['08'].apply(check_data)
    df1['39'] = df1['39'].apply(check_data)  # приводим к инту

    group_df1 = df1.groupby(['03']).agg({'08': sum, '39': sum})  # группируем
    group_df1 = group_df1.reset_index()  # переносим индексы
    group_df1.columns = ['Специальность', 'Трудоустроено в ОПК', 'Будут трудоустроены в ОПК']

    # приводим колонку 2 формы с числов выпускников к инту
    df2['04'] = df2['04'].apply(check_data)

    # Проверяем заполенение 2 формы, есть ли там вообще хоть что то
    quantity_now = group_df1['Трудоустроено в ОПК'].sum()  # сколько трудоустроено сейчас
    quantity_future = group_df1['Будут трудоустроены в ОПК'].sum()  # сколько будут трудоустроены
    # проверяем заполнение формы 2
    if (quantity_now != 0 or quantity_future != 0) and df2.shape[0] == 0:
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                            'В форме 1 есть выпускники трудоустроенные или которые будут трудоустроены в ОПК,\n'
                                            ' при этом форма 2 не заполнена. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
        return error_df

    # проверка по трудоустроенным и трудоустроенным в будущем
    cross_first_error_df = check_cross_first_error_df(df1.copy(), df2.copy(), name_file)
    error_df = pd.concat([error_df, cross_first_error_df], axis=0, ignore_index=True)

    # # проверка по целевикам
    check_cross_second_error_df = check_cross_second_error(df1.copy(), df2.copy(), name_file)
    error_df = pd.concat([error_df, check_cross_second_error_df], axis=0, ignore_index=True)

    return error_df


def check_cross_first_error_df(df1: pd.DataFrame, df2: pd.DataFrame, name_file):
    """
    Функция для првоерки соответствия количества указанных в форме 1 трудоустроенных и списка в форме 2
    проверки 1 и 2

    :param df1:
    :param df2:
    :param name_file:
    :return:
    """
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # проверяем наличие незаполненных ячеек в колонках 05 06
    # отбираем значения первой строки
    df1 = df1[df1['04'] == '01']
    df1 = df1.groupby(['03']).agg({'08': sum, '39': sum})  # группируем
    df1 = df1.reset_index()  # переносим индексы
    df1.columns = ['Специальность', 'Трудоустроено в ОПК', 'Будут трудоустроены в ОПК']

    etalon_05 = {'уже трудоустроены', 'будут трудоустроены'}  # эталонный состав колонки 05
    etalon_06 = {'заключили договор о целевом обучении', 'нет'}  # эталонный состав колонки 05
    # получаем состав колонок
    st_05 = set(df2['05'].unique())
    st_06 = set(df2['06'].unique())

    if not (st_05.issubset(etalon_05) or st_06.issubset(etalon_06)):
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                            'В колонках 05 или 06 формы 2 есть незаполненные ячейки,\n'
                                            ' или значения отличающиеся от требуемых. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # создаем 2 датафрейма, по колонке 05 трудоустроены и будут трудоустроены
    empl_now_df = df2[df2['05'] == 'уже трудоустроены']  # те что уже трудоустроены
    empl_future_df = df2[df2['05'] == 'будут трудоустроены']  # те что будут трудоустроены

    # проводим группировку
    empl_now_df_group = empl_now_df.groupby(['02']).agg({'04': sum})
    empl_future_df_group = empl_future_df.groupby(['02']).agg({'04': sum})

    df1_future = df1[df1[
                         'Будут трудоустроены в ОПК'] != 0]  # отбираем в форме 2 специальности по которым есть будущие трудоустроены выпускники
    check_df = empl_future_df_group.merge(df1_future, how='outer', left_on='02', right_on='Специальность')

    # находим строки где есть хотя бы один nan ,это значит что в формах есть разночтения по специальностям
    row_with_nan = check_df[check_df.isna().any(axis=1)]
    row_with_nan.fillna('Специальность не найдена', inplace=True)
    for row in row_with_nan.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'{row[2]} не совпадают данные !!! Отсутствуют данные по этой специальности либо в форме 1 либо в форме 2',
                                            'В форме 1 для этой специальности указаны выпускники которые будут трудоустроены, но в форме 2 такой специальности не найдено или наоборот. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # отбираем все строки где нет нан
    check_df = check_df[~check_df.isna().any(axis=1)]

    check_df['Результат'] = check_df['04'] == check_df['Будут трудоустроены в ОПК']
    check_df = check_df[~check_df['Результат']]

    # записываем где есть ошибки
    for row in check_df.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'{row[2]} не совпадают данные !!! по форме 1 для этой специальности будут трудоустроено {row[4]} чел.'
                                            f' в форме 2 по этой специальности найдено {int(row[1])} чел.',
                                            'Несовпадает количество выпускников которые будут трудоустроены в форме 1 и в форме 2. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # обрабатываем датафрейм с уже трудоустроенными
    df1_now = df1[df1[
                      'Трудоустроено в ОПК'] != 0]  # отбираем в форме 2 специальности по которым есть будущие трудоустроены выпускники
    check_df_now = empl_now_df_group.merge(df1_now, how='outer', left_on='02', right_on='Специальность')

    # находим строки где есть хотя бы один nan ,это значит что в формах есть разночтения по специальностям
    row_with_nan_now = check_df_now[check_df_now.isna().any(axis=1)]
    row_with_nan_now.fillna('Специальность не найдена', inplace=True)
    for row in row_with_nan_now.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'Перекрестная проверка трудоустроен/будет трудоустроен: {row[2]} не совпадают данные !!! Отсутствуют данные по этой специальности либо в форме 1 либо в форме 2',
                                            'В форме 1 для этой специальности указаны выпускники которые уже трудоустроены, но в форме 2 такой специальности не найдено или наоборот. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # отбираем все строки где нет нан
    check_df_now = check_df_now[~check_df_now.isna().any(axis=1)]

    check_df_now['Результат'] = check_df_now['04'] == check_df_now['Трудоустроено в ОПК']
    check_df_now = check_df_now[~check_df_now['Результат']]

    # записываем где есть ошибки
    for row in check_df_now.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'Перекрестная проверка трудоустроен/будет трудоустроен: {row[2]} не совпадают данные !!! по форме 1 для этой специальности трудоустроено {row[3]} чел.'
                                            f' в форме 2 по этой специальности найдено {int(row[1])} чел.',
                                            'Несовпадает количество выпускников которые трудоустроены в форме 1 и в форме 2. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    return error_df


def check_cross_second_error(df1: pd.DataFrame, df2: pd.DataFrame, name_file):
    """
    Функция для проверки  корректности заполнения показателей по целевому приему.
    :param df1:
    :param df2:
    :param name_file:
    :return:
    """
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # проверяем наличие незаполненных ячеек в колонках 05 06
    # отбираем значения первой строки
    df1 = df1[df1['04'] == '02']
    df1 = df1.groupby(['03']).agg({'08': sum, '39': sum})  # группируем
    df1 = df1.reset_index()  # переносим индексы
    df1.columns = ['Специальность', 'Трудоустроено в ОПК', 'Будут трудоустроены в ОПК']
    # Создаем 2 датафрейма из формы 1 целевики будущие и настоящие
    df1_now = df1[['Специальность', 'Трудоустроено в ОПК']]  # целевики трудоустроенные форма 1
    df1_now = df1_now[df1_now['Трудоустроено в ОПК'] != 0]
    df1_future = df1[['Специальность', 'Будут трудоустроены в ОПК']]  # целевики трудоустроенные форма 1
    df1_future = df1_future[df1_future['Будут трудоустроены в ОПК'] != 0]

    # создаем 2 датафрейма целевики уже трудоустроенные и целевики в будущем трудоустроенные
    # создаем датафрейм, по колонке 06 заключили договор о целевом обучении и уже трудоустроенные
    target_df_now = df2[df2['06'] == 'заключили договор о целевом обучении']  # целевики
    target_df_now = target_df_now[target_df_now['05'] == 'уже трудоустроены']  # целевики

    # проводим группировку
    target_df_now = target_df_now.groupby(['02']).agg({'04': sum})
    target_df_now = target_df_now.reset_index()

    union_df_now = df1_now.merge(target_df_now, how='outer', left_on='Специальность', right_on='02')
    #
    #
    # # находим строки где есть хотя бы один nan ,это значит что в формах есть разночтения по специальностям
    row_with_nan = union_df_now[union_df_now.isna().any(axis=1)]
    row_with_nan.fillna('Специальность не найдена', inplace=True)

    for row in row_with_nan.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'Перекрестная проверка целевиков (трудоустроен/будет трудоустроен): в форме 1 - {row[1]} а в форме 2 - {row[3]} не совпадают данные !!! Отсутствуют данные по этой специальности либо в форме 1 либо в форме 2',
                                            'В форме 1 для специальности указаны выпускники целевики которые  трудоустроены, но в форме 2 такой специальности не найдено или наоборот. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
    # # отбираем все строки где нет нан
    check_df_now = union_df_now[~union_df_now.isna().any(axis=1)]

    check_df_now['Результат'] = check_df_now['04'] == check_df_now['Трудоустроено в ОПК']
    check_df_now = check_df_now[~check_df_now['Результат']]
    # записываем где есть ошибки
    for row in check_df_now.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'Перекрестная проверка целевиков (трудоустроен/будет трудоустроен):В форме 1 - {row[1]} не совпадают данные !!! по форме 1 для этой специальности трудоустроено {row[2]} чел.'
                                            f' в форме 2 по этой специальности найдено {int(row[4])} чел.',
                                            'Несовпадает количество выпускников целевиков которые трудоустроены в форме 1 и в форме 2. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
    # создаем 2 датафрейма целевики уже трудоустроенные и целевики в будущем трудоустроенные
    # создаем датафрейм, по колонке 06 заключили договор о целевом обучении и уже трудоустроенные
    target_df_future = df2[df2['06'] == 'заключили договор о целевом обучении']  # целевики
    target_df_future = target_df_future[target_df_future['05'] == 'будут трудоустроены']  # целевики

    # проводим группировку
    target_df_future = target_df_future.groupby(['02']).agg({'04': sum})
    target_df_future = target_df_future.reset_index()

    union_df_future = df1_future.merge(target_df_future, how='outer', left_on='Специальность', right_on='02')
    #
    #
    # # находим строки где есть хотя бы один nan ,это значит что в формах есть разночтения по специальностям
    row_with_nan = union_df_future[union_df_future.isna().any(axis=1)]
    row_with_nan.fillna('Специальность не найдена', inplace=True)
    for row in row_with_nan.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'Перекрестная проверка целевиков (трудоустроен/будет трудоустроен): В форме 1- {row[1]} а в форме 2- {row[3]} !!! Отсутствуют данные по этой специальности либо в форме 1 либо в форме 2',
                                            'В форме 1 для специальности указаны выпускники целевики которые БУДУТ трудоустроены, но в форме 2 выпускников целевиков с такой специальностью которые БУДУт трудоустроены не найдено ИЛИ наоборот. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
    # # отбираем все строки где нет нан
    check_df_future = union_df_future[~union_df_future.isna().any(axis=1)]

    check_df_future['Результат'] = check_df_future['04'] == check_df_future['Будут трудоустроены в ОПК']
    check_df_future = check_df_future[~check_df_future['Результат']]
    # записываем где есть ошибки
    for row in check_df_future.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}',
                                            f'{row[1]} не совпадают данные !!! по форме 1 для этой специальности трудоустроено {row[2]} чел.'
                                            f' в форме 2 по этой специальности найдено {int(row[4])} чел.',
                                            'Несовпадает количество выпускников целевиков которые БУДУТ трудоустроены в форме 1 и в форме 2. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    return error_df


def check_horizont_sum_opk_all(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки простой горизонтальной суммы 06 = 07 +32,33,34,35,36,37,38,63,64,65,66:77
    :param df:
    :param name_file:
    :param tup_correct:
    :return:
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # конвертируем в инт
    drop_lst = ['08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
                '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '39', '40',
                '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53',
                '54', '55', '56', '57', '58', '59', '60', '61', '62']

    # удаляем колонки лишние колонки
    df.drop(columns=drop_lst, inplace=True)

    # # получаем сумму колонок
    df['Сумма'] = df.iloc[:, 1:].sum(axis=1)
    # # Проводим проверку
    df['Результат'] = df['06'] == df['Сумма']
    # # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: гр. 06 = гр.07 + сумма(всех колонок за исключением распределения по отраслям)'
    return temp_error_df


def check_horizont_chosen_sum_opk(df: pd.DataFrame, tup_checked_cols: list, name_itog_cols, name_file):
    """
    Функция для проверки равенства одиночных или небольших групп колонок
    tup_checked_cols колонки сумму которых нужно сравнить с name_itog_cols чтобы она не превышала это значение
    """
    # Считаем проверяемые колонки
    df['Сумма'] = df[tup_checked_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {x+9}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} == сумма {tup_checked_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_vertical_opk_all(df: pd.DataFrame, border, name_file, tup_correct):
    """
    Функция для проверки условия Количество целевиков не должно быть больше чем количество выпускников
    :param df:
    :param name_file:
    :param tup_correct:
    :return:
    """
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['01', '02'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['Результат'] = foo_df['02'] <= foo_df['01']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 02 <= стр. 01 '
    return temp_error_df


def create_check_tables_opk(high_level_dct: dict):
    """
        Функция для создания файла с данными по каждой специальности
        """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')

            temp_code_df = temp_code_df.stack()
            # название такое выбрал потому что было лень заменять значения из блокнота юпитера
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Суммарный выпуск 2023 г.'] = temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df['Трудоустроены на предприятия оборонно-промышленного комплекса*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df['Трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df['Трудоустроены на предприятия сельского хозяйства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Трудоустроены на предприятия металлургии'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df['Трудоустроены на предприятия железнодорожного транспорта'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
            temp_code_df['Трудоустроены на предприятия легкой промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df['Трудоустроены на предприятия химической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df['Трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Трудоустроены на предприятия фармацевтической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 16'))
            temp_code_df['Трудоустроены на предприятия отрасли информационных технологий'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df['Трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Трудоустроены на предприятия транспортной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 20'))
            temp_code_df['Трудоустроены на предприятия горнодобывающей отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df[
                'Трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 22'))
            temp_code_df['Трудоустроены на предприятия лесной промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df['Трудоустроены на предприятия строительной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df[
                'Трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 25'))
            temp_code_df['Трудоустроены на предприятия индустрии робототехники'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['Трудоустроены на предприятия в отрасли образования'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['Трудоустроены на предприятия в медицинской отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 28'))
            temp_code_df[
                'Трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
            temp_code_df['Трудоустроены на предприятия в отрасли искусства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
            temp_code_df['Трудоустроены на предприятия иная отрасль'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 31'))

            temp_code_df['Индивидуальные предприниматели'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
            temp_code_df['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессиональный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 33'))
            temp_code_df['Продолжили обучение'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 34'))
            temp_code_df['Проходят службу в армии по призыву'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 35'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 36'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 37'))
            temp_code_df[
                'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 38'))
            temp_code_df['Будут трудоустроены на предприятия оборонно-промышленного комплекса* '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 39'))
            temp_code_df['Будут трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 40'))
            temp_code_df['Будут трудоустроены на предприятия сельского хозяйства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 41'))
            temp_code_df['Будут трудоустроены на предприятия металлургии'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 42'))
            temp_code_df['Будут трудоустроены на предприятия железнодорожного транспорта'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 43'))
            temp_code_df['Будут трудоустроены на предприятия легкой промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 44'))
            temp_code_df['Будут трудоустроены на предприятия химической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 45'))
            temp_code_df[
                'Будут трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 46'))
            temp_code_df['Будут трудоустроены на предприятия фармацевтической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 47'))
            temp_code_df['Будут трудоустроены на предприятия отрасли информационных технологий'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 48'))
            temp_code_df[
                'Будут трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 49'))
            temp_code_df[
                'Будут трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 50'))

            temp_code_df['Будут трудоустроены на предприятия транспортной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 51'))
            temp_code_df['Будут трудоустроены на предприятия горнодобывающей отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 52'))
            temp_code_df[
                'Будут трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 53'))
            temp_code_df['Будут трудоустроены на предприятия лесной промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 54'))
            temp_code_df['Будут трудоустроены на предприятия строительной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 55'))
            temp_code_df[
                'Будут трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 56'))
            temp_code_df['Будут трудоустроены на предприятия индустрии робототехники'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 57'))
            temp_code_df['Будут трудоустроены на предприятия в отрасли образования'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 58'))
            temp_code_df['Будут трудоустроены на предприятия в медицинской отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 59'))
            temp_code_df[
                'Будут трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 60'))
            temp_code_df['Будут трудоустроены на предприятия в отрасли искусства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 61'))
            temp_code_df['Будут трудоустроены на предприятия иная отрасль'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 62'))
            temp_code_df['Будут индивидуальными предпринимателями'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 63'))
            temp_code_df['Будут самозанятыми'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 64'))
            temp_code_df['Будут продолжать обучение'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 65'))
            temp_code_df['Будут призваны в армию'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 66'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 67'))
            temp_code_df['Будут находиться в отпуске по уходу за ребенком'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 68'))
            temp_code_df['Неформальная занятость (теневой сектор экономики)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 69'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 70'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 71'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 72'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 73'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 74'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 75'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 76'))
            temp_code_df[
                'Иное в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 0,5% выпускников - укажите причины в графе "принимаемые меры"'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 77'))
            temp_code_df[
                'Принимаемые меры по содействию занятости, в том числе по трудоустройству выпускников на предприятия оборонно-промышленного комплекса тезисно - вид меры, охват выпускников мерой'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 78'))

            finish_df = temp_code_df.drop([0], axis=1)

            finish_df = finish_df.reset_index()

            finish_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из строки 01: имеют договор о целевом обучении'}

            finish_df['Наименование показателей (категория выпускников)'] = finish_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    wb.save(f'{path_to_end_folder_opk}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')


def processing_data_opk_employment():
    """
    Функция для обработки полной таблицы занятости выпускников в ОПК
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    all_form2 = pd.DataFrame(columns=['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11'])

    try:
        for file in os.listdir(path_folder_data_opk):
            if not file.startswith('~$') and file.endswith('.xls'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Файл с расширением XLS (СТАРЫЙ ФОРМАТ EXCEL)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                # Проверяем наличие листов с названиями Форма 1 и Форма 2
                wb_1 = openpyxl.load_workbook(f'{path_folder_data_opk}/{file}')
                if not {'Форма 1', 'Форма 2'}.issubset(set(wb_1.sheetnames)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте наличие листов с названием Форма 1 и Форма 2! Не должно быть пробелов в начале и в конце названия ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df_form1 = pd.read_excel(f'{path_folder_data_opk}/{file}', skiprows=8, dtype=str,
                                         sheet_name='Форма 1')  # общие данные

                # Находим строку с номерами колонок, так как вполне возможно в файле остались примеры
                temp_wb = openpyxl.load_workbook(f'{path_folder_data_opk}/{file}',
                                                 read_only=True)  # открываем файл в режиме чтения
                temp_ws = temp_wb['Форма 2']
                threshold_form2 = 5
                for row in temp_ws.iter_rows(0):  # перебираем значения в первой колонке
                    for cell in row:
                        if cell.value == '01':
                            threshold_form2 = cell.row
                temp_wb.close()  # закрываем файл чтобы потом не было ошибок
                form2_df = pd.read_excel(f'{path_folder_data_opk}/{file}', skiprows=threshold_form2 - 1, dtype=str,
                                         sheet_name='Форма 2')  # подробные данные по ОПК
                # создаем множество колонок наличие которых мы проверяем
                check_cols = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                              '16',
                              '17',
                              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                              '33',
                              '34', '35', '36',
                              '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51',
                              '52',
                              '53', '54', '55',
                              '56', '57', '58', '59', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70',
                              '71',
                              '72', '73', '74',
                              '75', '76', '77', '78', '79', '80'}
                if not check_cols.issubset(set(df_form1.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы на листе Форма 1.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 9 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # проверяем корректность формы 2
                check_cols_form2 = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11'}
                if not check_cols_form2.issubset(set(form2_df.columns)):
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Проверьте заголовок таблицы на листе Форма 2.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться ВЫШЕ списка трудоустроенных.! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df_form1 = df_form1[df_form1['04'] != '03']  # фильтруем строки с проверкой
                form2_df.dropna(axis=0, inplace=True, how='all')  # убираем все пустые строки
                # заполняем пустые строки в колонке 01
                form2_df['01'] = form2_df['01'].fillna('Не заполнено')
                form2_df = form2_df[
                    ~form2_df['01'].str.contains('Проверка', case=False)]  # фильруем строки с проверкой на листе 2
                form2_df = form2_df[
                    ~form2_df['01'].str.contains('Не заполнено', case=False)]  # фильруем строки с проверкой на листе 2

                df_form1 = df_form1.loc[:, '01':'78']  # отсекаем возможную первую колонку и колонки с примечаниями
                # получаем  часть с данными
                mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
                if mask[0]:  # если пустая строка идет первой то удаляем ее и обновляем маску
                    df_form1.drop(axis=0, index=0, inplace=True)
                    mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # Находим индекс первой пустой строки, если он есть,получаем список с значениями где есть пустые строки
                empty_row_index = np.where(df_form1.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df_form1 = df_form1.iloc[:row_index]
                quantity_spec = df_form1.shape[0] // 2  # получаем количество специальностей в файле

                check_two_rows_spec = df_form1['04'].tolist() == ['01',
                                                                  '02'] * quantity_spec  # проверяем чтобы колонка 04 состояла только из 01 и 02
                if not check_two_rows_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Возможно пропущена строка 01 или 02. Для каждой спец./проф. должны быть  только строки 01 и 02 не считая строки 03 с проверкой. Проверьте наличие ПУСТОЙ строки после таблицы! Последняя строка проверки должна быть заполнена полностью! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                check_code_lst = df_form1['03'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue

                # Проверка на непрерывность кода специальности, то есть на 2 строки должен быть только один код
                border_check_code = 0  # счетчик обработанных страниц
                quantity_check_code = len(check_code_lst) // 2  # получаем сколько специальностей в таблице
                flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 2 строки
                flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
                for i in range(quantity_check_code):
                    # получаем множество отбрасывая np.nan
                    temp_set = set([code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 2]])
                    print(temp_set)
                    if len(temp_set) != 1:
                        flag_error_code_spec = True
                    if ' ' in temp_set:
                        flag_error_space_spec = True
                    border_check_code += 2

                if flag_error_space_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if flag_error_code_spec:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД и Название СПЕЦИАЛЬНОСТИ/ПРОФЕССИИ НА КАЖДЫЕ 2 СТРОКИ (не считая строки с проверкой)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                """
                ПРОВЕРКИ
                в том числе проверка кода специальности

                """
                tup_correct = (10, 12)  # создаем кортеж  с поправками
                file_error_df = check_error_opk(df_form1.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)

                # проводим кросс проверки между 2 формами
                file_cross_error_df = check_cross_error_opk(df_form1.copy(), form2_df.copy(), name_file, tup_correct)
                error_df = pd.concat([error_df, file_cross_error_df], axis=0, ignore_index=True)

                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if file_cross_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки связанные с сравнением данных из формы 1 и формы 2!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df_form1['03'] = df_form1['03'].apply(extract_code_full)  # очищаем от текста в кодах
                if 'Не найден код специальности' in df_form1['03'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 03 Код специальности.Вместо кода присутствует дата,слитное написание кода и названия, вместе с кодом есть название,пробел перед кодом и т.п.!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                """
                Окончание блока проверок
                """
                all_form2 = pd.concat([all_form2, form2_df], axis=0, ignore_index=True)

                code_spec = [spec for spec in df_form1['03'].unique()]

                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 3)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(6, 79)]

                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}

                # Изменяем последний ключ на строковый поскольку там будут хранится примечания
                for row, value in spec_dict.items():
                    for col, data in value.items():
                        if col == 'Колонка 78':
                            spec_dict[row][col] = ''

                # Создаем словарь среднего уровня содержащй данные по всем специальностям
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

                high_level_dct[name_file] = copy.deepcopy(poo_dct)

                """
                В итоге получается такая структура
                {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                """

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк
                # Итерируемся по полученному датафрейму через itertuples
                for row in df_form1.itertuples():
                    # если счетчик колонок больше 15 то уменьшаем его до единицы
                    if idx_row > 2:
                        idx_row = 1
                        # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[3] is not np.nan) and (row[3] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[3] != current_code:
                            current_code = row[3]
                    data_row = row[6:79]  # получаем срез с нужными данными
                    for idx_col, value in enumerate(data_row, start=1):
                        if idx_col + 5 == 78:
                            # сохраняем примечания в строке
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 5}'] = f'{name_file} {check_data_note(value)};'

                        else:
                            high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                                f'Колонка {idx_col + 5}'] += check_data(value)
                    idx_row += 1
        # создаем файл для проверки. Специальности с разбивкой по организациям
        create_check_tables_opk(high_level_dct)

        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        if col == 'Колонка 78':
                            itog_df[code_spec][row][col] += check_data_note(value)
                        else:
                            itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()

        frame['Суммарный выпуск 2023 г.'] = frame[0].apply(lambda x: x.get('Колонка 6'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 7'))
        frame['Трудоустроены на предприятия оборонно-промышленного комплекса*'] = \
            frame[0].apply(lambda x: x.get('Колонка 8'))
        frame['Трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 9'))
        frame['Трудоустроены на предприятия сельского хозяйства'] = \
            frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Трудоустроены на предприятия металлургии'] = \
            frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Трудоустроены на предприятия железнодорожного транспорта'] = \
            frame[0].apply(lambda x: x.get('Колонка 12'))
        frame['Трудоустроены на предприятия легкой промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 13'))
        frame['Трудоустроены на предприятия химической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 14'))
        frame['Трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 15'))
        frame['Трудоустроены на предприятия фармацевтической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Трудоустроены на предприятия отрасли информационных технологий'] = \
            frame[0].apply(lambda x: x.get('Колонка 17'))
        frame['Трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Трудоустроены на предприятия транспортной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Трудоустроены на предприятия горнодобывающей отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 21'))
        frame[
            'Трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 22'))
        frame['Трудоустроены на предприятия лесной промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 23'))
        frame['Трудоустроены на предприятия строительной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 24'))
        frame[
            'Трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 25'))
        frame['Трудоустроены на предприятия индустрии робототехники'] = \
            frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['Трудоустроены на предприятия в отрасли образования'] = \
            frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['Трудоустроены на предприятия в медицинской отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 28'))
        frame[
            'Трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
            frame[0].apply(lambda x: x.get('Колонка 29'))
        frame['Трудоустроены на предприятия в отрасли искусства'] = \
            frame[0].apply(lambda x: x.get('Колонка 30'))
        frame['Трудоустроены на предприятия иная отрасль'] = \
            frame[0].apply(lambda x: x.get('Колонка 31'))

        frame['Индивидуальные предприниматели'] = \
            frame[0].apply(lambda x: x.get('Колонка 32'))
        frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессиональный доход)'] = \
            frame[0].apply(lambda x: x.get('Колонка 33'))
        frame['Продолжили обучение'] = \
            frame[0].apply(lambda x: x.get('Колонка 34'))
        frame['Проходят службу в армии по призыву'] = \
            frame[0].apply(lambda x: x.get('Колонка 35'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
            frame[0].apply(lambda x: x.get('Колонка 36'))
        frame['Находятся в отпуске по уходу за ребенком'] = \
            frame[0].apply(lambda x: x.get('Колонка 37'))
        frame[
            'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 38'))
        frame['Будут трудоустроены на предприятия оборонно-промышленного комплекса* '] = \
            frame[0].apply(lambda x: x.get('Колонка 39'))
        frame['Будут трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 40'))
        frame['Будут трудоустроены на предприятия сельского хозяйства'] = \
            frame[0].apply(lambda x: x.get('Колонка 41'))
        frame['Будут трудоустроены на предприятия металлургии'] = \
            frame[0].apply(lambda x: x.get('Колонка 42'))
        frame['Будут трудоустроены на предприятия железнодорожного транспорта'] = \
            frame[0].apply(lambda x: x.get('Колонка 43'))
        frame['Будут трудоустроены на предприятия легкой промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 44'))
        frame['Будут трудоустроены на предприятия химической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 45'))
        frame['Будут трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 46'))
        frame['Будут трудоустроены на предприятия фармацевтической отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 47'))
        frame['Будут трудоустроены на предприятия отрасли информационных технологий'] = \
            frame[0].apply(lambda x: x.get('Колонка 48'))
        frame['Будут трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 49'))
        frame[
            'Будут трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 50'))

        frame['Будут трудоустроены на предприятия транспортной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 51'))
        frame['Будут трудоустроены на предприятия горнодобывающей отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 52'))
        frame[
            'Будут трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 53'))
        frame['Будут трудоустроены на предприятия лесной промышленности'] = \
            frame[0].apply(lambda x: x.get('Колонка 54'))
        frame['Будут трудоустроены на предприятия строительной отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 55'))
        frame[
            'Будут трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
            frame[0].apply(lambda x: x.get('Колонка 56'))
        frame['Будут трудоустроены на предприятия индустрии робототехники'] = \
            frame[0].apply(lambda x: x.get('Колонка 57'))
        frame['Будут трудоустроены на предприятия в отрасли образования'] = \
            frame[0].apply(lambda x: x.get('Колонка 58'))
        frame['Будут трудоустроены на предприятия в медицинской отрасли'] = \
            frame[0].apply(lambda x: x.get('Колонка 59'))
        frame[
            'Будут трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
            frame[0].apply(lambda x: x.get('Колонка 60'))
        frame['Будут трудоустроены на предприятия в отрасли искусства'] = \
            frame[0].apply(lambda x: x.get('Колонка 61'))
        frame['Будут трудоустроены на предприятия иная отрасль'] = \
            frame[0].apply(lambda x: x.get('Колонка 62'))
        frame['Будут индивидуальными предпринимателями'] = \
            frame[0].apply(lambda x: x.get('Колонка 63'))
        frame['Будут самозанятыми'] = \
            frame[0].apply(lambda x: x.get('Колонка 64'))
        frame['Будут продолжать обучение'] = \
            frame[0].apply(lambda x: x.get('Колонка 65'))
        frame['Будут призваны в армию'] = \
            frame[0].apply(lambda x: x.get('Колонка 66'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
            frame[0].apply(lambda x: x.get('Колонка 67'))
        frame['Будут находиться в отпуске по уходу за ребенком'] = \
            frame[0].apply(lambda x: x.get('Колонка 68'))
        frame['Неформальная занятость (теневой сектор экономики)'] = \
            frame[0].apply(lambda x: x.get('Колонка 69'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
            frame[0].apply(lambda x: x.get('Колонка 70'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
            frame[0].apply(lambda x: x.get('Колонка 71'))
        frame['Иные причины нахождения под риском нетрудоустройства'] = \
            frame[0].apply(lambda x: x.get('Колонка 72'))
        frame['Смерть, тяжелое состояние здоровья'] = \
            frame[0].apply(lambda x: x.get('Колонка 73'))
        frame['Находятся под следствием, отбывают наказание'] = \
            frame[0].apply(lambda x: x.get('Колонка 74'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            frame[0].apply(lambda x: x.get('Колонка 75'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
            frame[0].apply(lambda x: x.get('Колонка 76'))
        frame[
            'Иное в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 0,5% выпускников - укажите причины в графе "принимаемые меры"'] = \
            frame[0].apply(lambda x: x.get('Колонка 77'))
        frame[
            'Принимаемые меры по содействию занятости, в том числе по трудоустройству выпускников на предприятия оборонно-промышленного комплекса тезисно - вид меры, охват выпускников мерой'] = \
            frame[0].apply(lambda x: x.get('Колонка 78'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из строки 01: имеют договор о целевом обучении'}

        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        # создаем итоговый отчет по ОПК
        opk_df = finish_df[['Суммарный выпуск 2023 г.',
                            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                            'Трудоустроены на предприятия оборонно-промышленного комплекса*',
                            'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                            'Будут трудоустроены на предприятия оборонно-промышленного комплекса* ']]
        opk_df = opk_df.sum(axis=0).to_frame()  # суммируем данные
        opk_df = opk_df.transpose()  # разворачиваем из колонки в строку

        # Создаем сумму по всем колонкам
        all_sum_df = finish_df.iloc[:, 2:].sum(axis=0).to_frame()  # суммируем данные
        all_sum_df = all_sum_df.transpose()  # разворачиваем из колонки в строку

        # добавляем строки с проверкой
        count = 0
        for i in range(2, len(finish_df) + 1, 2):
            new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
            new_row.iloc[:, 1] = 'Проверка (строка не редактируется) Строку требуется подставлять после каждого кода'
            new_row.iloc[:, 2:] = 'проверка пройдена'

            # Вставка новой строки через каждые 15 строк
            finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(
                drop=True)
            count += 1

        lst_number_row = ['01', '02', '03']
        multipler = len(finish_df) // 3  # получаем количество специальностей/профессий
        # вставляем новую колонку
        finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))

        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        # создаем сокращенный датафрейм по специальностям из плоной таблицы
        opk_finish_df = finish_df[['Код специальности', 'Номер строки',
                                   'Наименование показателей (категория выпускников)',
                                   'Суммарный выпуск 2023 г.',
                                   'Трудоустроены на предприятия оборонно-промышленного комплекса*',
                                   'Будут трудоустроены на предприятия оборонно-промышленного комплекса* ']]

        opk_finish_df.to_excel(
            f'{path_to_end_folder_opk}/Трудоустройство по специальностям (ОПК) от {current_time}.xlsx',
            index=False)
        finish_df.to_excel(
            f'{path_to_end_folder_opk}/Полная таблица Трудоустройство по отраслям от {current_time}.xlsx',
            index=False)

        # считаем сколько целевиков
        target_df = finish_df[finish_df['Номер строки'] == '02']
        target_df = target_df.iloc[:, 2:].sum(axis=0).to_frame()  # суммируем данные
        target_df = target_df.transpose()  # разворачиваем из колонки в строку
        # сохраняем результирующие датафреймы в один файл
        with pd.ExcelWriter(f'{path_to_end_folder_opk}/Итоги по ОПК,целевикам и всем колонкам {current_time}.xlsx',
                            engine='openpyxl') as writer:
            opk_df.to_excel(writer, sheet_name='Итог по ОПК')
            all_sum_df.to_excel(writer, sheet_name='Итог по всем колонкам')
            target_df.to_excel(writer, sheet_name='Итог по целевикам')

        # обрабатываем список из второй формы
        all_form2.columns = ['Субъект Российской Федерации', 'Код и наименование профессии, специальности',
                             'Наименование образовательной организации среднего профессионального образования, в которой обучался выпускник',
                             'Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса',
                             'Категория выпускников (выпускники, которые уже трудоустроены/выпускники, которые будут трудоустроены)',
                             'Наличие договора о целевом обучении (имеющие договор о целевом обучении/не имеющие договор о целевом обучении) с предприятием оборонно-промышленного комплекса',
                             'ИНН предприятия',
                             'Наименование предприятия/филиала предприятия/структурного подразделения предприятия оборонно-промышленного комплекса, на котором трудоустроен выпускник/планирует трудоустройство',
                             'Наименование профессии/должности, по которой трудоустраивается (планирует трудоустройство) выпускник',
                             'Трудоустройство в соответствии с освоенной профессией, специальностью (да/нет)',
                             'Сложности при трудоустройстве (например, при взаимодействии с предприятием, наличие рисков расторжения договора о целевом обучении и т.д.), реализуемые меры (описательная часть. В отсутствие сложностей - пропустите графу)']

        all_form2['Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса']=all_form2['Количество выпускников 2023 г. (включая ожидаемый выпуск) уже трудоустроенных либо планирующих трудоустройство на предприятия оборонно-промышленного комплекса'].apply(check_data)
        #проверяем на пустоту
        if all_form2.shape[0] != 0:
            # Создаем сводную таблицу
            #переименываем колонки для удобства
            all_form2.columns = ['Регион', 'Специальность', 'Наименование', 'Количество', 'Трудоустройство',
                                 'Целевой договор', 'ИНН', 'Предприятие', 'Должность', 'Трудоустройство по специальности',
                                 'Сложности']
            # делаем категориальными значения в некоторых колонках
            all_form2['Трудоустройство'] = all_form2['Трудоустройство'].astype('category')
            all_form2["Трудоустройство"].cat.set_categories(["уже трудоустроены", "будут трудоустроены"], inplace=True)
            all_form2['Целевой договор'] = all_form2['Целевой договор'].astype('category')
            all_form2["Целевой договор"].cat.set_categories(["нет", "заключили договор о целевом обучении"], inplace=True)
            all_form2['Трудоустройство по специальности'] = all_form2['Трудоустройство по специальности'].astype('category')
            all_form2["Трудоустройство по специальности"].cat.set_categories(["нет", "да"], inplace=True)

            out_svod_all_form2 = all_form2.pivot_table(index=['Специальность'],
                                                       values=['Количество'],
                                                       columns=['Трудоустройство', 'Целевой договор'],
                                                       aggfunc={'Количество': sum},
                                                       margins=True)

            out_svod_all_form2.fillna(0,inplace=True)
            out_svod_all_form2 = out_svod_all_form2.applymap(int)

            out_svod_all_form2.rename(index={'All': 'Итого'}, columns={'All': 'Итого'}, inplace=True)
        else:
            out_svod_all_form2 = pd.DataFrame(columns=['нет данных'])


        with pd.ExcelWriter(f'{path_to_end_folder_opk}/Общий список и сводная таблица по форме 2 от {current_time}.xlsx',
                            engine='openpyxl') as writer:
            all_form2.to_excel(writer, sheet_name='Общий список',index=False)
            out_svod_all_form2.to_excel(writer,sheet_name='Сводная таблица')


        error_df.to_excel(f'{path_to_end_folder_opk}/Ошибки ОПК от {current_time}.xlsx', index=False)
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Закройте открытые файлы Excel {e.args}')
    except:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'При обработке файла {name_file} возникла ошибка !!!\n'
                             f'Проверьте файл на соответствие шаблону.')
    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                                'Данные успешно обработаны.')


"""
Функции для нахождения разницы между 2 таблицами
"""


def select_first_diffrence():
    """
    Функция для файла с данными
    :return: Путь к файлу с данными
    """
    global data_first_diffrence
    # Получаем путь к файлу
    data_first_diffrence = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_second_diffrence():
    """
    Функция для файла с данными
    :return: Путь к файлу с данными
    """
    global data_second_diffrence
    # Получаем путь к файлу
    data_second_diffrence = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_diffrence():
    """
    Функия для выбора папки.Определенно вот это когда нибудь я перепишу на ООП
    :return:
    """
    global path_to_end_folder_diffrence
    path_to_end_folder_diffrence = filedialog.askdirectory()


def abs_diff(first_value, second_value):
    """
    Функция для подсчета абсолютной разницы между 2 значениями
    """
    try:
        return abs(float(first_value) - float(second_value))
    except:
        return None


def percent_diff(first_value, second_value):
    """
    функция для подсчета относительной разницы значений
    """
    try:
        # округляем до трех
        value = round(float(second_value) / float(first_value), 4) * 100
        return value
    except:
        return None


def change_perc_diff(first_value, second_value):
    """
    функция для подсчета процентного ихменения значений
    """
    try:
        value = (float(second_value) - float(first_value)) / float(first_value)
        return round(value, 4) * 100
    except:
        return None


def processing_diffrence():
    """
    Функция для вычисления разницы между двумя таблицами
    """
    # загружаем датафреймы
    try:
        dif_first_sheet_name = entry_first_sheet_name_diffrence.get()
        dif_second_sheet_name = entry_second_sheet_name_diffrence.get()

        df1 = pd.read_excel(data_first_diffrence, sheet_name=dif_first_sheet_name, dtype=str)
        df2 = pd.read_excel(data_second_diffrence, sheet_name=dif_second_sheet_name, dtype=str)

        # проверяем на соответсвие размеров
        if df1.shape != df2.shape:
            raise ShapeDiffierence

        # Проверям на соответсвие колонок
        if list(df1.columns) != list(df2.columns):
            diff_columns = set(df1.columns).difference(set(df2.columns))  # получаем отличающиеся элементы
            raise ColumnsDifference

        df_cols = df1.compare(df2,
                              result_names=('Первая таблица', 'Вторая таблица'))  # датафрейм с разницей по колонкам
        df_cols.index = list(
            map(lambda x: x + 2, df_cols.index))  # добавляем к индексу +2 чтобы соответствовать нумерации в экселе
        df_cols.index.name = '№ строки'  # переименовываем индекс

        df_rows = df1.compare(df2, align_axis=0,
                              result_names=('Первая таблица', 'Вторая таблица'))  # датафрейм с разницей по строкам
        lst_mul_ind = list(map(lambda x: (x[0] + 2, x[1]),
                               df_rows.index))  # добавляем к индексу +2 чтобы соответствовать нумерации в экселе
        index = pd.MultiIndex.from_tuples(lst_mul_ind, names=['№ строки', 'Таблица'])  # создаем мультиндекс
        df_rows.index = index

        # Создаем датафрейм с подсчетом разниц
        df_diff_cols = df_cols.copy()

        # получаем список колонок первого уровня
        temp_first_level_column = list(map(lambda x: x[0], df_diff_cols.columns))
        first_level_column = []
        [first_level_column.append(value) for value in temp_first_level_column if value not in first_level_column]

        # Добавляем колонки с абсолютной и относительной разницей
        count_columns = 2
        for name_column in first_level_column:
            # высчитываем абсолютную разницу
            df_diff_cols.insert(count_columns, (name_column, 'Разница между первым и вторым значением'),
                                df_diff_cols.apply(lambda x: abs_diff(x[name_column]['Первая таблица'],
                                                                      x[name_column]['Вторая таблица']), axis=1))

            # высчитываем отношение второго значения от первого
            df_diff_cols.insert(count_columns + 1, (name_column, '% второго от первого значения'),
                                df_diff_cols.apply(lambda x: percent_diff(x[name_column]['Первая таблица'],
                                                                          x[name_column]['Вторая таблица']), axis=1))

            # высчитываем процентное изменение
            df_diff_cols.insert(count_columns + 2, (name_column, 'Изменение в процентах'),
                                df_diff_cols.apply(lambda x: change_perc_diff(x[name_column]['Первая таблица'],
                                                                              x[name_column]['Вторая таблица']),
                                                   axis=1))

            count_columns += 5

        # записываем
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # делаем так чтобы записать на разные листы
        with pd.ExcelWriter(f'{path_to_end_folder_diffrence}/Разница между 2 таблицами {current_time}.xlsx') as writer:
            df_cols.to_excel(writer, sheet_name='По колонкам')
            df_rows.to_excel(writer, sheet_name='По строкам')
            df_diff_cols.to_excel(writer, sheet_name='Значение разницы')
    except ShapeDiffierence:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Не совпадают размеры таблиц, В первой таблице {df1.shape[0]}-стр. и {df1.shape[1]}-кол.\n'
                             f'Во второй таблице {df2.shape[0]}-стр. и {df2.shape[1]}-кол.')

    except ColumnsDifference:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Названия колонок в сравниваемых таблицах отличаются\n'
                             f'Колонок:{diff_columns}  нет во второй таблице !!!\n'
                             f'Сделайте названия колонок одинаковыми.')

    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except ValueError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'В файлах нет листа с таким названием!\n'
                             f'Проверьте написание названия листа')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    # except:
    #     logging.exception('AN ERROR HAS OCCURRED')
    #     messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.34',
    #                          'Возникла ошибка!!! Подробности ошибки в файле error.log')
    else:
        messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4',
                            'Таблицы успешно обработаны')


if __name__ == '__main__':
    window = Tk()
    window.title('Кассандра Подсчет данных по трудоустройству выпускников ver 3.4')
    window.geometry('750x860')
    window.resizable(False, False)

    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_employment = ttk.Frame(tab_control)
    tab_control.add(tab_employment, text='Подсчет по специальностям/профессиям')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_employment,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           'Трудоустройство выпускников. Подсчет по специальностям/профессиям')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_employment,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data = Button(tab_employment, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                             command=select_folder_data
                             )
    btn_choose_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_employment, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_employment, text='3) Обработать данные', font=('Arial Bold', 20),
                                  command=processing_data_employment
                                  )
    btn_proccessing_data.grid(column=0, row=4, padx=10, pady=10)

    """
    Вкладка для обработки формы №15
    """
    # Создаем вкладку обработки данных для Приложения 6
    tab_employment_modern = ttk.Frame(tab_control)
    tab_control.add(tab_employment_modern, text='Подсчет форм №15')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello_modern = Label(tab_employment_modern,
                             text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                                  'Трудоустройство выпускников. Подсчет по специальностям/профессиям Форма №15')
    lbl_hello_modern.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_modern = resource_path('logo.png')

    img_modern = PhotoImage(file=path_to_img_modern)
    Label(tab_employment_modern,
          image=img_modern
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data_modern = Button(tab_employment_modern, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                                    command=select_folder_data
                                    )
    btn_choose_data_modern.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_modern = Button(tab_employment_modern, text='2) Выберите конечную папку',
                                          font=('Arial Bold', 20),
                                          command=select_end_folder
                                          )
    btn_choose_end_folder_modern.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_data_modern = Button(tab_employment_modern, text='3) Обработать данные', font=('Arial Bold', 20),
                                         command=processing_data_employment_modern
                                         )
    btn_proccessing_data_modern.grid(column=0, row=4, padx=10, pady=10)

    """
    Вкладка для обработки отчетов центров карьеры
    """
    # Создаем вкладку обработки отчетов центров карьеры
    tab_ck_employment = ttk.Frame(tab_control)
    tab_control.add(tab_ck_employment, text='Отчет ЦК')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку
    # Создаем метку для описания назначения программы
    lbl_hello_ck = Label(tab_ck_employment,
                         text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                              'Обработка данных центров карьеры по трудоустроенным выпускникам')
    lbl_hello_ck.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_ck = resource_path('logo.png')

    img_ck = PhotoImage(file=path_to_img_ck)
    Label(tab_ck_employment,
          image=img_ck
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_ck_data = Button(tab_ck_employment, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                                command=select_folder_data_ck
                                )
    btn_choose_ck_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_ck_folder = Button(tab_ck_employment, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                      command=select_end_folder_ck
                                      )
    btn_choose_end_ck_folder.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_ck_data = Button(tab_ck_employment, text='3) Обработать данные', font=('Arial Bold', 20),
                                     command=processing_data_ck_employment
                                     )
    btn_proccessing_ck_data.grid(column=0, row=4, padx=10, pady=10)

    """
    Подсчет данных по трудоустройству ОПК
    """
    # Создаем вкладку обработки отчетов центров карьеры
    tab_opk_employment = ttk.Frame(tab_control)
    tab_control.add(tab_opk_employment, text='Отчет ОПК с отраслями')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку
    # Создаем метку для описания назначения программы
    lbl_hello_opk = Label(tab_opk_employment,
                          text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                               'Обработка данных по трудоустройству ОПК (по отраслям)\n'
                               'В обрабатываемых файлах должны быть листы Форма 1 и Форма 2,\n'
                               'В Форме 1 должно быть 80 колонок включая 2 колонки проверки\n'
                               ',внизу после окончания таблицы должна быть пустая строка.\n'
                               ' На 9 строке должна быть строка с номерами колонок.'
                               'В форме 2 должно быть 10 колонок')
    lbl_hello_opk.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_opk = resource_path('logo.png')

    img_opk = PhotoImage(file=path_to_img_opk)
    Label(tab_opk_employment,
          image=img_opk
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_opk_data = Button(tab_opk_employment, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                                 command=select_folder_data_opk
                                 )
    btn_choose_opk_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_opk_folder = Button(tab_opk_employment, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder_opk
                                       )
    btn_choose_end_opk_folder.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_opk_data = Button(tab_opk_employment, text='3) Обработать данные', font=('Arial Bold', 20),
                                      command=processing_data_opk_employment
                                      )
    btn_proccessing_opk_data.grid(column=0, row=4, padx=10, pady=10)

    """
    Разница двух таблиц
    """
    tab_diffrence = ttk.Frame(tab_control)
    tab_control.add(tab_diffrence, text='Разница 2 таблиц')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку разница 2 двух таблиц
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_diffrence,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           'Количество строк и колонок в таблицах должно совпадать\n'
                           'Названия колонок в таблицах должны совпадать'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_com = resource_path('logo.png')
    img_diffrence = PhotoImage(file=path_com)
    Label(tab_diffrence,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_diffrence = LabelFrame(tab_diffrence, text='Подготовка')
    frame_data_for_diffrence.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать  первый файл с данными
    btn_data_first_diffrence = Button(frame_data_for_diffrence, text='1) Выберите файл с первой таблицей',
                                      font=('Arial Bold', 10),
                                      command=select_first_diffrence
                                      )
    btn_data_first_diffrence.grid(column=0, row=3, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_first_sheet_name_diffrence = StringVar()
    # Описание поля
    label_first_sheet_name_diffrence = Label(frame_data_for_diffrence,
                                             text='2) Введите название листа, где находится первая таблица')
    label_first_sheet_name_diffrence.grid(column=0, row=4, padx=10, pady=10)
    # поле ввода имени листа
    first_sheet_name_entry_diffrence = Entry(frame_data_for_diffrence, textvariable=entry_first_sheet_name_diffrence,
                                             width=30)
    first_sheet_name_entry_diffrence.grid(column=0, row=5, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку Выбрать  второй файл с данными
    btn_data_second_diffrence = Button(frame_data_for_diffrence, text='3) Выберите файл со второй таблицей',
                                       font=('Arial Bold', 10),
                                       command=select_second_diffrence
                                       )
    btn_data_second_diffrence.grid(column=0, row=6, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_second_sheet_name_diffrence = StringVar()
    # Описание поля
    label_second_sheet_name_diffrence = Label(frame_data_for_diffrence,
                                              text='4) Введите название листа, где находится вторая таблица')
    label_second_sheet_name_diffrence.grid(column=0, row=7, padx=10, pady=10)
    # поле ввода
    second__sheet_name_entry_diffrence = Entry(frame_data_for_diffrence, textvariable=entry_second_sheet_name_diffrence,
                                               width=30)
    second__sheet_name_entry_diffrence.grid(column=0, row=8, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку выбора папки куда будет генерироваьться файл
    btn_select_end_diffrence = Button(frame_data_for_diffrence, text='5) Выберите конечную папку',
                                      font=('Arial Bold', 10),
                                      command=select_end_folder_diffrence
                                      )
    btn_select_end_diffrence.grid(column=0, row=10, padx=10, pady=10)

    # Создаем кнопку Обработать данные
    btn_data_do_diffrence = Button(tab_diffrence, text='6) Обработать таблицы', font=('Arial Bold', 20),
                                   command=processing_diffrence
                                   )
    btn_data_do_diffrence.grid(column=0, row=11, padx=10, pady=10)

    window.mainloop()
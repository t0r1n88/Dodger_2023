# -*- coding: utf-8 -*-
"""
Функции для проверки данных
"""
import struct

from cass_support_functions import *  # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np

pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re


def check_data(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0


def check_data_note(cell):
    if cell is np.nan:
        return 'Не заполнено'
    return str(cell)


def find_header_lenght(wb: openpyxl.Workbook, name_sheet: str, target_column: int, target_value: str) -> int:
    """
    Функция для поиска размера заголовка
    """
    header_lenght = None

    for row in wb[name_sheet].iter_rows(min_row=1, min_col=target_column, max_col=target_column):
        cell_value = row[0].value
        if cell_value == target_value:
            header_lenght = row[0].row - 1  # Отнимаем еденицу потому что будем использовать эту строку в качестве заголовка
            break

    return header_lenght

def replace_empty_with_word(value):
    if isinstance(value, str) and value.strip() == '':
        return 'Неправильно'
    else:
        return value




def base_check_file(file: str, path_folder_data: str, requred_columns_first_sheet: list,
                    requred_columns_second_sheet: list,requred_columns_prof_sheet:list):
    """
    Функция для базовой проверки файла мониторига занятости выпускников загружаемого в СССР. Расширение,наличие нужных листов
    """
    # Создаем словарь для базовой проверки файла (расширение, наличие листов, наличие колонок)
    """
    {Название листа:{Количество строк заголовка:int,'Обязательные колонки':список колонок,'Текст ошибки':'Описание ошибки'}}
    """
    checked_required_sheet = {'Выпуск-СПО': {'Количество строк заголовка': 4,
                                             'Название листа': 'Выпуск-СПО',
                                             'Реальное название листа': None,
                                             'Обязательные колонки': requred_columns_first_sheet,
                                             'Не найден лист': 'В файле не найден лист с названием Выпуск-СПО',
                                             'Нет колонок': 'На листе Выпуск-СПО не найдены колонки:'},
                              'Выпуск-Целевое': {'Количество строк заголовка': 3,
                                                 'Название листа': 'Выпуск-Целевое',
                                                 'Реальное название листа': None,
                                                 'Обязательные колонки': requred_columns_second_sheet,
                                                 'Не найден лист': 'В файле не найден лист с названием Выпуск-Целевое',
                                                 'Нет колонок': 'На листе Выпуск-Целевое не найдены колонки:'},
                              'Выпуск-Профессионалитет': {'Количество строк заголовка': None,
                                                 'Название листа': 'Профессионалитет',
                                                 'Реальное название листа': None,
                                                 'Обязательные колонки': requred_columns_prof_sheet,
                                                 'Не найден лист': 'В файле не найден лист с названием Профессионалитет',
                                                 'Нет колонок': 'На листе Выпуск-Профессионалитет не найдены колонки:'}
                              }

    _error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    if not file.startswith('~$') and not file.endswith('.xlsx'):
        # проверка файла на расширение xlsx
        name_file = file.split('.xls')[0]
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                            'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)
        return _error_df, checked_required_sheet
    if not file.startswith('~$') and file.endswith('.xlsx'):
        # Проверка файла на наличие требуемых листов
        name_file = file.split('.xlsx')[0]
        # получаем название первого листа
        temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}')
        lst_temp_sheets = temp_wb.sheetnames  # получаем листы в файле

        # Создаем переменные для названий листов
        name_spo_sheet = None
        name_target_sheet = None
        name_prof_sheet = None

        # Ищем лист содержащий слово Выпуск-СПО
        for sheet in lst_temp_sheets:
            if checked_required_sheet['Выпуск-СПО']['Название листа'] in sheet:
                name_spo_sheet = sheet

        if not name_spo_sheet:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Не найден лист содержащий название Выпуск-СПО']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)
            return _error_df, checked_required_sheet

        # Ищем лист содержащий слово Выпуск-Целевое
        for sheet in lst_temp_sheets:
            if checked_required_sheet['Выпуск-Целевое']['Название листа'] in sheet:
                name_target_sheet = sheet

        if not name_target_sheet:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Не найден лист содержащий название Выпуск-Целевое']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)
            return _error_df, checked_required_sheet

        # Ищем название листа с названием профессионалитет
        for sheet in lst_temp_sheets:

            if checked_required_sheet['Выпуск-Профессионалитет']['Название листа'] in sheet:
                name_prof_sheet = sheet


        # Ищем размер заголовка первого листа
        result_find_first_header = find_header_lenght(temp_wb, name_spo_sheet, 3,'1.1')
        if result_find_first_header:
            checked_required_sheet['Выпуск-СПО']['Количество строк заголовка'] = result_find_first_header

        # Ищем размер заголовка второго листа
        result_find_second_header = find_header_lenght(temp_wb, name_target_sheet, 2,'1')
        if result_find_second_header:
            checked_required_sheet['Выпуск-Целевое']['Количество строк заголовка'] = result_find_second_header

        # Ищем размер заголовка третьего листа
        if name_prof_sheet:
            result_find_third_header = find_header_lenght(temp_wb, name_prof_sheet, 2,'1')
            if result_find_third_header:
                checked_required_sheet['Выпуск-Профессионалитет']['Количество строк заголовка'] = result_find_third_header



        temp_wb.close()  # закрываем файл

        # проверяем наличие требуемых колонок на первом листе
        temp_df = pd.read_excel(f'{path_folder_data}/{file}', sheet_name=name_spo_sheet,
                                skiprows=checked_required_sheet['Выпуск-СПО']['Количество строк заголовка'])

        temp_df.columns = list(map(str, temp_df.columns))  # делаем названия колонок строковыми
        # находим разницу в колонках
        diff_cols = set(checked_required_sheet['Выпуск-СПО']['Обязательные колонки']).difference(set(temp_df.columns))
        if len(diff_cols) != 0:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                f'{checked_required_sheet["Выпуск-СПО"]["Нет колонок"]} {";".join(diff_cols)}.'
                                                f' Строка с номерами колонок должны быть на строке {checked_required_sheet["Выпуск-СПО"]["Количество строк заголовка"] + 1} в исходном файле']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)
            return _error_df, checked_required_sheet

        # Провереряем наличие требуемых колонок на втором листе
        temp_df = pd.read_excel(f'{path_folder_data}/{file}', sheet_name=name_target_sheet,
                                skiprows=checked_required_sheet['Выпуск-Целевое']['Количество строк заголовка'])

        temp_df.columns = list(map(str, temp_df.columns))  # делаем названия колонок строковыми
        # находим разницу в колонках
        diff_cols = set(checked_required_sheet['Выпуск-Целевое']['Обязательные колонки']).difference(
            set(temp_df.columns))
        if len(diff_cols) != 0:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                f'{checked_required_sheet["Выпуск-Целевое"]["Нет колонок"]} {";".join(diff_cols)}.'
                                                f' Строка с номерами колонок должны быть на строке {checked_required_sheet["Выпуск-Целевое"]["Количество строк заголовка"] + 1} в исходном файле']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)
            return _error_df, checked_required_sheet

        # получаем реальные названия листов
        checked_required_sheet['Выпуск-СПО']['Реальное название листа'] = name_spo_sheet
        checked_required_sheet['Выпуск-Целевое']['Реальное название листа'] = name_target_sheet
        checked_required_sheet['Выпуск-Профессионалитет']['Реальное название листа'] = name_prof_sheet


    return _error_df, checked_required_sheet


def check_sameness_column(checked_lst: list, check_range: int, begin_border: int, quantity_check_value: int,
                          tup_correct: tuple, correction: int,
                          name_file=None, name_column=None):
    """
    Функция для проверки заполнен ли определенный диапазон одинаковыми значениями
    checked_lst : список значений который нужно проверить на однородность в каждом диапазоне
    check_range : сколько значений в  првоеряемом диапазоне
    quantity_check_value : количество проверяемых диапазонов
    tup_correct : кортеж  где нулевой элемент это первая строка диапазона в экселе а первый элемент последняя строка в диапазоне
    это нужно чтобы точно указывать где искать ошибку в файле Excel
    correction : дополнительная поправка
    name_file : имя обрабатываемого файла Excel

    """
    # датафрейм для ошибок
    _error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    offset = 0  # сдвиг

    for i in range(quantity_check_value):
        temp_set = set(
            [value for value in checked_lst[begin_border:begin_border + check_range]])
        if len(temp_set) != 1:
            temp_error_df = pd.DataFrame(data=[[name_file,
                                                f'Диапазон строк {begin_border + first_correct + offset} - {begin_border + second_correct + offset}',
                                                f'В колонке {name_column} в указанном диапазоне обнаружены отличающиеся значения']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки', ])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)

        begin_border += check_range  # сдвигаем проверяемый диапазон в списке
        offset += correction  # добавляем поправку

    return _error_df


def check_blankness_column(checked_lst: list, check_range: int, begin_border: int, quantity_check_value: int,
                           tup_correct: tuple, correction: int,
                           name_file=None, name_column=None):
    """
    Функция для проверки есть ли в определенном диапазоне пустые значения
    checked_lst : список значений который нужно проверить на однородность в каждом диапазоне
    check_range : сколько значений в  првоеряемом диапазоне
    quantity_check_value : количество проверяемых диапазонов
    tup_correct : кортеж  где нулевой элемент это первая строка диапазона в экселе а первый элемент последняя строка в диапазоне
    это нужно чтобы точно указывать где искать ошибку в файле Excel
    correction : дополнительная поправка
    name_file : имя обрабатываемого файла Excel

    """
    # датафрейм для ошибок
    _error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    offset = 0  # сдвиг

    for i in range(quantity_check_value):
        temp_set = set(
            [value for value in checked_lst[begin_border:begin_border + check_range]])
        if np.nan in temp_set or ' ' in temp_set:
            temp_error_df = pd.DataFrame(data=[[name_file,
                                                f'Диапазон строк {begin_border + first_correct + offset} - {begin_border + second_correct + offset}',
                                                f'В колонке {name_column} в указанном диапазоне обнаружены незаполненные ячейки']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки', ])
            _error_df = pd.concat([_error_df, temp_error_df], axis=0, ignore_index=True)

        begin_border += check_range  # сдвигаем проверяемый диапазон в списке
        offset += correction  # добавляем поправку

    return _error_df


def check_first_error_temp(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки гр. 09 и гр. 10 < гр. 08
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]

    # Проводим проверку
    df['Результат'] = (df['08'] >= df['09']) & (df['08'] >= df['10'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 09 <= гр. 08  или гр. 10 <= гр. 08'
    return temp_error_df


def check_second_error_temp(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 07= гр.08 + сумма(с гр.11 по гр.32))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # конвертируем в инт
    all_sum_cols = list(df)
    # удаляем колонки 07, 09, 10
    all_sum_cols.remove('07')
    all_sum_cols.remove('09')
    all_sum_cols.remove('10')
    # получаем сумму колонок 08, 11:32
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['07'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 07 = гр.08 + сумма(с гр.11 по гр.32)'
    return temp_error_df


def check_third_error_temp(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр. 02 + стр. 04
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '06', ])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['Сумма'] = foo_df['02'] + foo_df['04']
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр. 02 + стр. 04 '
    return temp_error_df


"""
Проверки
"""


def check_first_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр 03 <= стр 02 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '03'])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['03'] = df.iloc[2, :]
    foo_df['Результат'] = foo_df['03'] <= foo_df['02']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 03 <= стр. 02 '
    return temp_error_df


def check_second_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр.02 и стр.04 и стр.05 <= стр.01
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '05', '01'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['05'] = df.iloc[4, :]

    foo_df['Результат'] = (foo_df['01'] >= foo_df['02']) & (foo_df['01'] >= foo_df['04']) & (
            foo_df['01'] >= foo_df['05'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр.02<= стр.01 или стр.04<= стр.01 или стр.05<= стр.01 '
    return temp_error_df

def check_third_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 05= сумма(с гр.06 по гр.27))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    all_sum_cols = list(df)  # получаем список колонок
    # удаляем колонку 05 с общей суммой
    all_sum_cols.remove('05')
    # получаем сумму колонок 06:27
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['05'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 05 = сумма(с гр.06 по гр.27)'
    return temp_error_df

def form_two_check_third_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 05= сумма(с гр.06 по гр.27))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    all_sum_cols = list(df)  # получаем список колонок
    remove_cols = ['гр.04','гр.06','гр.07','гр.23','гр.24']
    # удаляем лишние колонки
    all_sum_cols = [value for value in all_sum_cols if value not in remove_cols]
    # получаем сумму колонок 06:27
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['гр.04'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 04 = сумма(с гр.05 по гр.29) кроме гр.06,гр.07,гр.23,гр.24'
    return temp_error_df


"""
Функции проверки для формы 2
"""


def form_two_check_fourth_error(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр. 02 + стр. 04
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '06', ])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['Сумма'] = foo_df['02'] + foo_df['04']
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр. 02 + стр. 04 '
    return temp_error_df


def form_two_check_fifth_error(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13

    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['06', '07', '08', '09', '10', '11', '12', '13'])

    # Добавляем данные в датафрейм
    foo_df['06'] = df.iloc[5, :]
    foo_df['07'] = df.iloc[6, :]
    foo_df['08'] = df.iloc[7, :]
    foo_df['09'] = df.iloc[8, :]
    foo_df['10'] = df.iloc[9, :]
    foo_df['11'] = df.iloc[10, :]
    foo_df['12'] = df.iloc[11, :]
    foo_df['13'] = df.iloc[12, :]

    sum_col = ['07', '08', '09', '10', '11', '12', '13']
    foo_df['Сумма'] = foo_df[sum_col].sum(axis=1)
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13 '
    return temp_error_df


def form_two_check_sixth_error(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
       Функция для проверки правильности введеденных данных
       стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
       :param foo_df: копия датафрейма с данными из файла поо
       :return:датафрейм с ошибками
       """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['05', '06', '14'])

    # Добавляем данные в датафрейм
    foo_df['05'] = df.iloc[4, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['14'] = df.iloc[13, :]

    foo_df['Результат'] = (foo_df['14'] <= foo_df['05']) & (foo_df['14'] <= foo_df['06'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df[
        'index'].tolist()  # делаем список, прибавляем для того чтобы номера строк совпадали с строками в файле
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк  {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 14<=стр. 06, стр. 14<=стр 05'
    return temp_error_df


def form_three_check_third_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введенных данных Форма 3 ожидаемый выпуск
    (гр.05 = гр.06 + гр. 10-12 + гр. 15-20)
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    all_sum_cols = list(df)  # получаем список колонок
    # удаляем  лишние колонки при подсчете
    lst_remove_columns = ['05', '07', '08', '09', '13', '14', '21', '22']
    for rem_column in lst_remove_columns:
        all_sum_cols.remove(rem_column)

    # получаем сумму колонок
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['05'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр.05 = гр.06 + гр. 10-12 + гр. 15-20'
    return temp_error_df


def form_three_check_fourth_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введенных данных Форма 3 ожидаемый выпуск
    гр. 07+ гр. 08 + гр. 09 <= гр. 06
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]

    # получаем сумму колонок
    df['Сумма'] = df[['07', '08', '09', ]].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['06'] >= df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 07+ гр. 08 + гр. 09 <= гр. 06'
    return temp_error_df


def form_three_check_fifth_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введенных данных Форма 3 ожидаемый выпуск
    гр. 13 + гр. 14 <= 12
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # получаем сумму колонок
    df['Сумма'] = df[['13', '14']].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['12'] >= df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 13 + гр. 14 <= 12'
    return temp_error_df


def form_three_check_sixteen_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введенных данных Форма 3 ожидаемый выпуск
    гр. 21 + гр. 22 <= 20
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # получаем сумму колонок
    df['Сумма'] = df[['21', '22']].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['20'] >= df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 21 + гр. 22 <= 20'
    return temp_error_df


def check_fourth_error(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13

    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['06', '07', '08', '09', '10', '11', '12', '13'])

    # Добавляем данные в датафрейм
    foo_df['06'] = df.iloc[5, :]
    foo_df['07'] = df.iloc[6, :]
    foo_df['08'] = df.iloc[7, :]
    foo_df['09'] = df.iloc[8, :]
    foo_df['10'] = df.iloc[9, :]
    foo_df['11'] = df.iloc[10, :]
    foo_df['12'] = df.iloc[11, :]
    foo_df['13'] = df.iloc[12, :]

    sum_col = ['07', '08', '09', '10', '11', '12', '13']
    foo_df['Сумма'] = foo_df[sum_col].sum(axis=1)
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13 '
    return temp_error_df


def check_fifth_error(df: pd.DataFrame, name_file, border, tup_correct, correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['05', '06', '14'])

    # Добавляем данные в датафрейм
    foo_df['05'] = df.iloc[4, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['14'] = df.iloc[13, :]

    foo_df['Результат'] = (foo_df['14'] <= foo_df['05']) & (foo_df['14'] <= foo_df['06'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df[
        'index'].tolist()  # делаем список, прибавляем для того чтобы номера строк совпадали с строками в файле
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк  {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 14<=стр. 06, стр. 14<=стр 05'
    return temp_error_df


def check_sixth_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр 03 <= стр 02 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '03'])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['03'] = df.iloc[2, :]
    foo_df['Результат'] = foo_df['03'] <= foo_df['02']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 03 <= стр. 02 '
    return temp_error_df


def check_seventh_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр.02 и стр.04 и стр.05 < стр.01
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '05', '01'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['05'] = df.iloc[4, :]

    foo_df['Результат'] = (foo_df['01'] >= foo_df['02']) & (foo_df['01'] >= foo_df['04']) & (
            foo_df['01'] >= foo_df['05'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda
                x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct + correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр.02<= стр.01 или стр.04<= стр.01 или стр.05<= стр.01 '
    return temp_error_df


def check_error_form_one(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 3:26]  # получаем часть с числами
    df = df.applymap(check_data)  # заполняем пустые ячейки нулями

    # получаем количество датафреймов
    quantity = df.shape[0] // 5
    # счетчик для обработанных строк
    border = 0
    correction = 0  # поправка для учета строки c проверками
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 5, :]

        # Проводим проверку стр.03 <= стр.02
        first_error_df = check_first_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        second_error_df = check_second_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)
        #
        # Проводим проверку гр. 05=сумма(с гр.06 по гр.28)
        third_error_df = check_third_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 5
    # Возвращаем датафрейм с ошибками
    return error_df


def create_check_tables_form_one(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 5'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 12'))
            temp_code_df['Неформальная занятость (теневой сектор экономики)'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 13'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df[
                'Иные причины нахождения под риском нетрудоустройства (включая отсутствие проводимой с выпускниками работы по содействию их занятости)'] = \
            temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 20'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 21'))
            temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 22'))
            temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 23'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 25'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))

            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении'
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb


def check_error_form_two(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 3:29]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 14
    # счетчик для обработанных строк
    border = 0
    correction = 0  # поправка для учета строки 15 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 14, :]
        # Проводим проверку стр 03 <= стр 02
        first_error_df = check_first_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 <= стр.01
        second_error_df = check_second_error(temp_df.copy(), name_file, border, tup_correct, correction)

        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)
        # Проводим проверку гр. 04=сумма(с гр.05 по гр.29) кроме гр.06,гр.07,гр.23,гр.24
        third_error_df = form_two_check_third_error(temp_df.copy(), name_file, tup_correct)

        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр 06 = стр 02 + стр 04
        fourth_error_df = form_two_check_fourth_error(temp_df.copy(), name_file, border, tup_correct, correction)

        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр.13
        fifth_error_df = form_two_check_fifth_error(temp_df.copy(), name_file, border, tup_correct, correction)

        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        sixth_error_df = form_two_check_sixth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)


        # прибавляем border
        border += 14
        correction += 1
    # Возвращаем датафрейм с ошибками
    return error_df


def create_check_tables_form_two(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Суммарный выпуск'] = temp_code_df[0].apply(lambda x: x.get('Колонка 4'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 5'))
            temp_code_df['из них (из графы 05): продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df[
                'из них (из графы 05): трудоустроены по полученной профессии, специальности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df['Индиви-дуальные предприни-матели '] = temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df['Самозанятые (перешедшие на специальный налоговый режим - налог на профессио-нальный доход)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df[
                'Проходят службу в армии по призыву'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Проходят службу в армии по контракту, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии РФ, органах принудительного исполнения РФ '] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 11'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 12'))
            temp_code_df[
                'Находятся в отпуске по уходу за ребенком'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df[
                'Неформальная занятость (теневой сектор экономики)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице)'] = \
            temp_code_df[0].apply(
                lambda x: x.get('Колонка 15'))
            temp_code_df['Не имеют мотивации к трудоустройству и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот'] = temp_code_df[0].apply(lambda x: x.get('Колонка 16'))
            temp_code_df['Отсутствует спрос на специалистов в регионе, находятся в поиске работы'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 17'))
            temp_code_df[
                'Смерть, тяжелое состояние здоровья'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Находятся под следствием, отбывают наказание '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Переезд за пределы Российской Федерации (кроме переезда в иные регионы)'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 20'))
            temp_code_df['Ухаживают за больными родственниками (иные семейные обстоятельства)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df['будут трудоустроены (в соответствии с трудовым законодательством, законодательством об обязательном пенсионном страховании)'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 22'))
            temp_code_df['из них (из графы 22): продолжат обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df['из них (из графы 22): будут трудоустроены по полученной профессии, специальности'] = temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df[
                'будут осуществлять предприни-мательскую деятельность'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 25'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['будут в армии по контракту, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии РФ, органах принудительного исполнения РФ'] = temp_code_df[0].apply(lambda x: x.get('Колонка 28'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 29'))

            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении',
                   'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
                ,
                   'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
                   'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
                   'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
                   'Строка 12': 'расстройствами аутистического спектра',
                   'Строка 13': 'с инвалидностью вследствие  других причин',
                   'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb


def check_error_form_three(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 3:21]  # получаем часть с числами
    df = df.applymap(check_data)  # заполняем пустые ячейки нулями
    # получаем количество датафреймов
    quantity = df.shape[0] // 5
    # счетчик для обработанных строк
    border = 0
    correction = 0  # поправка для учета строки c проверками
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 5, :]
        # Проводим проверку стр.03 <= стр.02
        first_error_df = check_first_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        second_error_df = check_second_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)
        #
        # # Проводим проверку (гр.05 = гр.06 + гр. 10-12 + гр. 15-20)
        third_error_df = form_three_check_third_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07+ гр. 08 + гр. 09 <= гр. 06
        fourth_error_df = form_three_check_fourth_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 13 + гр. 14 <= 12
        fifth_error_df = form_three_check_fifth_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 21 + гр. 22 <= 20
        sixteen_error_df = form_three_check_sixteen_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, sixteen_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 5
    # Возвращаем датафрейм с ошибками
    return error_df


def create_check_tables_form_three(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Ожидаемый выпуск Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 5'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df[
                'совмещают обучение с трудоустройством по специальности, в том числе с переводом на индивидуальный учебный план'] = \
            temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'проходят оплачиваемую практику по специальности с заключением срочного трудового договора '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df[
                'трудоустроены в учебно-производственных комплексах, созданных на базе образовательных организаций, по специальности'] = \
            temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df['Индивидуальные предприни-матели '] = temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df['Планируют трудоустройство'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 12'))
            temp_code_df['трудоустроятся на предприятиях, в которых была пройдена практика'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 13'))
            temp_code_df[
                'трудоустроятся на других предприятиях, являющихся партнерами образовательной организации'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Планируют осуществлять предпринимательскую деятельность в форме индивидуального предпринимателя'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Планируют зарегистрироваться в качестве самозанятых'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Подлежат призыву в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df[
                'Планируют поступить в армию на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            temp_code_df[0].apply(
                lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Планируют продолжать обучение'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df[
                'Находятся под риском нетрудоустройства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 20'))
            temp_code_df[
                'Планируют переезд за пределы Российской Федерации(кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            temp_code_df[0].apply(
                lambda x: x.get('Колонка 21'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 22'))
            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении'
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb


def create_check_tables_may_2025(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.reset_index()
            temp_code_df.columns = ['Наименование файла','2', '3', '3.1', '3.2', '3.3',
                                 '4', '4.1', '4.2', '4.3', '5', '6', '7', '8', '9', '10', '11', '12',
                                 '13', '14', '15', '16', '17', '18']
            for r in dataframe_to_rows(temp_code_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40

    return wb


def create_check_tables_september_2025(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.reset_index()
            temp_code_df.columns = ['Наименование файла','2', '3', '3.1', '3.2', '3.3',
                                 '4', '4.1', '4.2', '4.3', '5', '6', '7', '8', '9', '10', '11', '12',
                                 '13', '14', '15', '16', '17','18']
            for r in dataframe_to_rows(temp_code_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40

    return wb

def create_check_tables_target_may_2025(df: pd.DataFrame):
    """
    Функция для создания файла с данными по каждой специальности
    """
    df['Код'] = df['1'].apply(extract_code_nose)  # очищаем от текста в кодах
    lst_unique_code = sorted(df['Код'].unique()) # список уникальных

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(lst_unique_code):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code in lst_unique_code:
        temp_df = df[df['Код'] == code]
        temp_df.drop(columns=['Код'],inplace=True)
        for r in dataframe_to_rows(temp_df, index=False, header=True):
            wb[code].append(r)
        wb[code].column_dimensions['A'].width = 40
        wb[code].column_dimensions['B'].width = 20


    return wb



def create_check_tables_nose_target_sept_2025(df: pd.DataFrame):
    """
    Функция для создания файла с данными по каждой специальности
    """
    lst_unique_code = sorted(df['1'].unique()) # список уникальных

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(lst_unique_code):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code in lst_unique_code:
        temp_df = df[df['1'] == code]
        temp_df.drop(columns=['1'],inplace=True)
        for r in dataframe_to_rows(temp_df, index=False, header=True):
            wb[code].append(r)
        wb[code].column_dimensions['A'].width = 40
        wb[code].column_dimensions['B'].width = 20


    return wb







def check_error_nose(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 4:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    correction = 0  # поправка для учета строки 16 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.03 <= стр.02
        sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
        correction += 1
    # Возвращаем датафрейм с ошибками
    return error_df


def check_error_base_mon(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных базового мониторинга 5 строк
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 6:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    correction = 0  # поправка для учета строки 16 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.03 <= стр.02
        sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
    # Возвращаем датафрейм с ошибками
    return error_df


def extract_code(value):
    """
    Функция для извлечения кода специальности
    """
    value = str(value)
    re_code = re.compile('^\d{2}?[.]\d{2}?[.]\d{2}')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.group()
    else:
        return 'error'


def extract_code_nose(value):
    """
    Функция для извлечения кода специальности из новой формы по нозологиям
    """
    value = str(value)
    re_code = re.compile('(\d{2}?[.]\d{2}?[.]\d{2})\D')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.groups()[0]
    else:
        return 'error'





def create_check_tables(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df[
                'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df[
                'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Неформальная занятость (нелегальная)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 20'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 22'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 25'))
            temp_code_df[
                'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 28'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 31'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
            temp_code_df['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 33'))

            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении',
                   'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
                ,
                   'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
                   'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
                   'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
                   'Строка 12': 'расстройствами аутистического спектра',
                   'Строка 13': 'с инвалидностью вследствие  других причин',
                   'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
                   'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb


"""
Функции для обработки отчетов ЦК
"""


def check_horizont_all_sum_error(df: pd.DataFrame, tup_exluded_cols: tuple, name_itog_cols, name_file):
    """
    Функция для проверки горизонтальных сумм по всей строке
    сумма в колонке 05 должна быть равна сумме всех колонок за исключением 07 и 15
    """
    # датафрейм для ощибок по горизонтали
    hor_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # получаем список колонок
    all_sum_cols = list(df)
    # удаляем колонки
    for name_cols in tup_exluded_cols:
        all_sum_cols.remove(name_cols)
    # удаляем итоговую колонку
    all_sum_cols.remove(name_itog_cols)

    # получаем сумму колонок за вычетом исключаемых и итоговой колонки
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} = сумма остальных гр. за искл.{tup_exluded_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_horizont_chosen_sum_error(df: pd.DataFrame, tup_checked_cols: list, name_itog_cols, name_file):
    """
    Функция для проверки равенства одиночных или небольших групп колонок
    tup_checked_cols колонки сумму которых нужно сравнить с name_itog_cols чтобы она не превышала это значение
    """
    # Считаем проверяемые колонки
    df['Сумма'] = df[tup_checked_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] >= df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} >= сумма {tup_checked_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_vertical_chosen_sum(df: pd.DataFrame, lst_checked_rows: list, itog_row, name_file):
    """
    Функция для проверки вертикальной суммы заданных строк сумма значений в tupl_checked_row должна быть равной ил меньше чем значение
    в itog_row
    """

    # обрабаотываем список строк чтобы привести его в читаемый вид
    lst_out_rows = list(map(lambda x: x + 1, lst_checked_rows))

    # делаем значения строковыми
    lst_out_rows = list(map(str, lst_out_rows))
    # Добавляем ноль в строки
    lst_out_rows = list(map(lambda x: '0' + x, lst_out_rows))
    # обрабатываем формат выходной строки
    out_itog_row = f'0{str(itog_row + 1)}'

    # создаем временный датафрейм
    foo_df = pd.DataFrame()
    # разворачиваем строки в колонки
    for idx_row in lst_checked_rows:
        foo_df[idx_row] = df.iloc[idx_row, :]

    # добавляем итоговую колонку
    foo_df[itog_row] = df.iloc[itog_row, :]

    # суммируем
    foo_df['Сумма'] = foo_df[lst_checked_rows].sum(axis=1)
    foo_df['Результат'] = foo_df[itog_row] >= foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    error_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    # Добавляем слово колонка
    error_df['index'] = error_df['index'].apply(lambda x: 'Колонка ' + str(x))
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    temp_error_df['Строка или колонка с ошибкой'] = error_df['index']
    temp_error_df[
        'Описание ошибки'] = f'Для указанной колонки сумма в строках {lst_out_rows} превышает значением в строке {out_itog_row} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! '
    temp_error_df['Название файла'] = name_file

    return temp_error_df


def check_error_ck(df: pd.DataFrame, name_file):
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # проводим горизонтальные проверки
    # проверка на общую сумму
    first_error_ck_df = check_horizont_all_sum_error(df.copy(), ('07', '15'), '05', name_file)
    error_df = pd.concat([error_df, first_error_ck_df], axis=0, ignore_index=True)
    # проверяем небольшие группы или одиночные колонки
    second_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['07'], '06', name_file)
    error_df = pd.concat([error_df, second_error_ck_df], axis=0, ignore_index=True)

    # проверяем колонки 14 и 15
    third_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['15'], '14', name_file)
    error_df = pd.concat([error_df, third_error_ck_df], axis=0, ignore_index=True)

    # Проводим вертикальные проверки
    # Сумма овз и целевиков не должна превышать общую численность выпускников. Строки с индексом 1 и 4 должныть меньше или равны строке с индексом 0
    fourth_error_ck_df = check_vertical_chosen_sum(df.copy(), [1, 4], 0, name_file)
    error_df = pd.concat([error_df, fourth_error_ck_df], axis=0, ignore_index=True)

    # Проверяем ОВЗ
    fifth_error_ck_df = check_vertical_chosen_sum(df.copy(), [2, 3], 1, name_file)
    error_df = pd.concat([error_df, fifth_error_ck_df], axis=0, ignore_index=True)

    return error_df


"""
Проверки мониторинг Май 2025
"""
def check_first_error_may_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 2 = сумма 3+4+ остальные колонки
    """
    lst_sum = ['3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
                           '13', '14', '15', '16', '17']

    # получаем сумму колонок 3,4:17
    df['Сумма'] = df[lst_sum].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['2'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие:на листе 1. Форма сбора Колонка 2 = сумма колонок (3,4,5 по 17)'
    return temp_error_df


def check_second_error_may_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 3 >= каждой из 3.1 3.2 3.3
    """

    # Проводим проверку
    df['Результат'] = (df['3'] >= df['3.1']) & (df['3'] >= df['3.2']) & (df['3'] >= df['3.3'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: на листе 1. Форма сбора Колонка 3 >= значений в колонках  (3.1, 3.2, 3.3)'
    return temp_error_df


def check_third_error_may_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 4 >= каждой из колонок 4.1 4.2 4.3
    """

    # Проводим проверку
    df['Результат'] = (df['4'] >= df['4.1']) & (df['4'] >= df['4.2']) & (df['4'] >= df['4.3'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: на листе 1. Форма сбора Колонка 4 >= значений в колонках(4.1, 4.2, 4.3)'
    return temp_error_df



def check_fourth_error_may_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 5 = сумма остальные колонки
    """
    lst_sum = ['6', '7', '8', '9', '10', '11',
                                   '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23',
                                   '24', '25', '26', '27']

    # получаем сумму колонок
    df['Сумма'] = df[lst_sum].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['5'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие:на листе 3. Целевики Колонка 5 = сумма колонок 6:27'
    return temp_error_df








def check_error_main_may_2025(df:pd.DataFrame,name_file:str):
    """
    Точка входа для проверки ошибок на основном листе
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.loc[:, '2':'17']
    df = df.applymap(check_data)
    # Проверяем 2 = сумма 3+4+ остальные колонки
    first_error_df = check_first_error_may_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

    # Проверяем 3 больше или равно  сумма 3.1 3.2 3.3
    second_error_df = check_second_error_may_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

    # Проверяем 4 больше или равно  сумма 4.1 4.2 4.3
    third_error_df = check_third_error_may_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

    return error_df



def check_error_target_may_2025(df:pd.DataFrame,name_file:str):
    """
    Точка входа для проверки ошибок на листе целевиков
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.loc[:, '2':'27']
    df = df.applymap(check_data)
    # Проверяем 5 = сумма остальные колонки
    fourth_error_df = check_fourth_error_may_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

    return error_df

def check_dight(value:str,quantity_tuple:tuple):
    """
    Функция для проверки количества цифр в ячейке
    """
    result = re.search(r'\d+',value)
    if result:
        out_value = result.group()
        if len(out_value) in quantity_tuple:
            return 'Правильно'
        else:
            return 'Неправильно'
    else:
        return 'Неправильно'


def check_kpp(row:pd.Series):
    """
    Функция для проверки КПП
    """
    lst_row = row.tolist()
    kpp = re.search(r'\d+',lst_row[1])
    inn = re.search(r'\d+',lst_row[0])
    if kpp:
        if inn:
            # есть есть цифры и в ИНН и в КППП
            if len(inn.group()) == 10 and len(kpp.group()) == 9:
                return 'Правильно'
            else:
                return 'Неправильно'
        else:
            return 'Неправильно'
    else:
        if inn:
            if len(inn.group()) == 12:
                return 'Правильно'
            else:
                return 'Неправильно'
        else:
            return 'Неправильно'





def check_first_error_september_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 2 = сумма 3+4+ остальные колонки
    """
    lst_sum = ['3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
                           '13', '14', '15', '16', '17']

    # получаем сумму колонок 3,4:17
    df['Сумма'] = df[lst_sum].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['2'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 5, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие:на листе 1. Форма сбора Колонка 2 = сумма колонок (3,4,5 по 17)'
    return temp_error_df


def check_second_error_september_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 3 >= каждой из 3.1 3.2 3.3
    """

    # Проводим проверку
    df['Результат'] = (df['3'] >= df['3.1']) & (df['3'] >= df['3.2']) & (df['3'] >= df['3.3'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 5, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: на листе 1. Форма сбора Колонка 3 >= значений в колонках  (3.1, 3.2, 3.3)'
    return temp_error_df


def check_third_error_september_2025(df:pd.DataFrame, name_file):
    """
    Функция для проверки 4 >= каждой из колонок 4.1 4.2 4.3
    """

    # Проводим проверку
    df['Результат'] = (df['4'] >= df['4.1']) & (df['4'] >= df['4.2']) & (df['4'] >= df['4.3'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 5, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: на листе 1. Форма сбора Колонка 4 >= значений в колонках(4.1, 4.2, 4.3)'
    return temp_error_df










def check_error_main_september_2025(df:pd.DataFrame,name_file:str):
    """
    Точка входа для проверки ошибок на основном листе
    """

    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.loc[:, '2':'17']
    df = df.applymap(check_data)

    # Проверяем 2 = сумма 3+4+ остальные колонки
    first_error_df = check_first_error_september_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

    # Проверяем 3 больше или равно  сумма 3.1 3.2 3.3
    second_error_df = check_second_error_september_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

    # Проверяем 4 больше или равно  сумма 4.1 4.2 4.3
    third_error_df = check_third_error_september_2025(df.copy(), name_file)
    error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

    return error_df



def check_contains_in_main_df(lst_spec:list,df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки наличия указанной специальности на основном листе
    """
    df['Результат'] = df['1'].apply(lambda x: x in lst_spec)
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = f'Не выполняется условие: указанная на листе {name_sheet} специальность отсутствует на листе 1. Форма сбора'
    return temp_error_df


def check_leaver_in_main_df(main_df:pd.DataFrame,df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки количество выпускников , значение на главное листе должно быть равно или больше чем сумма на листе
    """
    main_df = main_df[['1','2']]
    df = df.groupby(by='1').agg({'2':'count'}).reset_index() # группируем
    df.columns = ['Специальность','Проверяемый лист']
    main_df = pd.merge(main_df,df,how='inner',left_on='1',right_on='Специальность') # соединяем
    main_df['2'] = main_df['2'].apply(convert_to_int)
    main_df['Результат'] = main_df['2'] >= main_df['Проверяемый лист']

    main_df['Результат'] = main_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    main_df = main_df[main_df['Результат'] == 'Неправильно']
    lst_error = main_df['1'].tolist()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    temp_error_df['Строка или колонка с ошибкой'] = lst_error
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = f'Количество студентов на листе {name_sheet} с указанными специальностями больше чем общее количество студентов этих специальностей'
    return temp_error_df

def check_id(df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки заполненности ID в колонке 2
    """
    df['2'] = df['2'].fillna('Неправильно')
    df['2'] = df['2'].apply(replace_empty_with_word)

    df = df[df['2'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'На листе {name_sheet} не заполнена колонка уникальный номер выпускника'
    return temp_error_df



def check_dupl(df:pd.DataFrame,name_column_dupl:str,name_file:str,name_sheet:str):
    """
    Функция для проверки дубликатов ID в колонке 2
    """
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])


    temp_df = df[df[name_column_dupl].duplicated(keep=False)]  # получаем дубликаты
    if len(temp_df) == 0:
        return temp_error_df

    temp_df.insert(0, '№ строки дубликата ', list(map(lambda x: x + 3, list(temp_df.index))))
    temp_error_df['Строка или колонка с ошибкой'] = temp_df['№ строки дубликата ']
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'На листе {name_sheet} в колонке 2 (Уникальный номер выпускника) найдены дубликат' + ' ' + temp_df['2']
    return temp_error_df


def check_inn_target(df:pd.DataFrame,name_column:str,name_file:str,name_sheet:str):
    """
    Функция для проверки дубликатов ID в колонке 2
    """
    df['Результат'] = df[name_column].apply(lambda x:'Правильно' if len(x) in (10,12) else 'Неправильно')

    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = f'ИНН на листе {name_sheet} должен состоять из 10 или 12 цифр'
    return temp_error_df


def check_six_seven_target(df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки Если в графе 6 выбран вариант «Действует», заполните графу 7
    """
    lst_value = ['трудоустроен у работодателя, с которым заключен целевой договор',
                 'будет трудоустроен у работодателя, с которым заключен целевой договор',
                 'договор приостановлен (пролонгирован): продолжает обучение по согласованию с работодателем, заключившим договор о целевом обучении',
                 'договор приостановлен (пролонгирован): призван (будет призван) в Вооруженные Силы РФ',
                 'договор приостановлен (пролонгирован): находится (будет находиться) в отпуске по уходу за ребенком',
                 'договор приостановлен (пролонгирован): здоровье, требующее лечения, не позволяющее осуществлять трудовую деятельность',
                 ]

    df['Результат'] = df[['6','7']].apply(lambda x:'Правильно' if (x[0] == 'действует' and x[1] in lst_value) or (x[0] =='расторгнут' and pd.isna(x[1]))  else 'Неправильно',axis=1)

    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = f'Если на листе {name_sheet} в колонке 6 указано действует то в колонке 7 должно быть значение из списка указанных в шаблоне. Если в колонке 6 указано расторгнут то колонка 7 должна быть пустой.'
    return temp_error_df


def check_six_ten_target(df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки Если в графе 6 выбран вариант «расторгнут», заполните графу 8,9,10
    """
    df = df.applymap(lambda x:str.strip(x) if isinstance(x,str) else x) # очищаем от пробелов в начале и конце

    lst_eight = ['расторгнут по инициативе выпускника с возмещением заказчику расходов на меры поддержки',
                 'расторгнут по инициативе выпускника (освобождение от ответственности за неисполнение обязательств по договору)',
                 'расторгнут по инициативе работодателя с выплатой компенсации',
                 'расторгнут по инициативе работодателя (освобождение от ответственности за неисполнение обязательств по договору)',
                 'расторгнут по соглашению сторон или по независящим от сторон обстоятельствам',
                 ]

    lst_nine = ['отказ выпускника от трудоустройства в связи с низким уровнем заработной платы',
                'отказ выпускника от трудоустройства в связи с переездом, удаленностью места работы',
                 'отказ работодателя от трудоустройства в связи с неудовлетворенностью знаниями, умениями, навыками и компетенциями выпускника',
                 'отказ работодателя от трудоустройства в связи с отсутствием вакансий/сокращением штата предприятия',
                 'отказ работодателя от трудоустройства и выполнения условий договора (по иным причинам)',
                 'по независящим причинам: выпускник находится под следствием, отбывает наказание',
                 'по независящим причинам: выпускник осуществляет постоянный уход за ближайшим родственником',
                 'по независящим причинам: супруг/супруга выпускника - военнослужащий, проходит службу на другой территории',
                 'по независящим причинам: выпускнику присвоена инвалидность 1 или 2 группы',
                 'по независящим причинам: несоблюдение требований законодательства к работникам (медицинские противопоказания, судимость, отказ в допуске к государственной тайне, не прошел аккредитацию специалиста и др.).',
                 'по независящим причинам: смерть выпускника, тяжелое состояние здоровья'
                 ]

    lst_ten = ['трудоустроен у иного работодателя',
               'продолжает обучение',
               'призван (будет призван) в Вооруженные Силы РФ',
               'находятся (будут находиться) в отпуске по уходу за ребенком',
               'индивидуальный предприниматель',
               'самозанятый',
               'не может трудоустроиться: находится под следствием, отбывает наказание',
               'не может трудоустроиться: ухаживает за больными родственниками, иные семейные обстоятельства',
               'не может трудоустроиться: смерть выпускника, тяжелое состояние здоровья',
               'не имеет мотивации к ведению трудовой деятельности и не планирует трудоустраиваться',
               'переехал (планирует переезд) за пределы РФ',
               'неофициально трудоустроен',
               'отсутствует спрос на специалистов в регионе, находится в поиске работы',
               'зарегистрирован в центрах занятости в качестве безработного (получает пособие по безработице)',
               'не планирует трудоустраиваться, в том числе по причинам получения иных социальных льгот',
               ]
    df['Результат'] = df[['6','8','9','10']].apply(lambda x:'Правильно' if (x[0] == 'расторгнут' and x[1] in lst_eight and x[2] in lst_nine and x[3] in lst_ten) or (x[0] == 'действует' and pd.isna(x[1]) and pd.isna(x[2]) and pd.isna(x[3])) else 'Неправильно',axis=1)

    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = f'Если на листе {name_sheet}  в колонке 6 указано расторгнут то в колонках 8,9,10 должно быть значения из списка указанных в шаблоне. Если в колонке 6 указано действует то колонки 8,9,10 должны быть пустыми.'
    return temp_error_df





def check_employers_target(df:pd.DataFrame,name_file:str,name_sheet:str):
    """
    Функция для проверки Если в графе 8  выбран вариант «Работодатели», заполните графу 7
    """

    lst_emp_eight = ['расторгнут по инициативе работодателя с выплатой компенсации','расторгнут по инициативе работодателя (освобождение от ответственности за неисполнение обязательств по договору)']
    lst_emp_nine = ['отказ работодателя от трудоустройства в связи с неудовлетворенностью знаниями, умениями, навыками и компетенциями выпускника',
                'отказ работодателя от трудоустройства в связи с отсутствием вакансий/сокращением штата предприятия',
                'отказ работодателя от трудоустройства и выполнения условий договора (по иным причинам)']

    lst_work_eight = ['расторгнут по инициативе выпускника с возмещением заказчику расходов на меры поддержки','расторгнут по инициативе выпускника (освобождение от ответственности за неисполнение обязательств по договору)']
    lst_work_nine = ['отказ выпускника от трудоустройства в связи с низким уровнем заработной платы','отказ выпускника от трудоустройства в связи с переездом, удаленностью места работы']

    lst_all_eight =['расторгнут по соглашению сторон или по независящим от сторон обстоятельствам']
    lst_all_nine = ['по независящим причинам: выпускник находится под следствием, отбывает наказание','по независящим причинам: выпускник осуществляет постоянный уход за ближайшим родственником','по независящим причинам: супруг/супруга выпускника - военнослужащий, проходит службу на другой территории',
                    'по независящим причинам: выпускнику присвоена инвалидность 1 или 2 группы','по независящим причинам: несоблюдение требований законодательства к работникам (медицинские противопоказания, судимость, отказ в допуске к государственной тайне, не прошел аккредитацию специалиста и др.).',
                    'по независящим причинам: смерть выпускника, тяжелое состояние здоровья']


    # df['Результат'] = df[['8', '9']].apply(lambda x: 'Правильно' if () or (x[0] in lst_emp_eight and x[1] in lst_emp_nine) or (x[0] in lst_work_eight and x[1] in lst_work_nine) or (x[0] in lst_all_eight and x[1] in lst_all_nine) else 'Неправильно', axis=1)
    df['Результат'] = df[['6','8','9']].apply(lambda x: 'Правильно' if (x[0] =='действует') or (x[1] in lst_emp_eight and x[2] in lst_emp_nine) or (x[1] in lst_work_eight and x[2] in lst_work_nine) or (x[1] in lst_all_eight and x[2] in lst_all_nine) else 'Неправильно', axis=1)

    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 3, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = (f'Если на листе {name_sheet}  в колонке 8 выбраны варианты «Расторгнут по инициативе работодателя…» то в колонке 9 должны быть только варианты «Отказ работодателя…»;'
                              f'если в колонке 8 выбраны варианты «Расторгнут по инициативе выпускника…» то в колонке 9 должны быть только варианты «Отказ выпускника…»;'
                              f'если в колонке 8 выбраны варианты «Расторгнут соглашению сторон…» то в колонке 9 должны быть только варианты «По независящим причинам…».')
    return temp_error_df















def check_error_nose_september_2025(main_df:pd.DataFrame,nose_df:pd.DataFrame,name_file:str):
    """
    Точка входа для проверки ошибок на листе нозологий
    """

    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    lst_spec = main_df['1'].unique() # список уникальных специальностей

    # проверяем наличие специальности
    contains_error_df = check_contains_in_main_df(lst_spec,nose_df.copy(),name_file,'2. Нозологии')
    error_df = pd.concat([error_df, contains_error_df], axis=0, ignore_index=True)

    # проверяем количество выпускников с нозологиями или целевиков не должно превышать общее количество выпускников
    quantity_leaver_error_df = check_leaver_in_main_df(main_df.copy(),nose_df.copy(),name_file,'2. Нозологии')
    error_df = pd.concat([error_df, quantity_leaver_error_df], axis=0, ignore_index=True)

    # проверяем отсутствие идентификатора
    id_error_df = check_id(nose_df.copy(),name_file,'2. Нозологии')
    error_df = pd.concat([error_df, id_error_df], axis=0, ignore_index=True)

    # проверяем дубликаты ID
    dupl_error_df = check_dupl(nose_df.copy(),'2',name_file,'2. Нозологии')
    error_df = pd.concat([error_df, dupl_error_df], axis=0, ignore_index=True)

    return error_df



def check_error_target_september_2025(main_df:pd.DataFrame,target_df:pd.DataFrame,name_file:str):
    """
    Точка входа для проверки ошибок на листе целевиков
    """

    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    lst_spec = main_df['1'].unique() # список уникальных специальностей

    # проверяем наличие специальности
    contains_error_df = check_contains_in_main_df(lst_spec,target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, contains_error_df], axis=0, ignore_index=True)

    # проверяем количество выпускников с нозологиями или целевиков не должно превышать общее количество выпускников
    quantity_leaver_error_df = check_leaver_in_main_df(main_df.copy(),target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, quantity_leaver_error_df], axis=0, ignore_index=True)

    # проверяем отсутствие идентификатора
    id_error_df = check_id(target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, id_error_df], axis=0, ignore_index=True)

    # проверяем дубликаты ID
    dupl_error_df = check_dupl(target_df.copy(),'2',name_file,'3. Целевики')
    error_df = pd.concat([error_df, dupl_error_df], axis=0, ignore_index=True)

    # Проверяем ИНН
    inn_error_df = check_inn_target(target_df.copy(),'5',name_file,'3. Целевики')
    error_df = pd.concat([error_df, inn_error_df], axis=0, ignore_index=True)

    # Если в графе 6 выбран вариант «Действует», заполните графу 7
    six_seven_error_df = check_six_seven_target(target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, six_seven_error_df], axis=0, ignore_index=True)

    # # Если в графе 6 выбран вариант «расторгнут», заполните графу 8,9,10
    six_ten_error_df = check_six_ten_target(target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, six_ten_error_df], axis=0, ignore_index=True)

    # # Если в графе 8 выбран вариант работодатели, то в графе 9 тоже должны быть работодатели
    employers_error_df = check_employers_target(target_df.copy(),name_file,'3. Целевики')
    error_df = pd.concat([error_df, employers_error_df], axis=0, ignore_index=True)









    return error_df
























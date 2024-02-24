# -*- coding: utf-8 -*-
"""
Скрипт для обработки данных Формы 1 пятистрочной мониторинга занятости выпускников
"""
from check_functions import * # импортируем функции проверки
from support_functions import * # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random


def prepare_form_one_employment(path_folder_data:str,path_to_end_folder):
    """
    Фугкция для обработки данных формы 1 пять строк
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем словарь верхнего уровня для хранения пары ключ значение где ключ это код специальности а значение- код и наименование
    dct_code_and_name = dict()
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    tup_correct = (6, 10)  # создаем кортеж  с поправками где 6 это первая строка с данными а 10 строка где заканчивается первый диапазон

    try:
        for file in os.listdir(path_folder_data):
            if not file.startswith('~$') and not file.endswith('.xlsx'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                # получаем название первого листа
                temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}', read_only=True)
                lst_temp_sheets = temp_wb.sheetnames  # получаем листы в файле
                temp_wb.close()
                if 'Форма 1 пятистрочная' not in lst_temp_sheets: # проверяем наличие листа с названием в файле
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не найден лист с названием Форма 1 пятистрочная !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=4, dtype=str)
                df.columns = list(map(str,df.columns)) # делаем названия колонок строковыми
                # создаем множество колонок наличие которых мы проверяем
                check_cols = ['01','02', '03','04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15',
                              '16', '17',
                              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '30']
                if check_cols != list(df.columns):
                    diff_cols = set(list(df.columns)).difference(set(check_cols))
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{diff_cols}',
                                                        'Возможно старая версия формы сбора данных.Строка с номерами колонок (01,02,03,05 ... 28,30 как в исходной форме)\n должна находиться на 5 строке!\n'
                                                        ' указанные колонки являются лишними.Колонки с названимем Unnamed означаеют что на листе есть данные без заголовка в виде цифр на 5 строке  ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df = df.loc[:, '02':'27'] # отсекаем колонки с регионом и проверками
                # получаем  часть с данными
                mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # проверяем есть ли строка полностью состоящая из nan
                empty_row_index = np.where(df.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df = df.iloc[:row_index]
                #     # Проверка на размер таблицы, должно бьть кратно 5
                count_spec = df.shape[0] // 5  # количество специальностей
                check_code_lst = df['02'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке 02 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue
                # Проверка на непрерывность кода специальности, то есть на 5 строк должен быть только один код и на пустые ячейки
                border_check_code = 0  # начало отсчета
                quantity_check_code = len(check_code_lst) // 5  # получаем сколько специальностей в таблице
                correction = 0  # размер поправки на случай если есть строка проверки
                sameness_error_df = check_sameness_column(check_code_lst, 5, border_check_code, quantity_check_code,
                                                          tup_correct, correction, name_file, 'Код и наименование')

                blankness_error_df = check_blankness_column(check_code_lst, 5, border_check_code, quantity_check_code,
                                                            tup_correct, correction, name_file, 'Код и наименование')

                # проверяем на арифметические ошибки
                file_error_df = check_error_form_one(df.copy(), name_file, tup_correct)
                # добавляем в получившийся датафейм ошибки однородности диапазона
                file_error_df = pd.concat([file_error_df, sameness_error_df], axis=0, ignore_index=True)
                file_error_df = pd.concat([file_error_df, blankness_error_df], axis=0, ignore_index=True)

                for full_name in df['02'].tolist():
                    code = extract_code_nose(full_name) # получаем только цифры
                    dct_code_and_name[code] = full_name
                # очищаем от текста чтобы названия листов не обрезались
                df['02'] = df['02'].apply(extract_code_nose)  # очищаем от текста в кодах
                if 'error' in df['02'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 02 Код и наименование профессии/специальности.Вместо кода присутствует дата, и т.п. проверьте правильность заполнения колонки 02!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    file_error_df = pd.concat([file_error_df, temp_error_df], axis=0, ignore_index=True)
                # добавляем в основной файл с ошибками
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # Создание словаря для хранения данных файла
                code_spec = [spec for spec in df['02'].unique()] # получаем список специальностей которые есть в файле
                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 6)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(5, 28)]  # раньше было 7
                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}
                high_level_dct[name_file] = copy.deepcopy(poo_dct)
        #             """
        #             В итоге получается такая структура
        #             {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк

                # Итерируемся по полученному датафрейму через itertuples
                for row in df.itertuples():
                    # если счетчик колонок больше 15 то уменьшаем его до единицы
                    if idx_row > 5:
                        idx_row = 1
                    # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[1] is not np.nan) and (row[1] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[1] != current_code:
                            current_code = row[1]
                    data_row = row[4:27]  # получаем срез с нужными данными


                    for idx_col, value in enumerate(data_row, start=1):
                        high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                            f'Колонка {idx_col + 4}'] += check_data(value)
                    #
                    idx_row += 1
        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)
        wb_check_tables = create_check_tables_form_one(high_level_dct)  # проверяем данные по каждой специальности
        wb_check_tables.save(
            f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')

        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}
        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')

        stack_df = out_df.stack()
        # название такое выбрал потому что было лень заменять значения из блокнота юпитера
        frame = stack_df.to_frame()
        frame['Всего'] = frame[0].apply(lambda x: x.get('Колонка 5'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 6'))
        frame['Индивидуальные предприниматели'] = frame[0].apply(lambda x: x.get('Колонка 7'))
        frame[
            'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
            frame[0].apply(lambda x: x.get('Колонка 8'))
        frame['Продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 9'))
        frame['Проходят службу в армии по призыву'] = frame[0].apply(lambda x: x.get('Колонка 10'))
        frame[
            'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 11'))
        frame['Находятся в отпуске по уходу за ребенком'] = frame[0].apply(
            lambda x: x.get('Колонка 12'))
        frame['Неформальная занятость (теневой сектор экономики)'] = frame[0].apply(lambda x: x.get('Колонка 13'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
            frame[0].apply(lambda x: x.get('Колонка 14'))
        frame[
            'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
            frame[0].apply(lambda x: x.get('Колонка 15'))
        frame[
            'Иные причины нахождения под риском нетрудоустройства (включая отсутствие проводимой с выпускниками работы по содействию их занятости)'] = \
        frame[0].apply(
            lambda x: x.get('Колонка 16'))
        frame['Смерть, тяжелое состояние здоровья'] = frame[0].apply(lambda x: x.get('Колонка 17'))
        frame['Находятся под следствием, отбывают наказание'] = frame[0].apply(
            lambda x: x.get('Колонка 18'))
        frame[
            'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame[
            'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
            frame[0].apply(lambda x: x.get('Колонка 20'))
        frame['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = frame[0].apply(
            lambda x: x.get('Колонка 21'))
        frame['будут трудоустроены'] = frame[0].apply(lambda x: x.get('Колонка 22'))
        frame['будут осуществлять предпринимательскую деятельность'] = frame[0].apply(
            lambda x: x.get('Колонка 23'))
        frame['будут самозанятыми'] = frame[0].apply(lambda x: x.get('Колонка 24'))
        frame['будут призваны в армию'] = frame[0].apply(lambda x: x.get('Колонка 25'))
        frame[
            'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
            frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['будут продолжать обучение'] = frame[0].apply(lambda x: x.get('Колонка 27'))

        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)

        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении'

               }
        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        finish_df['Код специальности'] = finish_df['Код специальности'].apply(lambda x:dct_code_and_name[x]) # делаем код чтобы отображался код и наименование
        # Создаем файл в котором будут отображаться листы с 5 и одной строками
        one_row_finish_df = pd.DataFrame(columns=finish_df.columns) # для одной строки по каждой специальности
        lst_code_spec = finish_df['Код специальности'].unique()  # получаем список специальностей
        for code_spec in lst_code_spec:
            temp_df = finish_df[finish_df['Код специальности'] == code_spec]
            one_row_finish_df = pd.concat([one_row_finish_df, temp_df.iloc[:1, :]], axis=0, ignore_index=True)

        with pd.ExcelWriter(f'{path_to_end_folder}/Полная таблица Форма 1 пятистрочная от {current_time}.xlsx') as writer:
            finish_df.to_excel(writer, sheet_name='5 строк', index=False)
            one_row_finish_df.to_excel(writer, sheet_name='1 строка (Всего выпускников)', index=False)

        # Создаем документ
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 50
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ Форма 1 пятистрочная от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    # except:
    #     messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
    #                          f'При обработке файла {name_file} возникла ошибка !!!\n'
    #                          f'Проверьте файл на соответсвие шаблону')

    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                                'Данные успешно обработаны.Ошибок не обнаружено')


if __name__ == '__main__':
    main_data_folder = 'data/example/testing'
    main_result_folder = 'data/result'
    prepare_form_one_employment(main_data_folder,main_result_folder)

    print('Lindy Booth')
# -*- coding: utf-8 -*-
"""
Функции для проверки данных
"""
from support_functions import * # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time

pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random

def check_data(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0


def check_data_note(cell):
    if cell is np.nan:
        return 'Не заполнено'
    return str(cell)

def check_sameness_column(checked_lst:list,check_range:int,begin_border:int,quantity_check_value:int,tup_correct:tuple,correction:int,
                          name_file=None,name_column=None):
    """
    Функция для проверки заполнен ли определенный диапазон одинаковыми значениями
    checked_lst : список значений который нужно проверить на однородность в каждом диапазоне
    check_range : сколько значений в  првоеряемом диапазоне
    quantity_check_value : количество проверяемых диапазонов
    tup_correct : кортеж  где нулевой элемент это первая строка диапазона в экселе а первый элемент последняя строка в диапазоне
    это нужно чтобы точно указывать где искать ошибку в файле Excel
    correction : дополнительная поправка
    name_file : имя обрабатываемого файла Excel

    """
    # датафрейм для ошибок
    _error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    offset = 0 # сдвиг

    for i in range(quantity_check_value):
        temp_set = set(
            [value for value in checked_lst[begin_border:begin_border + check_range]])
        if len(temp_set) != 1:
            temp_error_df = pd.DataFrame(data=[[name_file,f'Диапазон строк {begin_border + first_correct + offset} - {begin_border + second_correct +offset}',
                                                f'В колонке {name_column} в указанном диапазоне обнаружены отличающиеся значения']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
            _error_df = pd.concat([_error_df,temp_error_df],axis=0,ignore_index=True)

        begin_border +=  check_range # сдвигаем проверяемый диапазон в списке
        offset += correction # добавляем поправку

    return _error_df


def check_blankness_column(checked_lst:list,check_range:int,begin_border:int,quantity_check_value:int,tup_correct:tuple,correction:int,
                          name_file=None,name_column=None):
    """
    Функция для проверки есть ли в определенном диапазоне пустые значения
    checked_lst : список значений который нужно проверить на однородность в каждом диапазоне
    check_range : сколько значений в  првоеряемом диапазоне
    quantity_check_value : количество проверяемых диапазонов
    tup_correct : кортеж  где нулевой элемент это первая строка диапазона в экселе а первый элемент последняя строка в диапазоне
    это нужно чтобы точно указывать где искать ошибку в файле Excel
    correction : дополнительная поправка
    name_file : имя обрабатываемого файла Excel

    """
    # датафрейм для ошибок
    _error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    offset = 0 # сдвиг

    for i in range(quantity_check_value):
        temp_set = set(
            [value for value in checked_lst[begin_border:begin_border + check_range]])
        if np.nan in temp_set or ' ' in temp_set:
            temp_error_df = pd.DataFrame(data=[[name_file,f'Диапазон строк {begin_border + first_correct + offset} - {begin_border + second_correct +offset}',
                                                f'В колонке {name_column} в указанном диапазоне обнаружены незаполненные ячейки']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
            _error_df = pd.concat([_error_df,temp_error_df],axis=0,ignore_index=True)

        begin_border +=  check_range # сдвигаем проверяемый диапазон в списке
        offset += correction # добавляем поправку

    return _error_df











def check_first_error_temp(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки гр. 09 и гр. 10 < гр. 08
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]

    # Проводим проверку
    df['Результат'] = (df['08'] >= df['09']) & (df['08'] >= df['10'])
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 09 <= гр. 08  или гр. 10 <= гр. 08'
    return temp_error_df


def check_second_error_temp(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 07= гр.08 + сумма(с гр.11 по гр.32))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # конвертируем в инт
    all_sum_cols = list(df)
    # удаляем колонки 07, 09, 10
    all_sum_cols.remove('07')
    all_sum_cols.remove('09')
    all_sum_cols.remove('10')
    # получаем сумму колонок 08, 11:32
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['07'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 07 = гр.08 + сумма(с гр.11 по гр.32)'
    return temp_error_df


def check_third_error_temp(df: pd.DataFrame, name_file, border, tup_correct,correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр. 02 + стр. 04
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '06', ])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['Сумма'] = foo_df['02'] + foo_df['04']
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр. 02 + стр. 04 '
    return temp_error_df

"""
Проверки
"""
def check_first_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр 03 <= стр 02 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '03'])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['03'] = df.iloc[2, :]
    foo_df['Результат'] = foo_df['03'] <= foo_df['02']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 03 <= стр. 02 '
    return temp_error_df

def check_second_error(df: pd.DataFrame, name_file, border, tup_correct: tuple, correction):
    """
    Функция для проверки правильности введеденных данных
    стр.02 и стр.04 и стр.05 < стр.01
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '05', '01'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['05'] = df.iloc[4, :]

    foo_df['Результат'] = (foo_df['01'] >= foo_df['02']) & (foo_df['01'] >= foo_df['04']) & (
                foo_df['01'] >= foo_df['05'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр.02<= стр.01 или стр.04<= стр.01 или стр.05<= стр.01 '
    return temp_error_df


def check_third_error(df: pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки правильности введеденных данных
    (гр. 05= сумма(с гр.06 по гр.27))
    :param df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    all_sum_cols = list(df) # получаем список колонок
    # удаляем колонку 05 с общей суммой
    all_sum_cols.remove('05')
    # получаем сумму колонок 06:27
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df['05'] == df['Сумма']
    # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 05 = сумма(с гр.06 по гр.27)'
    return temp_error_df










def check_fourth_error(df: pd.DataFrame, name_file, border, tup_correct,correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13

    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['06', '07', '08', '09', '10', '11', '12', '13'])

    # Добавляем данные в датафрейм
    foo_df['06'] = df.iloc[5, :]
    foo_df['07'] = df.iloc[6, :]
    foo_df['08'] = df.iloc[7, :]
    foo_df['09'] = df.iloc[8, :]
    foo_df['10'] = df.iloc[9, :]
    foo_df['11'] = df.iloc[10, :]
    foo_df['12'] = df.iloc[11, :]
    foo_df['13'] = df.iloc[12, :]

    sum_col = ['07', '08', '09', '10', '11', '12', '13']
    foo_df['Сумма'] = foo_df[sum_col].sum(axis=1)
    foo_df['Результат'] = foo_df['06'] == foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct + correction} - {border + second_correct +correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13 '
    return temp_error_df


def check_fifth_error(df: pd.DataFrame, name_file, border, tup_correct,correction):
    """
    Функция для проверки правильности введеденных данных
    стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]

    foo_df = pd.DataFrame(columns=['05', '06', '14'])

    # Добавляем данные в датафрейм
    foo_df['05'] = df.iloc[4, :]
    foo_df['06'] = df.iloc[5, :]
    foo_df['14'] = df.iloc[13, :]

    foo_df['Результат'] = (foo_df['14'] <= foo_df['05']) & (foo_df['14'] <= foo_df['06'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()

    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df[
        'index'].tolist()  # делаем список, прибавляем для того чтобы номера строк совпадали с строками в файле
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк  {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 14<=стр. 06, стр. 14<=стр 05'
    return temp_error_df


def check_sixth_error(df: pd.DataFrame, name_file, border, tup_correct: tuple,correction):
    """
    Функция для проверки правильности введеденных данных
    стр 03 <= стр 02 (<= означает "меньше или равно")
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '03'])

    # Добавляем данные в датафрейм
    foo_df['02'] = df.iloc[1, :]
    foo_df['03'] = df.iloc[2, :]
    foo_df['Результат'] = foo_df['03'] <= foo_df['02']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 03 <= стр. 02 '
    return temp_error_df


def check_seventh_error(df: pd.DataFrame, name_file, border, tup_correct: tuple,correction):
    """
    Функция для проверки правильности введеденных данных
    стр.02 и стр.04 и стр.05 < стр.01
    :param foo_df: копия датафрейма с данными из файла поо
    : param tup_correction кортеж с поправочными границами для того чтобы диапазон строки с ошибкой корректно считался
    :return:датафрейм с ошибками
    """
    # получаем поправки на диапазон
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['02', '04', '05', '01'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['04'] = df.iloc[3, :]
    foo_df['05'] = df.iloc[4, :]

    foo_df['Результат'] = (foo_df['01'] >= foo_df['02']) & (foo_df['01'] >= foo_df['04']) & (
                foo_df['01'] >= foo_df['05'])
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct+correction} - {border + second_correct+correction}, колонка {str(x)}',
            raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = 'Не выполняется условие: стр.02<= стр.01 или стр.04<= стр.01 или стр.05<= стр.01 '
    return temp_error_df

def check_error_form_one(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 3:26] # получаем часть с числами
    df = df.applymap(check_data) # заполняем пустые ячейки нулями

    # получаем количество датафреймов
    quantity = df.shape[0] // 5
    # счетчик для обработанных строк
    border = 0
    correction = 0 # поправка для учета строки c проверками
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 5, :]

        # Проводим проверку стр.03 <= стр.02
        first_error_df = check_first_error(temp_df.copy(), name_file,border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        second_error_df = check_second_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)
        #
        # Проводим проверку гр. 05=сумма(с гр.06 по гр.28)
        third_error_df = check_third_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 5
    # Возвращаем датафрейм с ошибками
    return error_df





def create_check_tables_form_one(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 5'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 12'))
            temp_code_df['Неформальная занятость (теневой сектор экономики)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства (включая отсутствие проводимой с выпускниками работы по содействию их занятости)'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 20'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 21'))
            temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 22'))
            temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 23'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 25'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))


            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении'
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb


def check_error_form_two(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 3:26]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    correction = 0 # поправка для учета строки 16 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку стр 03 <= стр 02
        first_error_df = check_first_error(temp_df.copy(), name_file,border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        second_error_df = check_second_error(temp_df.copy(), name_file, border, tup_correct, correction)
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)
        # Проводим проверку гр. 05=сумма(с гр.06 по гр.28)
        third_error_df = check_third_error(temp_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)


        # # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        # fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # # добавляем результат проверки в датафрейм
        # error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)
        #
        # # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        # fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # # добавляем результат проверки в датафрейм
        # error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)
        #
        # # Проводим проверку стр.03 <= стр.02
        # sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)
        #
        # # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        # seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
        correction +=1
    # Возвращаем датафрейм с ошибками
    return error_df





def check_error_nose(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных нозологий
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 4:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    correction = 0 # поправка для учета строки 16 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.03 <= стр.02
        sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct,correction)
        error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
        correction +=1
    # Возвращаем датафрейм с ошибками
    return error_df

def check_error_base_mon(df: pd.DataFrame, name_file, tup_correct: tuple):
    """
    Функция для проверки данных базового мониторинга 5 строк
    tup_correct - кортеж  с поправками для того чтобы диапазон строк с ошибкой корректно отображался
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    df = df.iloc[:, 6:32]
    df = df.applymap(check_data)

    # получаем количество датафреймов
    quantity = df.shape[0] // 15
    # счетчик для обработанных строк
    border = 0
    correction = 0 # поправка для учета строки 16 чтобы диапазон ошибки отображался правильно
    for i in range(1, quantity + 1):
        temp_df = df.iloc[border:border + 15, :]
        # Проводим проверку гр. 09 и гр. 10 <= гр. 08
        first_error_df = check_first_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, first_error_df], axis=0, ignore_index=True)

        # Проводим проверку гр. 07= гр.08 + сумма(с гр.11 по гр.32)
        second_error_df = check_second_error(temp_df.copy(), name_file, tup_correct)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, second_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр. 02 + стр. 04
        third_error_df = check_third_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, third_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 06 = стр.07 + стр.08 + стр.09 + стр.10 + стр.11 + стр.12 + стр. 13
        fourth_error_df = check_fourth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fourth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр. 14<=стр. 06, стр. 14<=стр 05 (<= означает "меньше или равно")
        fifth_error_df = check_fifth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        # добавляем результат проверки в датафрейм
        error_df = pd.concat([error_df, fifth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.03 <= стр.02
        sixth_error_df = check_sixth_error(temp_df.copy(), name_file, border, tup_correct,correction)
        error_df = pd.concat([error_df, sixth_error_df], axis=0, ignore_index=True)

        # Проводим проверку стр.02 и стр.04 и стр.05 < стр.01
        seventh_error_df = check_seventh_error(temp_df.copy(), name_file, border, tup_correct,correction)
        error_df = pd.concat([error_df, seventh_error_df], axis=0, ignore_index=True)

        # прибавляем border

        border += 15
    # Возвращаем датафрейм с ошибками
    return error_df


def extract_code(value):
    """
    Функция для извлечения кода специальности
    """
    value = str(value)
    re_code = re.compile('^\d{2}?[.]\d{2}?[.]\d{2}')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.group()
    else:
        return 'error'

def extract_code_nose(value):
    """
    Функция для извлечения кода специальности из новой формы по нозологиям
    """
    value = str(value)
    re_code = re.compile('(\d{2}?[.]\d{2}?[.]\d{2})\D')  # создаем выражение для поиска кода специальности
    result = re.search(re_code, value)
    if result:
        return result.groups()[0]
    else:
        return 'error'


def create_check_tables(high_level_dct: dict):
    """
    Функция для создания файла с данными по каждой специальности
    """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')
            temp_code_df = temp_code_df.stack()
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Всего'] = temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df[
                'В том числе (из трудоустроенных): в соответствии с освоенной профессией, специальностью (исходя из осуществляемой трудовой функции)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df[
                'В том числе (из трудоустроенных): работают на протяжении не менее 4-х месяцев на последнем месте работы'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Индивидуальные предприниматели'] = temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df[
                'Самозанятые (перешедшие на специальный налоговый режим  - налог на профессио-нальный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
            temp_code_df['Продолжили обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df['Проходят службу в армии по призыву'] = temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 16'))
            temp_code_df['Неформальная занятость (нелегальная)'] = temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 20'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 22'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df['Выпускники из числа иностранных граждан, которые не имеют СНИЛС'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 25'))
            temp_code_df[
                'Иное (в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 1 из 200 выпускников - укажите причины в гр. 33 '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['будут трудоустроены'] = temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['будут осуществлять предпринимательскую деятельность'] = temp_code_df[0].apply(
                lambda x: x.get('Колонка 28'))
            temp_code_df['будут самозанятыми'] = temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
            temp_code_df['будут призваны в армию'] = temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 31'))
            temp_code_df['будут продолжать обучение'] = temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
            temp_code_df['Принимаемые меры по содействию занятости (тезисно - вид меры, охват выпускников мерой)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 33'))

            finish_code_spec_df = temp_code_df.drop([0], axis=1)

            finish_code_spec_df = finish_code_spec_df.reset_index()

            finish_code_spec_df.rename(
                columns={'level_0': 'Название файла', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
                   'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
                   'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
                   'Строка 5': 'Имеют договор о целевом обучении',
                   'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
                ,
                   'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
                   'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
                   'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
                   'Строка 12': 'расстройствами аутистического спектра',
                   'Строка 13': 'с инвалидностью вследствие  других причин',
                   'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
                   'Строка 15': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): принимали участие в чемпионате «Абилимпикс»',
                   }
            finish_code_spec_df['Наименование показателей (категория выпускников)'] = finish_code_spec_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_code_spec_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40
    return wb





"""
Функции для обработки отчетов ЦК
"""


def check_horizont_all_sum_error(df: pd.DataFrame, tup_exluded_cols: tuple, name_itog_cols, name_file):
    """
    Функция для проверки горизонтальных сумм по всей строке
    сумма в колонке 05 должна быть равна сумме всех колонок за исключением 07 и 15
    """
    # датафрейм для ощибок по горизонтали
    hor_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # получаем список колонок
    all_sum_cols = list(df)
    # удаляем колонки
    for name_cols in tup_exluded_cols:
        all_sum_cols.remove(name_cols)
    # удаляем итоговую колонку
    all_sum_cols.remove(name_itog_cols)

    # получаем сумму колонок за вычетом исключаемых и итоговой колонки
    df['Сумма'] = df[all_sum_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] == df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} = сумма остальных гр. за искл.{tup_exluded_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_horizont_chosen_sum_error(df: pd.DataFrame, tup_checked_cols: list, name_itog_cols, name_file):
    """
    Функция для проверки равенства одиночных или небольших групп колонок
    tup_checked_cols колонки сумму которых нужно сравнить с name_itog_cols чтобы она не превышала это значение
    """
    # Считаем проверяемые колонки
    df['Сумма'] = df[tup_checked_cols].sum(axis=1)
    # Проводим проверку
    df['Результат'] = df[name_itog_cols] >= df['Сумма']
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + 1, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка 0{str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df[
        'Описание ошибки'] = f'Не выполняется условие: гр. {name_itog_cols} >= сумма {tup_checked_cols} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!'
    return temp_error_df


def check_vertical_chosen_sum(df: pd.DataFrame, lst_checked_rows: list, itog_row, name_file):
    """
    Функция для проверки вертикальной суммы заданных строк сумма значений в tupl_checked_row должна быть равной ил меньше чем значение
    в itog_row
    """

    # обрабаотываем список строк чтобы привести его в читаемый вид
    lst_out_rows = list(map(lambda x: x + 1, lst_checked_rows))

    # делаем значения строковыми
    lst_out_rows = list(map(str, lst_out_rows))
    # Добавляем ноль в строки
    lst_out_rows = list(map(lambda x: '0' + x, lst_out_rows))
    # обрабатываем формат выходной строки
    out_itog_row = f'0{str(itog_row + 1)}'

    # создаем временный датафрейм
    foo_df = pd.DataFrame()
    # разворачиваем строки в колонки
    for idx_row in lst_checked_rows:
        foo_df[idx_row] = df.iloc[idx_row, :]

    # добавляем итоговую колонку
    foo_df[itog_row] = df.iloc[itog_row, :]

    # суммируем
    foo_df['Сумма'] = foo_df[lst_checked_rows].sum(axis=1)
    foo_df['Результат'] = foo_df[itog_row] >= foo_df['Сумма']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # получаем датафрейм с ошибками и извлекаем индекс
    error_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    # Добавляем слово колонка
    error_df['index'] = error_df['index'].apply(lambda x: 'Колонка ' + str(x))
    # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    temp_error_df['Строка или колонка с ошибкой'] = error_df['index']
    temp_error_df[
        'Описание ошибки'] = f'Для указанной колонки сумма в строках {lst_out_rows} превышает значением в строке {out_itog_row} ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! '
    temp_error_df['Название файла'] = name_file

    return temp_error_df
def check_error_ck(df: pd.DataFrame, name_file):
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки'])

    # проводим горизонтальные проверки
    # проверка на общую сумму
    first_error_ck_df = check_horizont_all_sum_error(df.copy(), ('07', '15'), '05', name_file)
    error_df = pd.concat([error_df, first_error_ck_df], axis=0, ignore_index=True)
    # проверяем небольшие группы или одиночные колонки
    second_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['07'], '06', name_file)
    error_df = pd.concat([error_df, second_error_ck_df], axis=0, ignore_index=True)

    # проверяем колонки 14 и 15
    third_error_ck_df = check_horizont_chosen_sum_error(df.copy(), ['15'], '14', name_file)
    error_df = pd.concat([error_df, third_error_ck_df], axis=0, ignore_index=True)

    # Проводим вертикальные проверки
    # Сумма овз и целевиков не должна превышать общую численность выпускников. Строки с индексом 1 и 4 должныть меньше или равны строке с индексом 0
    fourth_error_ck_df = check_vertical_chosen_sum(df.copy(), [1, 4], 0, name_file)
    error_df = pd.concat([error_df, fourth_error_ck_df], axis=0, ignore_index=True)

    # Проверяем ОВЗ
    fifth_error_ck_df = check_vertical_chosen_sum(df.copy(), [2, 3], 1, name_file)
    error_df = pd.concat([error_df, fifth_error_ck_df], axis=0, ignore_index=True)

    return error_df

"""
Проверки ОПК
"""


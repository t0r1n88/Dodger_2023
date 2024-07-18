# -*- coding: utf-8 -*-
"""
Модуль для обработки таблиц мониторинга занятости выпускников используемого для загрузки на сайт СССР
"""
from check_functions import base_check_file, extract_code_nose
from support_functions import convert_to_int

from mon_grad_check_functions import create_check_tables_mon_grad, check_error_mon_grad_spo, check_error_mon_grad_target

import pandas as pd
import numpy as np
import os
import warnings
from tkinter import messagebox
import time

pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def prepare_graduate_employment(path_folder_data: str, path_result_folder: str):
    """
    Функция для обработки мониторинга занятости
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    requred_columns_first_sheet = ['1', '1.1', '1.2', '2', '3', '3.1', '3.2', '4', '5', '6', '7', '8', '9', '10',
                                   '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
                                   '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
                                   '31', '32', '32.1', '32.2', '33', '34', '35', '36', '37', '38', '39', '40',
                                   '41', '42', '43', '44', '45', '46', '47', '48', '49', '50',
                                   '51', '52', '53', '54', '55', '56', '57', '58', '59', '60',
                                   '61', '62', '63', '64', '65', '66', '67', '68', '69', '70',
                                   '71', '72', '73']

    text_required_columns_first_sheet = ['73']

    requred_columns_second_sheet = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']
    columns_for_out_second_df = ['Название файла', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']

    # создаем базовый датафрейм для данных второго листа
    main_second_df = pd.DataFrame(columns=columns_for_out_second_df)

    try:
        for file in os.listdir(path_folder_data):
            if not file.startswith('~$'):

                # Проверяем файл на расширение, наличие нужных листов и колонок
                file_error_df, check_required_dct = base_check_file(file, path_folder_data, requred_columns_first_sheet,
                                                                    requred_columns_second_sheet)
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if len(file_error_df) != 0:
                    continue
                print(file)
                name_file = file.split('.xlsx')[0]
                # Обрабатываем данные с листа Выпуск-СПО
                df_first_sheet = pd.read_excel(f'{path_folder_data}/{file}',
                                               sheet_name=check_required_dct['Выпуск-СПО']['Реальное название листа'],
                                               skiprows=check_required_dct['Выпуск-СПО'][
                                                   'Количество строк заголовка'])  # Считываем прошедший базовую проверку файл
                if len(df_first_sheet) == 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Лист Выпуск-СПО не заполнен']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df_first_sheet.columns = list(map(str, df_first_sheet.columns))  # делаем названия колонок строковыми
                # Если есть колонка с нулем где записан регион или название техникума то удаляем
                if '0' in df_first_sheet.columns:
                    df_first_sheet.drop(columns=['0'], inplace=True)
                # удаляем строки с суммами
                df_first_sheet = df_first_sheet[df_first_sheet['1'].notna()]

                # Приводим все колонки кроме первой к инту
                df_first_sheet[requred_columns_first_sheet[1:-1]] = df_first_sheet[
                    requred_columns_first_sheet[1:-1]].applymap(convert_to_int)
                # очищаем первую колонку от пробельных символов вначаче и конце
                df_first_sheet['1'] = df_first_sheet['1'].apply(lambda x: x.strip() if isinstance(x, str) else x)
                """
                Начинаем проверку содержания файла и арифметических формул
                """
                # Проверяем правильность заполнения колонки 1
                df_first_sheet['Код'] = df_first_sheet['1'].apply(extract_code_nose)  # очищаем от текста в кодах
                if 'error' in df_first_sheet['Код'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 1 Код и наименование профессии/специальности.Вместо кода присутствует дата, и т.п. проверьте правильность заполнения колонки 1!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df_first_sheet.drop(columns=['Код'], inplace=True)

                # Убираем лишние колонки перед отправкой датафрейма на проверку
                checked_first_sheet_df = df_first_sheet.copy().drop(columns=text_required_columns_first_sheet)
                checked_first_sheet_df.drop(
                    columns=[name_column for name_column in checked_first_sheet_df.columns if 'Unnamed' in name_column],
                    inplace=True)

                file_error_df = check_error_mon_grad_spo(checked_first_sheet_df,
                                                         name_file, check_required_dct['Выпуск-СПО'][
                                                             'Количество строк заголовка'] + 1)  # отправляем на проверку без 73 и Unnamed
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if len(file_error_df) != 0:
                    continue

                """
                Обрабатываем лист Выпуск- Целевое
                """
                # Обрабатываем данные с листа Выпуск-Целевое
                df_second_sheet = pd.read_excel(f'{path_folder_data}/{file}',
                                                sheet_name=check_required_dct['Выпуск-Целевое'][
                                                    'Реальное название листа'],
                                                skiprows=check_required_dct['Выпуск-Целевое'][
                                                    'Количество строк заголовка'], dtype=str)

                # проводим обработку только если лист заполнен данными
                if len(df_second_sheet) != 0:
                    # удаляем строки с суммами и пустые строки
                    df_second_sheet = df_second_sheet[df_second_sheet['1'].notna()]
                    lst_int_columns_second_sheet = ['5', '6', '7', '8', '9', '10', '11', '12']
                    df_second_sheet[lst_int_columns_second_sheet] = df_second_sheet[
                        lst_int_columns_second_sheet].applymap(convert_to_int)
                    file_error_df = check_error_mon_grad_target(df_first_sheet.copy(), df_second_sheet.copy(),
                                                                name_file,
                                                                check_required_dct['Выпуск-Целевое'][
                                                                    'Количество строк заголовка'] + 1)  # отправляем на проверку без 73 и Unnamed
                    error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                    if len(file_error_df) != 0:
                        continue
                    df_second_sheet.insert(0, 'Название файла', name_file)  # добавляем колонку для названия файла
                    main_second_df = pd.concat([main_second_df, df_second_sheet], axis=0, ignore_index=True)

                # Заполняем словарь данными
                # перебираем список словарей
                # Создание словаря для хранения данных файла
                code_spec = [spec for spec in
                             df_first_sheet['1'].unique()]  # получаем список специальностей которые есть в файле
                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for spec in code_spec:
                    spec_dict[spec] = {key: 0 for key in requred_columns_first_sheet}

                high_level_dct[name_file] = copy.deepcopy(spec_dict)

                # превращаем в словарь
                first_sheet_lst_dct = df_first_sheet.to_dict(orient='records')

                # добавляем данные
                for dct in first_sheet_lst_dct:
                    high_level_dct[name_file][dct['1']].update(dct)

        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)

        wb_check_tables = create_check_tables_mon_grad(high_level_dct)  # проверяем данные по каждой специальности
        if len(wb_check_tables.sheetnames) != 0:
            wb_check_tables.save(
                f'{path_result_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')
        else:
            empty_wb = openpyxl.Workbook()
            empty_wb.save(f'{path_result_folder}/Отсутствуют файлы без ошибок {current_time}.xlsx')
        # Обрабатываем конечный файл
        # Создаем словарь в котором будут храниться словари по специальностям
        code_spec_dct = {}

        # Создаем ключи для словаря
        for poo, spec_data in high_level_dct.items():
            for name_spec, row in spec_data.items():
                code_spec_dct[name_spec] = {'1': name_spec}
                code_spec_dct[name_spec].update({key: 0 for key in row.keys() if key != '1'})
                code_spec_dct[name_spec]['73'] = ''
        # Суммируем значения из словаря
        for poo, spec_data in high_level_dct.items():
            for name_spec, row in spec_data.items():
                for key, value in row.items():
                    if 'Unnamed' not in key and key != '1':
                        if key == '73':
                            code_spec_dct[name_spec][key] += f';{str(value)}'
                        else:
                            code_spec_dct[name_spec][key] += value
        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_code_spec_dct = sorted(code_spec_dct.items())
        code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}
        # Создаем датафрейм
        out_df = pd.DataFrame.from_dict(code_spec_dct, orient='index')
        # Удаляем лишние колонки
        out_df.drop(columns=[name_column for name_column in out_df.columns if 'Unnamed' in name_column], inplace=True)

        # Сохраняем файл
        wb = openpyxl.Workbook()
        wb.create_sheet('Выпуск-СПО', index=0)
        wb.create_sheet('Выпуск-Целевое', index=1)

        # Записываем в файл
        for r in dataframe_to_rows(out_df, index=False, header=True):
            wb['Выпуск-СПО'].append(r)
        wb['Выпуск-СПО'].column_dimensions['A'].width = 40

        for r in dataframe_to_rows(main_second_df, index=False, header=True):
            wb['Выпуск-Целевое'].append(r)
        wb['Выпуск-Целевое'].column_dimensions['A'].width = 30

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(f'{path_result_folder}/Итоговый файл от {current_time}.xlsx')

        # out_df.to_excel(f'{path_result_folder}/Итоговый файл от {current_time}.xlsx', index=False)

        # Сохраняем файл с ошибками
        error_df.to_excel(f'{path_result_folder}/Ошибки {current_time}.xlsx', index=False)

    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    except Exception as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'При обработке файла {name_file} возникла ошибка {e.args} !!!\n'
                             f'Проверьте файл на соответствие шаблону')
    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                                'Данные успешно обработаны.Ошибок не обнаружено')


if __name__ == '__main__':
    main_data_folder = 'c:/Users/1/PycharmProjects/Dodger_2023/data/Мониторинг занятости выпускников/Файлы/'
    main_result_folder = 'c:/Users/1/PycharmProjects/Dodger_2023/data/Мониторинг занятости выпускников/Результат'
    prepare_graduate_employment(main_data_folder, main_result_folder)

    print('Lindy Booth')

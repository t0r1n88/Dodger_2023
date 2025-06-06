# -*- coding: utf-8 -*-
"""
Скрипт для обработки данных Формы 2 нозология (15 строк) мониторинга занятости выпускников
"""
import numpy as np

from cass_support_functions import * # импортируем вспомогательные функции и исключения
from cass_check_functions import (check_error_main_may_2025,check_error_target_may_2025,
                                  extract_code_nose,check_data,create_check_tables_may_2025,create_check_tables_target_may_2025,check_dight,check_kpp)
# проверка основного листа
import pandas as pd
import copy
import os
import warnings
from tkinter import messagebox
import time
pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class NotCorrectFile(Exception):
    """
    Исключения для обработки случая когда нет ни одного корректного файла
    """
    pass


def prepare_may_2025(path_folder_data:str,path_to_end_folder):
    """
    Фугкция для обработки данных мониторинга 2025
    :return:
    """
    try:
        # создаем словарь верхнего уровня для каждого поо
        high_level_dct = {}
        # создаем базовый датафрейм для целевиков
        main_target_df = pd.DataFrame(columns=['Наименование файла','1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11',
                                       '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23',
                                       '24', '25', '26', '27'])
        # создаем словарь верхнего уровня для хранения пары ключ значение где ключ это код специальности а значение- код и наименование
        dct_code_and_name = dict()

        # создаем словарь верхнего уровня для хранения пары ключ значение где ключ это код специальности а значение- код и наименование
        target_dct_code_and_name = dict()


        # создаем датафрейм для регистрации ошибок
        error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

        # Создаем датафрейм для контроля дублирующихся кодов специальностей
        main_dupl_df = pd.DataFrame(columns=['Название файла','Полное наименование'])

        for file in os.listdir(path_folder_data):
            if not file.startswith('~$') and not file.endswith('.xlsx'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                # получаем название первого листа
                temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}', read_only=True)
                lst_temp_sheets = temp_wb.sheetnames  # получаем листы в файле
                temp_wb.close()
                if '1. Форма сбора' not in lst_temp_sheets:  # проверяем наличие листа с названием в файле
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не найден лист с названием 1. Форма сбора !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                if '3. Целевики' not in lst_temp_sheets:  # проверяем наличие листа с названием в файле
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не найден лист с названием 3. Целевики !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # Считываем данные листа с общими данными
                try:
                    df = pd.read_excel(f'{path_folder_data}/{file}', sheet_name='1. Форма сбора', skiprows=2, dtype=str)
                except:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не удалось прочитать файл! Отключите фильтры в файле, проверьте файл на целостность и соответствие шаблону']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df.columns = list(map(str, df.columns))  # делаем названия колонок строковыми
                # создаем множество колонок наличие которых мы проверяем
                check_cols = {'1', '1.1', '1.2', '2', '3', '3.1', '3.2', '3.3',
                               '4', '4.1', '4.2', '4.3', '5', '6', '7', '8', '9', '10', '11', '12',
                               '13', '14', '15', '16', '17', '18'}
                diff_cols = check_cols.difference(list(df.columns))
                if len(diff_cols) != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{diff_cols}',
                                                        'На листе 1. Форма сбора не найдены указанные колонки. Проверьте соответствие файла шаблону с сайта, строка с номерами колонок должна быть на третьй строке ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # Считываем данные листа с целевиками
                try:
                    target_df = pd.read_excel(f'{path_folder_data}/{file}', sheet_name='3. Целевики', skiprows=2, dtype=str)
                except:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не удалось прочитать файл! Отключите фильтры в файле, проверьте файл на целостность и соответствие шаблону']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                target_df.columns = list(map(str, target_df.columns))  # делаем названия колонок строковыми
                # создаем множество колонок наличие которых мы проверяем
                check_cols_target = {'1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11',
                                       '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23',
                                       '24', '25', '26', '27'}
                diff_cols_target = check_cols_target.difference(list(target_df.columns))
                if len(diff_cols_target) != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{diff_cols_target}',
                                                        'На листе 3. Целевики не найдены указанные колонки. Проверьте соответствие файла шаблону с сайта, строка с номерами колонок должна быть на третьй строке ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df = df[df['1'].notna()]  # убираем возможные наны из за лишних строк
                df = df[df['1'].str.strip() != ''] # убираем строки где только пробелы
                target_df = target_df[target_df['1'].notna()]  # убираем возможные наны из за лишних строк
                target_df = target_df[target_df['1'].str.strip() != ''] # убираем строки где только пробелы

                # Проверяем на заполнение лист с общими данными
                if len(df) == 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'Отсутствуют коды специальностей в колонке 1',
                                                        'Лист 1. Форма сбора пустой. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # отсекаем возможный первый столбец с данными ПОО,начинаем датафрейм с колонки 1 и отсекаем колонки с проверками
                df = df.loc[:, '1':'18']
                target_df = target_df.loc[:, '1':'27']


                # проверяем на арифметические ошибки основной лист
                file_error_df = check_error_main_may_2025(df.copy(), name_file)

                # проверяем на арифметические ошибки лист целевиков
                file_error_target_df = check_error_target_may_2025(target_df.copy(), name_file)

                # Добавляем в датафрейм для проверки на дубликаты
                temp_dupl_df = df['1'].to_frame()
                temp_dupl_df['Название файла'] = name_file
                temp_dupl_df = temp_dupl_df.reindex(columns=['Название файла', '1'])
                temp_dupl_df.columns = ['Название файла', 'Полное наименование']
                temp_dupl_df.drop_duplicates(subset='Полное наименование', inplace=True)

                main_dupl_df = pd.concat([main_dupl_df, temp_dupl_df])



                # добавляем в словарь в полные имена из кода и наименования

                # делаем копию для перекрестной проверки с листом целевиков
                cross_check_df = df.copy()


                for full_name in df['1'].tolist():
                    code = extract_code_nose(full_name)  # получаем только цифры
                    dct_code_and_name[code] = full_name
                # очищаем от текста чтобы названия листов не обрезались
                df['1'] = df['1'].apply(extract_code_nose)  # очищаем от текста в кодах

                if 'error' in df['1'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 1 Код и наименование профессии/специальности на листе 1. Форма сбора.Вместо кода присутствует дата,в колонке есть слово Итого и т.п. проверьте правильность заполнения колонки 1!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    file_error_df = pd.concat([file_error_df, temp_error_df], axis=0, ignore_index=True)



                # добавляем в словарь в полные имена из кода и наименования
                for full_name in target_df['1'].tolist():
                    code = extract_code_nose(full_name)  # получаем только цифры
                    target_dct_code_and_name[code] = full_name
                # очищаем от текста чтобы названия листов не обрезались
                # Проверяем правильность написания кодов
                checked_code_target_df = target_df.copy()

                checked_code_target_df['1'] = checked_code_target_df['1'].apply(extract_code_nose)  # очищаем от текста в кодах
                if 'error' in checked_code_target_df['1'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке 1 Код и наименование профессии/специальности на листе 3. Целевики.Вместо кода присутствует дата, слово Итого и т.п. проверьте правильность заполнения колонки 1!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    file_error_target_df = pd.concat([file_error_target_df, temp_error_df], axis=0, ignore_index=True)



                # добавляем в основной файл с ошибками
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                error_df = pd.concat([error_df, file_error_target_df], axis=0, ignore_index=True)
                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                if file_error_target_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue



                # для ошибок в инн кпп или количестве
                second_target_error__df = pd.DataFrame(
                    columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
                # Обрабатываем лист целевиков
                if len(target_df) != 0:
                    second_target_error__df = pd.DataFrame(
                        columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

                    # Делаем проверку наличие специальностей в обоих листах
                    code_df = set(cross_check_df['1'].unique()) # специальности которые есть на главном листе
                    code_target_df = set(target_df['1'].unique()) # специальности целевиков
                    diff_code = code_target_df.difference(code_df)
                    if len(diff_code) != 0:
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            f'На листе 3. Целевики обнаружены специальности {diff_code} которых нет на листе 1. Форма сбора!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        second_target_error__df = pd.concat([second_target_error__df, temp_error_df], axis=0, ignore_index=True)



                    # проверяем превышение количества целевиков
                    cross_check_df['2'] = cross_check_df['2'].apply(check_data)
                    cross_code_df = cross_check_df.groupby(by='1').agg({'2':'sum'})
                    target_df['5'] = target_df['5'].apply(check_data)
                    target_code_df = target_df.groupby(by='1').agg({'5':'sum'})

                    check_target_df = target_code_df.join(cross_code_df)
                    check_target_df = check_target_df.reset_index()
                    check_target_df.columns= ['Профессия/специальность','Количество целевиков','Общий выпуск']
                    check_target_df = check_target_df.reindex(columns=['Профессия/специальность','Общий выпуск','Количество целевиков'])

                    check_target_df['Результат'] = check_target_df['Общий выпуск'] >= check_target_df['Количество целевиков']
                    check_target_df['Результат'] = check_target_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
                    # получаем датафрейм с ошибками и извлекаем индекс
                    temp_target_error_df = check_target_df[check_target_df['Результат'] == 'Неправильно']
                    if len(temp_target_error_df) != 0:
                        error_code_target = temp_target_error_df['Профессия/специальность'].tolist()
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            f'На листе 3. Целевики обнаружены специальности {error_code_target} численность выпускников которых превышает количество для этих специальностей указанное на листе 1. Форма сбора на листе 1. Форма сбора!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        second_target_error__df = pd.concat([second_target_error__df, temp_error_df], axis=0, ignore_index=True)


                    # Проверяем длину ИНН и КПП
                    inn_kp_df = target_df.copy()
                    inn_kp_df=inn_kp_df.applymap(str)
                    # Проверяем длину ИНН
                    inn_kp_df['Результат_ИНН'] = inn_kp_df['3'].apply(lambda x:check_dight(x,(10,12)))
                    # получаем датафрейм с ошибками и извлекаем индекс
                    inn_df = inn_kp_df[inn_kp_df['Результат_ИНН'] == 'Неправильно'].reset_index()
                    if len(inn_df) !=0:
                        # создаем датафрейм дял добавления в ошибки
                        temp_error_df = pd.DataFrame(
                            columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
                        # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
                        raw_lst_index = inn_df['index'].tolist()  # делаем список
                        finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
                        finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
                        temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
                        temp_error_df['Название файла'] = name_file
                        temp_error_df[
                            'Описание ошибки'] = 'Количество цифр в ИНН (колонка 3 на листе 3. Целевики) на указанной строке не равно 10 (юрлица) или 12 (в случае ИП)'
                        second_target_error__df = pd.concat([second_target_error__df, temp_error_df], axis=0, ignore_index=True)

                    # Проверяем длину КПП
                    inn_kp_df['Результат_КПП'] = inn_kp_df[['3','4']].apply(lambda x:check_kpp(x),axis=1)
                    # получаем датафрейм с ошибками и извлекаем индекс
                    kpp_df = inn_kp_df[inn_kp_df['Результат_КПП'] == 'Неправильно'].reset_index()
                    if len(kpp_df) !=0:
                        # создаем датафрейм дял добавления в ошибки
                        temp_error_df = pd.DataFrame(
                            columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
                        # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
                        raw_lst_index = kpp_df['index'].tolist()  # делаем список
                        finish_lst_index = list(map(lambda x: x + 4, raw_lst_index))
                        finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
                        temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
                        temp_error_df['Название файла'] = name_file
                        temp_error_df[
                            'Описание ошибки'] = 'Количество цифр в КПП (колонка 4 на листе 3. Целевики) на указанной строке не равно 9 для юридических лиц или указан КПП для ИП'
                        second_target_error__df = pd.concat([second_target_error__df, temp_error_df], axis=0, ignore_index=True)





                if second_target_error__df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, second_target_error__df], axis=0, ignore_index=True)
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                # После всех проверок добавляем в данные.
                # Создание словаря для хранения данных файла
                code_spec = [spec for spec in
                             df['1'].unique()]  # получаем список специальностей которые есть в файле
                # Названия колонок
                column_cat = [f'Колонка {i}' for i in range(1, 24)]

                spec_dct = {key: 0 for key in column_cat}

                high_level_dct[name_file] = {code: copy.deepcopy(spec_dct) for code in code_spec}

                # Создание словаря для хранения данных с основного листа
                for row in df.itertuples():
                    data_row = row[4:27]  # получаем срез с нужными данными колонки в которых есть числа
                    for idx_col, value in enumerate(data_row, start=1):
                        high_level_dct[name_file][row[1]][f'Колонка {idx_col}'] += check_data(value)

                # добавляем целевиков
                target_df.insert(0, 'Наименование файла', name_file)
                main_target_df = pd.concat([main_target_df, target_df])






        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)

            # Сохраняем ошибки
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ Май 2025 от {current_time}.xlsx')

        if len(high_level_dct) == 0:
            raise NotCorrectFile

        # Общий список всех специальностей что встречаются
        main_dupl_df.sort_values(by='Полное наименование', inplace=True)
        # Уникальный специальности
        unique_spec_df = main_dupl_df.drop_duplicates(subset='Полное наименование')
        unique_spec_df.drop(columns='Название файла', inplace=True)

        # Дубликаты
        dupl_df = unique_spec_df.copy()
        dupl_df['Полное наименование'] = dupl_df['Полное наименование'].apply(str.strip)
        dupl_df = dupl_df.drop_duplicates(subset=['Полное наименование'])

        dupl_df['Код специальности'] = dupl_df['Полное наименование'].apply(extract_code_nose)
        out_dupl_df = dupl_df[dupl_df['Код специальности'].duplicated(keep=False)]  # получаем дубликаты

        # Свод
        svod_df_spec = main_dupl_df['Полное наименование'].value_counts().to_frame()
        svod_df_spec = svod_df_spec.reset_index()
        svod_df_spec.columns = ['Специальность/профессия', 'Количество ПОО в которых она есть']
        svod_df_spec.sort_values(by='Специальность/профессия',ascending=True,inplace=True)

        with pd.ExcelWriter(f'{path_to_end_folder}/Дублирующиеся коды от {current_time}.xlsx') as writer:
            out_dupl_df.to_excel(writer, sheet_name='Дублирующиеся коды', index=False)
            unique_spec_df.to_excel(writer, sheet_name='Уникальные', index=False)
            svod_df_spec.to_excel(writer, sheet_name='Свод по количеству', index=False)
            main_dupl_df.to_excel(writer, sheet_name='Полный список', index=False)


        # # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            # ПОО и словарь со специальностями
            for code_spec, _ in spec.items():
                # специальность и словарь с данными
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dct) for key in all_spec_code}
        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            # ПОО и словарь со специальностями
            for code_spec, data in spec.items():
                # специальность и словарь с данными
                for col, col_data in data.items():
                    itog_df[code_spec][col] += col_data

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        finish_df = pd.DataFrame.from_dict(itog_df, orient='index')
        finish_df = finish_df.reset_index()
        finish_df.insert(1,'1.1',np.nan)
        finish_df.insert(2,'1.2',np.nan)
        finish_df.columns = ['1', '1.1', '1.2', '2', '3', '3.1', '3.2', '3.3',
                               '4', '4.1', '4.2', '4.3', '5', '6', '7', '8', '9', '10', '11', '12',
                               '13', '14', '15', '16', '17', '18']

        finish_df['1'] = finish_df['1'].apply(
            lambda x: dct_code_and_name[x])  # делаем код чтобы отображался код и наименование

        wb_check_tables = create_check_tables_may_2025(high_level_dct)  # проверяем данные по каждой специальности
        if 'Sheet' in wb_check_tables.sheetnames:
            del wb_check_tables['Sheet']
        wb_check_tables.save(
            f'{path_to_end_folder}/Данные для проверки заполнения общего выпуска от {current_time}.xlsx')




        """
        Обработка листа с целевиками
        """
        if len(main_target_df) != 0:
            wb_check_target_tables = create_check_tables_target_may_2025(main_target_df.copy())  # проверяем данные по каждой специальности
            if 'Sheet' in wb_check_target_tables.sheetnames:
                del wb_check_target_tables['Sheet']
            wb_check_target_tables.save(
                f'{path_to_end_folder}/Данные для проверки заполнения целевиков от {current_time}.xlsx')


        finish_df.sort_values(by='1',ascending=True,inplace=True)
        main_target_df.sort_values(by='1',ascending=True,inplace=True)

        if len(main_target_df) != 0:
            # Создаем третий лист для свода по целевикам
            svod_target = main_target_df.groupby(by='1').agg({'5':'sum'})
            svod_target = svod_target.reset_index()
            svod_target.columns = ['Специальность/профессия','Количество']
            svod_target.sort_values(by='Специальность/профессия', ascending=True, inplace=True)
            with pd.ExcelWriter(f'{path_to_end_folder}/Мониторинг Май 2025 от {current_time}.xlsx') as writer:
                finish_df.to_excel(writer, sheet_name='1. Форма сбора', index=False)
                main_target_df.to_excel(writer, sheet_name='3. Целевики', index=False)
                svod_target.to_excel(writer, sheet_name='Свод по целевикам', index=False)

        else:
            with pd.ExcelWriter(f'{path_to_end_folder}/Мониторинг Май 2025 от {current_time}.xlsx') as writer:
                finish_df.to_excel(writer, sheet_name='1. Форма сбора', index=False)
                main_target_df.to_excel(writer, sheet_name='3. Целевики', index=False)


    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    except NotCorrectFile as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено ни одного файла БЕЗ ошибок. Откройте файл с ошибками и исправьте указанные ошибки!')
    # except:
    #     messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
    #                          f'При обработке файла {name_file} возникла ошибка !!!\n'
    #                          f'Проверьте файл на соответствие шаблону исходной формы 2 нозологии)')

    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                                'Данные успешно обработаны.Ошибок не обнаружено')













if __name__ == '__main__':
    main_path_folder_data = 'data/Май Мониторинг'
    main_path_end_folder = 'data/РЕЗУЛЬТАТ'
    prepare_may_2025(main_path_folder_data,main_path_end_folder)
    print('Lindy Booth')




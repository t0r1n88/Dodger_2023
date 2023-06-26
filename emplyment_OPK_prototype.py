# -*- coding: UTF-8 -*-
import pandas as pd
import numpy as np
import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
# pd.options.mode.chained_assignment = None  # default='warn'
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import random


def check_data(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0


def check_data_note(cell):
    if cell is np.nan:
        return 'Не заполнено'
    return str(cell)

def extract_code_full(value):
    """
    Функция для извлечения кода специальности из ячейки в которой соединены и код и название специалньости
    """
    value = str(value)
    re_code = re.compile('\d{2}?[.]\d{2}?[.]\d{2}')  # создаем выражение для поиска кода специальности
    result = re.search(re_code,value)
    if result:
        return result.group()
    else:
        return 'Не найден код специальности'

"""
Проверка ошибок
"""
def check_error_opk(df1:pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки таблиц по ОПК
    :param df1: датафрейм форма 1
    :param df2: датафрейм форма 2
    :param name_file: название обрабатываемого файла
    :param tup_correct: значаение корректировки
    :return: датафрейм с найденными ошибками
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # делаем датафреймы для простых проверок
    df1 = df1.iloc[:, 5:77]
    df1 = df1.applymap(check_data)

    # получаем количество датафреймов
    quantity = df1.shape[0] // 2
    # счетчик для обработанных строк
    border = 0
    for i in range(1, quantity + 1):
        temp_df = df1.iloc[border:border + 2, :]
    # Проверяем корректность заполнения формы 1
        first_error_opk = check_horizont_sum_opk_all(temp_df.copy(),name_file,tup_correct) # проверяем сумму по строкам
        error_df = pd.concat([error_df, first_error_opk], axis=0, ignore_index=True)

        # проверяем условие  по колонкам строка 02 не должна быть больше строки 01
        second_error_opk = check_vertical_opk_all(temp_df.copy(),border,name_file,tup_correct)
        error_df = pd.concat([error_df, second_error_opk], axis=0, ignore_index=True)

        border += 2




    return error_df

def check_cross_error_opk(df1:pd.DataFrame,df2:pd.DataFrame, name_file, tup_correct):
    """
    Функция для проверки значений между формой 1 и формой 2
    :param df1:
    :param df2:
    :param name_file:
    :param tup_correct:
    :return:
    """
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    df1['08'] = df1['08'].apply(check_data)
    df1['39'] = df1['39'].apply(check_data) # приводим к инту

    group_df1 = df1.groupby(['03']).agg({'08':sum,'39':sum}) # группируем
    group_df1 = group_df1.reset_index() # переносим индексы
    group_df1.columns = ['Специальность','Трудоустроено в ОПК','Будут трудоустроены в ОПК']

    # приводим колонку 2 формы с числов выпускников к инту
    df2['04'] = df2['04'].apply(check_data)

    # Проверяем заполенение 2 формы, есть ли там вообще хоть что то
    quantity_now = group_df1['Трудоустроено в ОПК'].sum() # сколько трудоустроено сейчас
    quantity_future = group_df1['Будут трудоустроены в ОПК'].sum() # сколько будут трудоустроены
    # проверяем заполнение формы 2
    if (quantity_now != 0 or quantity_future !=0) and df2.shape[0] == 0:
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                            'В форме 1 есть выпускники трудоустроенные или которые будут трудоустроены в ОПК,\n'
                                            ' при этом форма 2 не заполнена. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
        return error_df

    cross_first_error_df =check_cross_first_error_df(group_df1.copy(),df2.copy(),name_file,)
    error_df = pd.concat([error_df, cross_first_error_df], axis=0, ignore_index=True)




    return error_df

def check_cross_first_error_df(df1:pd.DataFrame,df2:pd.DataFrame, name_file):
    """
    Функция для првоерки соответствия количества указанных в форме 1 трудоустроенных и списка в форме 2
    проверки 1 и 2

    :param df1:
    :param df2:
    :param name_file:
    :return:
    """
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # проверяем наличие незаполненных ячеек в колонках 05 06
    etalon_05 = {'уже трудоустроены','будут трудоустроены'} # эталонный состав колонки 05
    etalon_06 = {'заключили договор о целевом обучении','нет'} # эталонный состав колонки 05
    # получаем состав колонок
    st_05 = set(df2['05'].unique())
    st_06 = set(df2['06'].unique())
    if not (st_05.issubset(etalon_05) or st_06.issubset(etalon_06)):
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                            'В колонках 05 или 06 формы 2 есть незаполненные ячейки,\n'
                                            ' или значения отличающиеся от требуемых. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    # Проверяем

    # создаем 2 датафрейма, по колонке 05 трудоустроены и будут трудоустроены
    empl_now_df = df2[df2['05'] =='уже трудоустроены'] # те что уже трудоустроены
    empl_future_df = df2[df2['05'] =='будут трудоустроены'] # те что будут трудоустроены

    # проводим группировку
    empl_now_df_group = empl_now_df.groupby(['02']).agg({'04':sum})
    empl_future_df_group = empl_future_df.groupby(['02']).agg({'04':sum})

    df1_future = df1[df1['Будут трудоустроены в ОПК'] !=0] # отбираем в форме 2 специальности по которым есть будущие трудоустроены выпускники
    check_df = empl_future_df_group.merge(df1_future,how='outer',left_on='02',right_on='Специальность')
    check_df['Результат'] = check_df['04'] == check_df['Будут трудоустроены в ОПК']
    check_df = check_df[~check_df['Результат']]

    print(check_df)
    # записываем где есть ошибки
    for row in check_df.itertuples():
        temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{row[2]} не совпадают данные !!! по форме 1 для этой специальности будут трудоустроено {int(row[4])}'
                                                            f' в форме 2 по этой специальности найдено {int(row[1])}',
                                            'Несовпадает количество выпускников которые будут трудоустроены в форме 1 и в форме 2. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                              'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

    #check_df.to_excel('trs.xlsx',index=False)








    return error_df














def check_horizont_sum_opk_all(df:pd.DataFrame,name_file,tup_correct):
    """
    Функция для проверки простой горизонтальной суммы 06 = 07 +32,33,34,35,36,37,38,63,64,65,66:77
    :param df:
    :param name_file:
    :param tup_correct:
    :return:
    """
    # получаем строку диапазона
    first_correct = tup_correct[0]
    # конвертируем в инт
    drop_lst = ['08','09','10','11','12','13','14','15','16','17','18','19','20',
                '21','22','23','24','25','26','27','28','29','30','31','39','40',
                '41','42','43','44','45','46','47','48','49','50','51','52','53',
                '54','55','56','57','58','59','60','61','62']

    # удаляем колонки лишние колонки
    df.drop(columns=drop_lst,inplace=True)

    # # получаем сумму колонок
    df['Сумма'] = df.iloc[:,1:].sum(axis=1)
    # # Проводим проверку
    df['Результат'] = df['06'] == df['Сумма']
    # # заменяем булевые значения на понятные
    df['Результат'] = df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')
    # # получаем датафрейм с ошибками и извлекаем индекс
    df = df[df['Результат'] == 'Неправильно'].reset_index()
    # # создаем датафрейм дял добавления в ошибки
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = df['index'].tolist()  # делаем список
    finish_lst_index = list(map(lambda x: x + first_correct, raw_lst_index))
    finish_lst_index = list(map(lambda x: f'Строка {str(x)}', finish_lst_index))
    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: гр. 06 = гр.07 + сумма(всех колонок за исключением распределения по отраслям)'
    return temp_error_df


def check_vertical_opk_all(df:pd.DataFrame,border,name_file,tup_correct):
    """
    Функция для проверки условия Количество целевиков не должно быть больше чем количество выпускников
    :param df:
    :param name_file:
    :param tup_correct:
    :return:
    """
    first_correct = tup_correct[0]
    second_correct = tup_correct[1]
    foo_df = pd.DataFrame(columns=['01', '02'])

    # Добавляем данные в датафрейм
    foo_df['01'] = df.iloc[0, :]
    foo_df['02'] = df.iloc[1, :]
    foo_df['Результат'] = foo_df['02'] <= foo_df['01']
    foo_df['Результат'] = foo_df['Результат'].apply(lambda x: 'Правильно' if x else 'Неправильно')

    foo_df = foo_df[foo_df['Результат'] == 'Неправильно'].reset_index()
    temp_error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])
    # обрабатываем индексы строк с ошибками чтобы строки совпадали с файлом excel
    raw_lst_index = foo_df['index'].tolist()  # делаем список
    finish_lst_index = list(
        map(lambda x: f'Диапазон строк {border + first_correct} - {border + second_correct}, колонка {str(x)}', raw_lst_index))

    temp_error_df['Строка или колонка с ошибкой'] = finish_lst_index
    temp_error_df['Название файла'] = name_file
    temp_error_df['Описание ошибки'] = 'Не выполняется условие: стр. 02 <= стр. 01 '
    return temp_error_df




def create_check_tables_opk(high_level_dct: dict):
    """
        Функция для создания файла с данными по каждой специальности
        """
    # Создаем словарь в котором будут храниться словари по специальностям
    code_spec_dct = {}

    # инвертируем словарь так чтобы код специальности стал внешним ключом а названия файлов внутренними
    for poo, spec_data in high_level_dct.items():
        for code_spec, data in spec_data.items():
            if code_spec not in code_spec_dct:
                code_spec_dct[code_spec] = {f'{poo}': high_level_dct[poo][code_spec]}
            else:
                code_spec_dct[code_spec].update({f'{poo}': high_level_dct[poo][code_spec]})

    # Сортируем получившийся словарь по возрастанию для удобства использования
    sort_code_spec_dct = sorted(code_spec_dct.items())
    code_spec_dct = {dct[0]: dct[1] for dct in sort_code_spec_dct}

    # Создаем файл
    wb = openpyxl.Workbook()
    # Создаем листы
    for idx, code_spec in enumerate(code_spec_dct.keys()):
        if code_spec != 'nan':
            wb.create_sheet(title=code_spec, index=idx)

    for code_spec in code_spec_dct.keys():
        if code_spec != 'nan':
            temp_code_df = pd.DataFrame.from_dict(code_spec_dct[code_spec], orient='index')

            temp_code_df = temp_code_df.stack()
            # название такое выбрал потому что было лень заменять значения из блокнота юпитера
            temp_code_df = temp_code_df.to_frame()

            temp_code_df['Суммарный выпуск 2023 г.'] = temp_code_df[0].apply(lambda x: x.get('Колонка 6'))
            temp_code_df[
                'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 7'))
            temp_code_df['Трудоустроены на предприятия оборонно-промышленного комплекса*'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 8'))
            temp_code_df['Трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 9'))
            temp_code_df['Трудоустроены на предприятия сельского хозяйства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 10'))
            temp_code_df['Трудоустроены на предприятия металлургии'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 11'))
            temp_code_df['Трудоустроены на предприятия железнодорожного транспорта'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 12'))
            temp_code_df['Трудоустроены на предприятия легкой промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 13'))
            temp_code_df['Трудоустроены на предприятия химической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 14'))
            temp_code_df['Трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 15'))
            temp_code_df['Трудоустроены на предприятия фармацевтической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 16'))
            temp_code_df['Трудоустроены на предприятия отрасли информационных технологий'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 17'))
            temp_code_df['Трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 18'))
            temp_code_df[
                'Трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 19'))
            temp_code_df['Трудоустроены на предприятия транспортной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 20'))
            temp_code_df['Трудоустроены на предприятия горнодобывающей отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 21'))
            temp_code_df[
                'Трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 22'))
            temp_code_df['Трудоустроены на предприятия лесной промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 23'))
            temp_code_df['Трудоустроены на предприятия строительной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 24'))
            temp_code_df[
                'Трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 25'))
            temp_code_df['Трудоустроены на предприятия индустрии робототехники'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 26'))
            temp_code_df['Трудоустроены на предприятия в отрасли образования'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 27'))
            temp_code_df['Трудоустроены на предприятия в медицинской отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 28'))
            temp_code_df[
                'Трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 29'))
            temp_code_df['Трудоустроены на предприятия в отрасли искусства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 30'))
            temp_code_df['Трудоустроены на предприятия иная отрасль'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 31'))

            temp_code_df['Индивидуальные предприниматели'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 32'))
            temp_code_df['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессиональный доход)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 33'))
            temp_code_df['Продолжили обучение'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 34'))
            temp_code_df['Проходят службу в армии по призыву'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 35'))
            temp_code_df[
                'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 36'))
            temp_code_df['Находятся в отпуске по уходу за ребенком'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 37'))
            temp_code_df[
                'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 38'))
            temp_code_df['Будут трудоустроены на предприятия оборонно-промышленного комплекса* '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 39'))
            temp_code_df['Будут трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 40'))
            temp_code_df['Будут трудоустроены на предприятия сельского хозяйства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 41'))
            temp_code_df['Будут трудоустроены на предприятия металлургии'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 42'))
            temp_code_df['Будут трудоустроены на предприятия железнодорожного транспорта'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 43'))
            temp_code_df['Будут трудоустроены на предприятия легкой промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 44'))
            temp_code_df['Будут трудоустроены на предприятия химической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 45'))
            temp_code_df[
                'Будут трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 46'))
            temp_code_df['Будут трудоустроены на предприятия фармацевтической отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 47'))
            temp_code_df['Будут трудоустроены на предприятия отрасли информационных технологий'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 48'))
            temp_code_df[
                'Будут трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 49'))
            temp_code_df[
                'Будут трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 50'))

            temp_code_df['Будут трудоустроены на предприятия транспортной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 51'))
            temp_code_df['Будут трудоустроены на предприятия горнодобывающей отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 52'))
            temp_code_df[
                'Будут трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 53'))
            temp_code_df['Будут трудоустроены на предприятия лесной промышленности'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 54'))
            temp_code_df['Будут трудоустроены на предприятия строительной отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 55'))
            temp_code_df[
                'Будут трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 56'))
            temp_code_df['Будут трудоустроены на предприятия индустрии робототехники'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 57'))
            temp_code_df['Будут трудоустроены на предприятия в отрасли образования'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 58'))
            temp_code_df['Будут трудоустроены на предприятия в медицинской отрасли'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 59'))
            temp_code_df[
                'Будут трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 60'))
            temp_code_df['Будут трудоустроены на предприятия в отрасли искусства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 61'))
            temp_code_df['Будут трудоустроены на предприятия иная отрасль'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 62'))
            temp_code_df['Будут индивидуальными предпринимателями'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 63'))
            temp_code_df['Будут самозанятыми'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 64'))
            temp_code_df['Будут продолжать обучение'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 65'))
            temp_code_df['Будут призваны в армию'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 66'))
            temp_code_df[
                'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 67'))
            temp_code_df['Будут находиться в отпуске по уходу за ребенком'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 68'))
            temp_code_df['Неформальная занятость (теневой сектор экономики)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 69'))
            temp_code_df[
                'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 70'))
            temp_code_df[
                'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 71'))
            temp_code_df['Иные причины нахождения под риском нетрудоустройства'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 72'))
            temp_code_df['Смерть, тяжелое состояние здоровья'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 73'))
            temp_code_df['Находятся под следствием, отбывают наказание'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 74'))
            temp_code_df[
                'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 75'))
            temp_code_df[
                'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 76'))
            temp_code_df[
                'Иное в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 0,5% выпускников - укажите причины в графе "принимаемые меры"'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 77'))
            temp_code_df[
                'Принимаемые меры по содействию занятости, в том числе по трудоустройству выпускников на предприятия оборонно-промышленного комплекса тезисно - вид меры, охват выпускников мерой'] = \
                temp_code_df[0].apply(lambda x: x.get('Колонка 78'))

            finish_df = temp_code_df.drop([0], axis=1)

            finish_df = finish_df.reset_index()

            finish_df.rename(
                columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
                inplace=True)

            dct = {'Строка 1': 'Всего (общая численность выпускников)',
                   'Строка 2': 'из строки 01: имеют договор о целевом обучении'}

            finish_df['Наименование показателей (категория выпускников)'] = finish_df[
                'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

            for r in dataframe_to_rows(finish_df, index=False, header=True):
                wb[code_spec].append(r)
            wb[code_spec].column_dimensions['A'].width = 20
            wb[code_spec].column_dimensions['B'].width = 40

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    wb.save(f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')


path_folder_data = 'data/ОПК Воронеж'
path_to_end_folder = 'data'

# создаем словарь верхнего уровня для каждого поо
high_level_dct = {}
# создаем датафрейм для регистрации ошибок
error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

for file in os.listdir(path_folder_data):
    if not file.startswith('~$') and file.endswith('.xlsx'):
        name_file = file.split('.xlsx')[0]
        print(name_file)
        # Проверяем наличие листов с названиями Форма 1 и Форма 2
        wb_1 = openpyxl.load_workbook(f'{path_folder_data}/{file}')
        if not {'Форма 1','Форма 2'}.issubset(set(wb_1.sheetnames)):
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Проверьте наличие листов с названием Форма 1 и Форма 2! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue



        df_form1 = pd.read_excel(f'{path_folder_data}/{file}', skiprows=8, dtype=str,
                                 sheet_name='Форма 1')  # общие данные

        # Находим строку с номерами колонок, так как вполне возможно в файле остались примеры
        temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}',read_only=True) # открываем файл в режиме чтения
        temp_ws = temp_wb['Форма 2']
        threshold_form2 = 5
        for row in temp_ws.iter_rows(0): # перебираем значения в первой колонке
            for cell in row:
                if cell.value == '01':
                    threshold_form2 = cell.row
        temp_wb.close() # закрываем файл чтобы потом не было ошибок
        form2_df = pd.read_excel(f'{path_folder_data}/{file}', skiprows=threshold_form2-1, dtype=str,
                                 sheet_name='Форма 2')  # подробные данные по ОПК
        # создаем множество колонок наличие которых мы проверяем
        check_cols = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16',
                      '17',
                      '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33',
                      '34', '35', '36',
                      '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51', '52',
                      '53', '54', '55',
                      '56', '57', '58', '59', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70', '71',
                      '72', '73', '74',
                      '75', '76', '77', '78', '79', '80'}
        if not check_cols.issubset(set(df_form1.columns)):
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Проверьте заголовок таблицы на листе Форма 1.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 9 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue

        # проверяем корректность формы 2
        check_cols_form2 = {'01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11'}
        if not check_cols_form2.issubset(set(form2_df.columns)):
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Проверьте заголовок таблицы на листе Форма 2.Строка с номерами колонок (01,02,03 и т.д. как в исходной форме)\n должна находиться на 5 строке! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue
        df_form1 = df_form1[df_form1['04'] != '03']  # фильтруем строки с проверкой
        form2_df.dropna(axis=0,inplace=True,how='all') # убираем все пустые строки
        form2_df = form2_df[
            ~form2_df['01'].str.contains('Проверка', case=False)]  # фильруем строки с проверкой на листе 2

        df_form1 = df_form1.loc[:, '01':'78']  # отсекаем возможную первую колонку и колонки с примечаниями
        # получаем  часть с данными
        mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
        if mask[0]:# если пустая строка идет первой то удаляем ее и обновляем маску
            df_form1.drop(axis=0,index=0,inplace=True)
            mask = pd.isna(df_form1).all(axis=1)  # создаем маску для строк с пропущенными значениями
        # Находим индекс первой пустой строки, если он есть,получаем список с значениями где есть пустые строки
        empty_row_index = np.where(df_form1.isna().all(axis=1))
        if empty_row_index[0].tolist():
            row_index = empty_row_index[0][0]
            df_form1 = df_form1.iloc[:row_index]
        quantity_spec = df_form1.shape[0] // 2  # получаем количество специальностей в файле

        check_two_rows_spec = df_form1['04'].tolist() == ['01',
                                                          '02'] * quantity_spec  # проверяем чтобы колонка 04 состояла только из 01 и 02
        if not check_two_rows_spec:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Проверьте правильность заполнения колонки 04. Для каждой спец./проф. должны быть  только строки 01 и 02 не считая строки с проверкой. Также возможно под таблицей есть суммирующая строка ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue

        check_code_lst = df_form1['03'].tolist()  # получаем список кодов специальностей
        # Проверка на то чтобы в колонке 03 в первой строке не было пустой ячейки
        if True in mask.tolist():
            if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'В колонке 03 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue

        # Проверка на непрерывность кода специальности, то есть на 2 строки должен быть только один код
        border_check_code = 0  # счетчик обработанных страниц
        quantity_check_code = len(check_code_lst) // 2  # получаем сколько специальностей в таблице
        flag_error_code_spec = False  # чекбокс для ошибки несоблюдения расстояния в 2 строки
        flag_error_space_spec = False  # чекбокс для ошибки заполнения кода специальности пробелом
        for i in range(quantity_check_code):
            # получаем множество отбрасывая np.nan
            temp_set = set([code_spec for code_spec in check_code_lst[border_check_code:border_check_code + 2]])
            if len(temp_set) != 1:
                flag_error_code_spec = True
            if ' ' in temp_set:
                flag_error_space_spec = True
            border_check_code += 2

        if flag_error_space_spec:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Обнаружены ячейки заполненные пробелом в колонке 03 !!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue

        if flag_error_code_spec:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'ДОЛЖЕН БЫТЬ ОДИНАКОВЫЙ КОД СПЕЦИАЛЬНОСТИ НА КАЖДЫЕ 2 СТРОКИ (не считая строки с проверкой)!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue
        """
        ПРОВЕРКИ
        в том числе проверка кода специальности

        """
        tup_correct = (10, 12)  # создаем кортеж  с поправками
        file_error_df = check_error_opk(df_form1.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)

        # проводим кросс проверки между 2 формами
        file_cross_error_df = check_cross_error_opk(df_form1.copy(),form2_df.copy(), name_file, tup_correct)
        error_df = pd.concat([error_df, file_cross_error_df], axis=0, ignore_index=True)

        if file_error_df.shape[0] != 0:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue


        df_form1['03'] = df_form1['03'].apply(extract_code_full)  # очищаем от текста в кодах
        if 'Не найден код специальности' in df_form1['03'].values:
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                'Некорректные значения в колонке 03 Код специальности.Вместо кода присутствует дата, вместе с кодом есть название,пробел перед кодом и т.п.!!!']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue

        code_spec = [spec for spec in df_form1['03'].unique()]

        # Создаем список для строк
        row_cat = [f'Строка {i}' for i in range(1, 3)]
        # Создаем список для колонок
        column_cat = [f'Колонка {i}' for i in range(6, 79)]

        # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
        spec_dict = {}
        for row in row_cat:
            spec_dict[row] = {key: 0 for key in column_cat}

        # Изменяем последний ключ на строковый поскольку там будут хранится примечания
        for row, value in spec_dict.items():
            for col, data in value.items():
                if col == 'Колонка 78':
                    spec_dict[row][col] = ''

        # Создаем словарь среднего уровня содержащй данные по всем специальностям
        poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}

        high_level_dct[name_file] = copy.deepcopy(poo_dct)

        """
        В итоге получается такая структура
        {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

        """

        current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

        idx_row = 1  # счетчик обработанных строк
        # Итерируемся по полученному датафрейму через itertuples
        for row in df_form1.itertuples():
            # если счетчик колонок больше 15 то уменьшаем его до единицы
            if idx_row > 2:
                idx_row = 1
                # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
            if (row[3] is not np.nan) and (row[3] != ' '):
                # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                if row[3] != current_code:
                    current_code = row[3]
            data_row = row[6:79]  # получаем срез с нужными данными
            for idx_col, value in enumerate(data_row, start=1):
                if idx_col + 5 == 78:
                    # сохраняем примечания в строке
                    high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                        f'Колонка {idx_col + 5}'] = f'{name_file} {check_data_note(value)};'

                else:
                    high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                        f'Колонка {idx_col + 5}'] += check_data(value)
            idx_row += 1
# создаем файл для проверки. Специальности с разбивкой по организациям
create_check_tables_opk(high_level_dct)

# получаем уникальные специальности
all_spec_code = set()
for poo, spec in high_level_dct.items():
    for code_spec, data in spec.items():
        all_spec_code.add(code_spec)

itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}

# Складываем результаты неочищенного словаря
for poo, spec in high_level_dct.items():
    for code_spec, data in spec.items():
        for row, col_data in data.items():
            for col, value in col_data.items():
                if col == 'Колонка 78':
                    itog_df[code_spec][row][col] += check_data_note(value)
                else:
                    itog_df[code_spec][row][col] += value

# Сортируем получившийся словарь по возрастанию для удобства использования
sort_itog_dct = sorted(itog_df.items())
itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

out_df = pd.DataFrame.from_dict(itog_df, orient='index')

stack_df = out_df.stack()
# название такое выбрал потому что было лень заменять значения из блокнота юпитера
frame = stack_df.to_frame()

frame['Суммарный выпуск 2023 г.'] = frame[0].apply(lambda x: x.get('Колонка 6'))
frame[
    'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
    frame[0].apply(lambda x: x.get('Колонка 7'))
frame['Трудоустроены на предприятия оборонно-промышленного комплекса*'] = \
    frame[0].apply(lambda x: x.get('Колонка 8'))
frame['Трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 9'))
frame['Трудоустроены на предприятия сельского хозяйства'] = \
    frame[0].apply(lambda x: x.get('Колонка 10'))
frame['Трудоустроены на предприятия металлургии'] = \
    frame[0].apply(lambda x: x.get('Колонка 11'))
frame['Трудоустроены на предприятия железнодорожного транспорта'] = \
    frame[0].apply(lambda x: x.get('Колонка 12'))
frame['Трудоустроены на предприятия легкой промышленности'] = \
    frame[0].apply(lambda x: x.get('Колонка 13'))
frame['Трудоустроены на предприятия химической отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 14'))
frame['Трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 15'))
frame['Трудоустроены на предприятия фармацевтической отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 16'))
frame['Трудоустроены на предприятия отрасли информационных технологий'] = \
    frame[0].apply(lambda x: x.get('Колонка 17'))
frame['Трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 18'))
frame['Трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 19'))
frame['Трудоустроены на предприятия транспортной отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 20'))
frame['Трудоустроены на предприятия горнодобывающей отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 21'))
frame[
    'Трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 22'))
frame['Трудоустроены на предприятия лесной промышленности'] = \
    frame[0].apply(lambda x: x.get('Колонка 23'))
frame['Трудоустроены на предприятия строительной отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 24'))
frame['Трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 25'))
frame['Трудоустроены на предприятия индустрии робототехники'] = \
    frame[0].apply(lambda x: x.get('Колонка 26'))
frame['Трудоустроены на предприятия в отрасли образования'] = \
    frame[0].apply(lambda x: x.get('Колонка 27'))
frame['Трудоустроены на предприятия в медицинской отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 28'))
frame[
    'Трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
    frame[0].apply(lambda x: x.get('Колонка 29'))
frame['Трудоустроены на предприятия в отрасли искусства'] = \
    frame[0].apply(lambda x: x.get('Колонка 30'))
frame['Трудоустроены на предприятия иная отрасль'] = \
    frame[0].apply(lambda x: x.get('Колонка 31'))

frame['Индивидуальные предприниматели'] = \
    frame[0].apply(lambda x: x.get('Колонка 32'))
frame['Самозанятые (перешедшие на специальный налоговый режим  - налог на профессиональный доход)'] = \
    frame[0].apply(lambda x: x.get('Колонка 33'))
frame['Продолжили обучение'] = \
    frame[0].apply(lambda x: x.get('Колонка 34'))
frame['Проходят службу в армии по призыву'] = \
    frame[0].apply(lambda x: x.get('Колонка 35'))
frame[
    'Проходят службу в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
    frame[0].apply(lambda x: x.get('Колонка 36'))
frame['Находятся в отпуске по уходу за ребенком'] = \
    frame[0].apply(lambda x: x.get('Колонка 37'))
frame[
    'Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
    frame[0].apply(lambda x: x.get('Колонка 38'))
frame['Будут трудоустроены на предприятия оборонно-промышленного комплекса* '] = \
    frame[0].apply(lambda x: x.get('Колонка 39'))
frame['Будут трудоустроены на предприятия машиностроения (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 40'))
frame['Будут трудоустроены на предприятия сельского хозяйства'] = \
    frame[0].apply(lambda x: x.get('Колонка 41'))
frame['Будут трудоустроены на предприятия металлургии'] = \
    frame[0].apply(lambda x: x.get('Колонка 42'))
frame['Будут трудоустроены на предприятия железнодорожного транспорта'] = \
    frame[0].apply(lambda x: x.get('Колонка 43'))
frame['Будут трудоустроены на предприятия легкой промышленности'] = \
    frame[0].apply(lambda x: x.get('Колонка 44'))
frame['Будут трудоустроены на предприятия химической отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 45'))
frame['Будут трудоустроены на предприятия атомной отрасли (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 46'))
frame['Будут трудоустроены на предприятия фармацевтической отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 47'))
frame['Будут трудоустроены на предприятия отрасли информационных технологий'] = \
    frame[0].apply(lambda x: x.get('Колонка 48'))
frame['Будут трудоустроены на предприятия радиоэлектроники (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 49'))
frame[
    'Будут трудоустроены на предприятия топливно-энергетического комплекса (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 50'))

frame['Будут трудоустроены на предприятия транспортной отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 51'))
frame['Будут трудоустроены на предприятия горнодобывающей отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 52'))
frame[
    'Будут трудоустроены на предприятия отрасли электротехнической промышленности (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 53'))
frame['Будут трудоустроены на предприятия лесной промышленности'] = \
    frame[0].apply(lambda x: x.get('Колонка 54'))
frame['Будут трудоустроены на предприятия строительной отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 55'))
frame[
    'Будут трудоустроены на предприятия отрасли электронной промышленности (кроме оборонно-промышленного комплекса)'] = \
    frame[0].apply(lambda x: x.get('Колонка 56'))
frame['Будут трудоустроены на предприятия индустрии робототехники'] = \
    frame[0].apply(lambda x: x.get('Колонка 57'))
frame['Будут трудоустроены на предприятия в отрасли образования'] = \
    frame[0].apply(lambda x: x.get('Колонка 58'))
frame['Будут трудоустроены на предприятия в медицинской отрасли'] = \
    frame[0].apply(lambda x: x.get('Колонка 59'))
frame[
    'Будут трудоустроены на предприятия в отрасли сферы услуг, туризма, торговли, организациях финансового сектора, правоохранительной сферы и управления, средств массовой информации'] = \
    frame[0].apply(lambda x: x.get('Колонка 60'))
frame['Будут трудоустроены на предприятия в отрасли искусства'] = \
    frame[0].apply(lambda x: x.get('Колонка 61'))
frame['Будут трудоустроены на предприятия иная отрасль'] = \
    frame[0].apply(lambda x: x.get('Колонка 62'))
frame['Будут индивидуальными предпринимателями'] = \
    frame[0].apply(lambda x: x.get('Колонка 63'))
frame['Будут самозанятыми'] = \
    frame[0].apply(lambda x: x.get('Колонка 64'))
frame['Будут продолжать обучение'] = \
    frame[0].apply(lambda x: x.get('Колонка 65'))
frame['Будут призваны в армию'] = \
    frame[0].apply(lambda x: x.get('Колонка 66'))
frame[
    'будут в армии на контрактной основе, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии Российской Федерации, органах принудительного исполнения Российской Федерации**'] = \
    frame[0].apply(lambda x: x.get('Колонка 67'))
frame['Будут находиться в отпуске по уходу за ребенком'] = \
    frame[0].apply(lambda x: x.get('Колонка 68'))
frame['Неформальная занятость (теневой сектор экономики)'] = \
    frame[0].apply(lambda x: x.get('Колонка 69'))
frame[
    'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице) и не планируют трудоустраиваться'] = \
    frame[0].apply(lambda x: x.get('Колонка 70'))
frame[
    'Не имеют мотивации к трудоустройству (кроме зарегистрированных в качестве безработных) и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот '] = \
    frame[0].apply(lambda x: x.get('Колонка 71'))
frame['Иные причины нахождения под риском нетрудоустройства'] = \
    frame[0].apply(lambda x: x.get('Колонка 72'))
frame['Смерть, тяжелое состояние здоровья'] = \
    frame[0].apply(lambda x: x.get('Колонка 73'))
frame['Находятся под следствием, отбывают наказание'] = \
    frame[0].apply(lambda x: x.get('Колонка 74'))
frame[
    'Переезд за пределы Российской Федерации (кроме переезда в иные регионы - по ним регион должен располагать сведениями)'] = \
    frame[0].apply(lambda x: x.get('Колонка 75'))
frame[
    'Не могут трудоустраиваться в связи с уходом за больными родственниками, в связи с иными семейными обстоятельствами'] = \
    frame[0].apply(lambda x: x.get('Колонка 76'))
frame[
    'Иное в первую очередь выпускники распределяются по всем остальным графам. Данная графа предназначена для очень редких случаев. Если в нее включено более 0,5% выпускников - укажите причины в графе "принимаемые меры"'] = \
    frame[0].apply(lambda x: x.get('Колонка 77'))
frame[
    'Принимаемые меры по содействию занятости, в том числе по трудоустройству выпускников на предприятия оборонно-промышленного комплекса тезисно - вид меры, охват выпускников мерой'] = \
    frame[0].apply(lambda x: x.get('Колонка 78'))

finish_df = frame.drop([0], axis=1)

finish_df = finish_df.reset_index()

finish_df.rename(
    columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
    inplace=True)

dct = {'Строка 1': 'Всего (общая численность выпускников)',
       'Строка 2': 'из строки 01: имеют договор о целевом обучении'}

finish_df['Наименование показателей (категория выпускников)'] = finish_df[
    'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

# генерируем текущее время
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)

# создаем итоговый отчет по ОПК
opk_df = finish_df[['Суммарный выпуск 2023 г.','Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                    'Трудоустроены на предприятия оборонно-промышленного комплекса*','Будут трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)',
                    'Будут трудоустроены на предприятия оборонно-промышленного комплекса* ']]
opk_df = opk_df.sum(axis=0).to_frame() # суммируем данные
opk_df = opk_df.transpose() # разворачиваем из колонки в строку
opk_df.to_excel(f'{path_to_end_folder}/Итог по ОПК {current_time}.xlsx', index=False)

#Создаем сумму по всем колонкам
all_sum_df = finish_df.iloc[:,2:].sum(axis=0).to_frame() # суммируем данные
all_sum_df = all_sum_df.transpose() # разворачиваем из колонки в строку
all_sum_df.to_excel(f'{path_to_end_folder}/Сумма по всем категориям {current_time}.xlsx', index=False)

# добавляем строки с проверкой
count = 0
for i in range(2, len(finish_df) + 1, 2):
    new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
    new_row.iloc[:, 1] = 'Проверка (строка не редактируется) Строку требуется подставлять после каждого кода'
    new_row.iloc[:, 2:] = 'проверка пройдена'

    # Вставка новой строки через каждые 15 строк
    finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(drop=True)
    count += 1

lst_number_row = ['01', '02', '03']
multipler = len(finish_df) // 3  # получаем количество специальностей/профессий
# вставляем новую колонку
finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))

finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
finish_df.to_excel(f'{path_to_end_folder}/Полная таблица Трудоустройство ОПК от {current_time}.xlsx', index=False)
error_df.to_excel(f'{path_to_end_folder}/Ошибки ОПК от {current_time}.xlsx',index=False)

print(error_df)

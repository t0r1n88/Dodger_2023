# -*- coding: utf-8 -*-
"""
Скрипт для обработки данных Формы 2 нозология (15 строк) мониторинга занятости выпускников
"""
from cass_check_functions import * # импортируем функции проверки
from cass_support_functions import * # импортируем вспомогательные функции и исключения
import pandas as pd
import numpy as np
import os
import warnings
from tkinter import messagebox
import time
pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def prepare_form_two_employment(path_folder_data:str,path_to_end_folder):
    """
    Фугкция для обработки данных формы 2 15 строк нозология
    :return:
    """
    # создаем словарь верхнего уровня для каждого поо
    high_level_dct = {}
    # создаем словарь верхнего уровня для хранения пары ключ значение где ключ это код специальности а значение- код и наименование
    dct_code_and_name = dict()
    # создаем датафрейм для регистрации ошибок
    error_df = pd.DataFrame(columns=['Название файла', 'Строка или колонка с ошибкой', 'Описание ошибки', ])

    tup_correct = (6, 19)  # создаем кортеж  с поправками где 6 это первая строка с данными а 19 строка где заканчивается первый диапазон
    try:
        for file in os.listdir(path_folder_data):
            if not file.startswith('~$') and not file.endswith('.xlsx'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                    'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)
                # получаем название первого листа
                temp_wb = openpyxl.load_workbook(f'{path_folder_data}/{file}', read_only=True)
                lst_temp_sheets = temp_wb.sheetnames  # получаем листы в файле
                temp_wb.close()
                if 'Форма нозологии' not in lst_temp_sheets:  # проверяем наличие листа с названием в файле
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Не найден лист с названием Форма нозологии !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
                df = pd.read_excel(f'{path_folder_data}/{file}',sheet_name='Форма нозологии', skiprows=4, dtype=str)
                df.columns = list(map(str, df.columns))  # делаем названия колонок строковыми
                # создаем множество колонок наличие которых мы проверяем
                check_cols = ['гр.01', 'гр.02', 'гр.03', 'гр.04', 'гр.05', 'гр.06', 'гр.07', 'гр.08', 'гр.09', 'гр.10', 'гр.11', 'гр.12', 'гр.13', 'гр.14', 'гр.15',
                              'гр.16', 'гр.17',
                              'гр.18', 'гр.19', 'гр.20', 'гр.21', 'гр.22', 'гр.23', 'гр.24', 'гр.25', 'гр.26', 'гр.27', 'гр.28', 'гр.29', 'гр.30', 'гр.31', 'гр.32']
                if check_cols != list(df.columns):
                    diff_cols = set(list(df.columns)).difference(set(check_cols))
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{diff_cols}',
                                                        'Возможно старая версия формы сбора данных.Строка с номерами колонок (гр.01,гр.02,гр.03,гр.05 ... гр.31,гр.32 как в исходной форме)\n должна находиться на 5 строке!\n'
                                                        ' указанные колонки являются лишними.Колонки с названимем Unnamed означаеют что на листе есть данные без заголовка в виде цифр на 5 строке  ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue

                df = df[df['гр.03'].notna()] # убираем возможные наны из за лишних строк
                df = df[~df['гр.03'].str.contains('Проверка')]  # фильтруем строки с проверками

                # отсекаем возможный первый столбец с данными ПОО,начинаем датафрейм с колонки 01 и отсекаем колонки с проверками
                df = df.loc[:, 'гр.02':'гр.29']


                # получаем  часть с данными
                mask = pd.isna(df).all(axis=1)  # создаем маску для строк с пропущенными значениями
                # проверяем есть ли строка полностью состоящая из nan
                empty_row_index = np.where(df.isna().all(axis=1))
                if empty_row_index[0].tolist():
                    row_index = empty_row_index[0][0]
                    df = df.iloc[:row_index]
                # Проверка на лишние или недостающие строки в конце таблицы
                remains = len(df) % 14
                if remains !=0 :
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        f'На каждую специальность/профессию должно приходиться по 14 строк не считая строки проверки.Найдено {len(df)} строк с данными, остаток при делении на 14 равен {remains}. Возможно какие то строки удалены или под таблицей есть лишние строки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue


                #     # Проверка на размер таблицы, должно бьть кратно 14
                count_spec = df.shape[0] // 14  # количество специальностей
                df = df.iloc[:count_spec * 14, :]  # отбрасываем строки проверки
                #
                check_code_lst = df['гр.02'].tolist()  # получаем список кодов специальностей
                # Проверка на то чтобы в колонке 02 в первой строке не было пустой ячейки
                if True in mask.tolist():
                    if check_code_lst[0] is np.nan or check_code_lst[0] == ' ':
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                            'В колонке гр.02 на первой строке не заполнен код специальности. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла', 'Строка или колонка с ошибкой',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue
                # Проверка на непрерывность кода специальности, то есть на 14 строк должен быть только один код и на пустые ячейки
                border_check_code = 0  # начало отсчета
                quantity_check_code = len(check_code_lst) // 14  # получаем сколько специальностей в таблице
                correction = 1  # размер поправки на случай если есть строка проверки
                sameness_error_df = check_sameness_column(check_code_lst, 14, border_check_code, quantity_check_code,
                                                          tup_correct, correction, name_file, 'Код и наименование')

                blankness_error_df = check_blankness_column(check_code_lst, 14, border_check_code, quantity_check_code,
                                                            tup_correct, correction, name_file, 'Код и наименование')


                # Добавляем колонку с номерами строк
                lst_number_row = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14']
                df.insert(1,'03',lst_number_row * count_spec)
                # проверяем на арифметические ошибки
                file_error_df = check_error_form_two(df.copy(), name_file, tup_correct)
                # добавляем в получившийся датафейм ошибки однородности диапазона
                file_error_df = pd.concat([file_error_df, sameness_error_df], axis=0, ignore_index=True)
                file_error_df = pd.concat([file_error_df, blankness_error_df], axis=0, ignore_index=True)
                # добавляем в словарь в полные имена из кода и наименования
                for full_name in df['гр.02'].tolist():
                    code = extract_code_nose(full_name)  # получаем только цифры
                    dct_code_and_name[code] = full_name
                # очищаем от текста чтобы названия листов не обрезались
                df['гр.02'] = df['гр.02'].apply(extract_code_nose)  # очищаем от текста в кодах
                if 'error' in df['гр.02'].values:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'Некорректные значения в колонке гр.02 Код и наименование профессии/специальности.Вместо кода присутствует дата, и т.п. проверьте правильность заполнения колонки 02!!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    file_error_df = pd.concat([file_error_df, temp_error_df], axis=0, ignore_index=True)
                # добавляем в основной файл с ошибками
                error_df = pd.concat([error_df, file_error_df], axis=0, ignore_index=True)
                if file_error_df.shape[0] != 0:
                    temp_error_df = pd.DataFrame(data=[[f'{name_file}', '',
                                                        'В файле обнаружены ошибки!!! ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                                 columns=['Название файла', 'Строка или колонка с ошибкой',
                                                          'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue
        #
                # Создание словаря для хранения данных файла
                code_spec = [spec for spec in df['гр.02'].unique()]  # получаем список специальностей которые есть в файле
                # Создаем список для строк
                row_cat = [f'Строка {i}' for i in range(1, 15)]
                # Создаем список для колонок
                column_cat = [f'Колонка {i}' for i in range(4, 30)]  # 4 это порядковый номер колонки суммарный выпуск а 30 это последняя колонка с цифрами
                # Создаем словарь нижнего уровня содержащий в себе все данные для каждой специальности
                spec_dict = {}
                for row in row_cat:
                    spec_dict[row] = {key: 0 for key in column_cat}
                poo_dct = {key: copy.deepcopy(spec_dict) for key in code_spec}
                high_level_dct[name_file] = copy.deepcopy(poo_dct)
                #             """
                #             В итоге получается такая структура
                #             {БРИТ:{13.01.10:{Строка 1:{Колонка 1:0}}},ТСИГХ:{22.01.10:{Строка 1:{Колонка 1:0}}}}

                current_code = 'Ошибка проверьте правильность заполнения кодов специальностей'  # чекбокс для проверки заполнения кода специальности

                idx_row = 1  # счетчик обработанных строк

                # Итерируемся по полученному датафрейму через itertuples
                for row in df.itertuples():
                    # если счетчик колонок больше 14 то уменьшаем его до единицы
                    if idx_row > 14:
                        idx_row = 1
                    # Проверяем на незаполненные ячейки и ячейки заполненные пробелами
                    if (row[1] is not np.nan) and (row[1] != ' '):
                        # если значение ячейки отличается от текущего кода специальности то обновляем значение текущего кода
                        if row[1] != current_code:
                            current_code = row[1]
                    data_row = row[4:30]  # получаем срез с нужными данными колонки в которых есть числа

                    for idx_col, value in enumerate(data_row, start=1):
                        high_level_dct[name_file][current_code][f'Строка {idx_row}'][
                            f'Колонка {idx_col + 3}'] += check_data(value)
                    #
                    idx_row += 1
        #
        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)
        wb_check_tables = create_check_tables_form_two(high_level_dct)  # проверяем данные по каждой специальности
        if 'Sheet' in wb_check_tables.sheetnames:
            del wb_check_tables['Sheet']
        wb_check_tables.save(
            f'{path_to_end_folder}/Данные для проверки правильности заполнения файлов от {current_time}.xlsx')
        #
        # получаем уникальные специальности
        all_spec_code = set()
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                all_spec_code.add(code_spec)

        itog_df = {key: copy.deepcopy(spec_dict) for key in all_spec_code}
        # Складываем результаты неочищенного словаря
        for poo, spec in high_level_dct.items():
            for code_spec, data in spec.items():
                for row, col_data in data.items():
                    for col, value in col_data.items():
                        itog_df[code_spec][row][col] += value

        # Сортируем получившийся словарь по возрастанию для удобства использования
        sort_itog_dct = sorted(itog_df.items())
        itog_df = {dct[0]: dct[1] for dct in sort_itog_dct}

        out_df = pd.DataFrame.from_dict(itog_df, orient='index')
        #
        stack_df = out_df.stack()
        frame = stack_df.to_frame()
        frame['Суммарный выпуск'] = frame[0].apply(lambda x: x.get('Колонка 4'))
        frame[
            'Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании)'] = \
            frame[0].apply(lambda x: x.get('Колонка 5'))
        frame['из них (из графы 05): продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 6'))
        frame[
            'из них (из графы 05): трудоустроены по полученной профессии, специальности'] = \
            frame[0].apply(lambda x: x.get('Колонка 7'))
        frame['Индиви-дуальные предприни-матели '] = frame[0].apply(lambda x: x.get('Колонка 8'))
        frame['Самозанятые (перешедшие на специальный налоговый режим - налог на профессио-нальный доход)'] = frame[0].apply(lambda x: x.get('Колонка 9'))
        frame[
            'Проходят службу в армии по призыву'] = \
            frame[0].apply(lambda x: x.get('Колонка 10'))
        frame['Проходят службу в армии по контракту, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии РФ, органах принудительного исполнения РФ '] = frame[0].apply(
            lambda x: x.get('Колонка 11'))
        frame['Продолжили обучение'] = frame[0].apply(lambda x: x.get('Колонка 12'))
        frame[
            'Находятся в отпуске по уходу за ребенком'] = \
            frame[0].apply(lambda x: x.get('Колонка 13'))
        frame[
            'Неформальная занятость (теневой сектор экономики)'] = \
            frame[0].apply(lambda x: x.get('Колонка 14'))
        frame[
            'Зарегистрированы в центрах занятости в качестве безработных (получают пособие по безработице)'] = \
            frame[0].apply(
                lambda x: x.get('Колонка 15'))
        frame['Не имеют мотивации к трудоустройству и не планируют трудоустраиваться, в том числе по причинам получения иных социальных льгот'] = frame[0].apply(lambda x: x.get('Колонка 16'))
        frame['Отсутствует спрос на специалистов в регионе, находятся в поиске работы'] = frame[0].apply(
            lambda x: x.get('Колонка 17'))
        frame[
            'Смерть, тяжелое состояние здоровья'] = \
            frame[0].apply(lambda x: x.get('Колонка 18'))
        frame[
            'Находятся под следствием, отбывают наказание '] = \
            frame[0].apply(lambda x: x.get('Колонка 19'))
        frame['Переезд за пределы Российской Федерации (кроме переезда в иные регионы)'] = frame[0].apply(
            lambda x: x.get('Колонка 20'))
        frame['Ухаживают за больными родственниками (иные семейные обстоятельства)'] = frame[0].apply(lambda x: x.get('Колонка 21'))
        frame['будут трудоустроены (в соответствии с трудовым законодательством, законодательством об обязательном пенсионном страховании)'] = frame[0].apply(
            lambda x: x.get('Колонка 22'))
        frame['из них (из графы 22): продолжат обучение'] = frame[0].apply(lambda x: x.get('Колонка 23'))
        frame['из них (из графы 22): будут трудоустроены по полученной профессии, специальности'] = frame[0].apply(lambda x: x.get('Колонка 24'))
        frame[
            'будут осуществлять предприни-мательскую деятельность'] = \
            frame[0].apply(lambda x: x.get('Колонка 25'))
        frame['будут самозанятыми'] = frame[0].apply(lambda x: x.get('Колонка 26'))
        frame['будут призваны в армию'] = frame[0].apply(lambda x: x.get('Колонка 27'))
        frame['будут в армии по контракту, в органах внутренних дел, Государственной противопожарной службе, органах по контролю за оборотом наркотических средств и психотропных веществ, учреждениях и органах уголовно-исполнительной системы, войсках национальной гвардии РФ, органах принудительного исполнения РФ'] = frame[0].apply(lambda x: x.get('Колонка 28'))
        frame['будут продолжать обучение'] = frame[0].apply(lambda x: x.get('Колонка 29'))
        #
        finish_df = frame.drop([0], axis=1)

        finish_df = finish_df.reset_index()

        finish_df.rename(
            columns={'level_0': 'Код специальности', 'level_1': 'Наименование показателей (категория выпускников)'},
            inplace=True)
        # создаем словарь для замены слов Строка на правильные обозначения
        dct = {'Строка 1': 'Всего (общая численность выпускников)',
               'Строка 2': 'из общей численности выпускников (из строки 01): лица с ОВЗ',
               'Строка 3': 'из числа лиц с ОВЗ (из строки 02): инвалиды и дети-инвалиды',
               'Строка 4': 'Инвалиды и дети-инвалиды (кроме учтенных в строке 03)',
               'Строка 5': 'Имеют договор о целевом обучении',
               'Строка 6': 'Автосумма строк 02 и 04 - Всего (общая численность выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов) '
            ,
               'Строка 7': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): с нарушениями: зрения',
               'Строка 8': 'слуха', 'Строка 9': 'опорно-двигательного аппарата',
               'Строка 10': 'тяжелыми нарушениями речи', 'Строка 11': 'задержкой психического развития',
               'Строка 12': 'расстройствами аутистического спектра',
               'Строка 13': 'с инвалидностью вследствие  других причин',
               'Строка 14': 'из общей численности выпускников из числа лиц с ОВЗ, инвалидов и детей-инвалидов (из строки 06): имеют договор о целевом обучении',
               }
        #
        finish_df['Наименование показателей (категория выпускников)'] = finish_df[
            'Наименование показателей (категория выпускников)'].apply(lambda x: dct[x])

        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        finish_df['Код специальности'] = finish_df['Код специальности'].apply(
            lambda x: dct_code_and_name[x])  # делаем код чтобы отображался код и наименование

        # добавляем строки с проверкой
        count = 0
        for i in range(14, len(finish_df) + 1, 14):
            new_row = finish_df.iloc[i - 1 + count, :].to_frame().transpose().copy()
            new_row.iloc[:, 1] = 'Проверка (строка не редактируется) - для специальностей'
            new_row.iloc[:, 2:] = 'проверка пройдена'
        #
            # Вставка новой строки через каждые 14 строк
            finish_df = pd.concat([finish_df.iloc[:i + count], new_row, finish_df.iloc[i + count:]]).reset_index(
                drop=True)
            count += 1
        lst_number_row = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15'
                          ]
        multipler = len(finish_df) // 15  # получаем количество специальностей/профессий
        # вставляем новую колонку
        finish_df.insert(1, 'Номер строки', pd.Series(lst_number_row * multipler))

        finish_df = finish_df[finish_df['Код специальности'] != 'nan']  # отбрасываем nan
        # Соединяем колонки с номером строки и названием строки
        finish_df['Наименование показателей (категория выпускников)'] = finish_df['Номер строки'] + '. ' + finish_df['Наименование показателей (категория выпускников)']
        finish_df.drop(columns=['Номер строки'],inplace=True)

        #
        # Создаем датафреймы с 1 и 5 строками
        small_finish_df = pd.DataFrame(columns=finish_df.columns)
        one_finish_df = pd.DataFrame(columns=finish_df.columns)
        lst_code_spec = finish_df['Код специальности'].unique()  # получаем список специальностей

        for code_spec in lst_code_spec:
            temp_df = finish_df[finish_df['Код специальности'] == code_spec]
            small_finish_df = pd.concat([small_finish_df, temp_df.iloc[:5, :]], axis=0, ignore_index=True)
            one_finish_df = pd.concat([one_finish_df, temp_df.iloc[:1, :]], axis=0, ignore_index=True)

        with pd.ExcelWriter(f'{path_to_end_folder}/Форма 2 нозологии (15 строк) от {current_time}.xlsx') as writer:
            finish_df.to_excel(writer, sheet_name='Нозологии 15 строк', index=False)
            small_finish_df.to_excel(writer, sheet_name='5 строк', index=False)
            one_finish_df.to_excel(writer, sheet_name='1 строка (Всего выпускников)', index=False)

            # Создаем документ
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{path_to_end_folder}/ОШИБКИ Форма 2 нозологии (15 строк) от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except PermissionError as e:
        messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                             f'Закройте открытые файлы Excel {e.args}')
    # except:
    #     messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
    #                          f'При обработке файла {name_file} возникла ошибка !!!\n'
    #                          f'Проверьте файл на соответствие шаблону исходной формы 2 нозологии)')

    else:
        if error_df.shape[0] != 0:
            messagebox.showerror('Кассандра Подсчет данных по трудоустройству выпускников',
                                 'Обнаружены ошибки в обрабатываемых файлах.\n'
                                 'Названия файлов с ошибками и ошибки вы можете найти в файле Ошибки.\n'
                                 'Исправьте ошибки и запустите повторную обработку для того чтобы получить полный результат.')
        else:
            messagebox.showinfo('Кассандра Подсчет данных по трудоустройству выпускников',
                                'Данные успешно обработаны.Ошибок не обнаружено')


if __name__ == '__main__':
    main_data_folder = 'data/example/15'
    main_data_folder = 'data/Нозология 2025'
    main_result_folder = 'data/РЕЗУЛЬТАТ'
    prepare_form_two_employment(main_data_folder,main_result_folder)

    print('Lindy Booth')